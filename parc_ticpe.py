# -*- coding: utf-8 -*-
"""Relevé Excel des heures de fonctionnement des engins de chantier (carrière).

Rapport « type TICPE » pour les aides spécifiques aux carrières. Par défaut,
un classeur multi-années : une feuille « Récap » (engins × années, toutes
années confondues) puis une feuille détaillée par année civile, de l'année de
la première saisie d'odomètre jusqu'à aujourd'hui. Avec ?du=&au=, une seule
feuille sur la période demandée. Heures = dernier relevé de la période −
dernier relevé avant la période. Servi par app.py :
GET /parc-ticpe.xlsx?k=<maquignon.parc_pdf_key>[&du=YYYY-MM-DD&au=YYYY-MM-DD].
"""
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# fleet.vehicle.model.category des engins de chantier
CATEGORIES_ENGINS = {169: "Chargeuse", 170: "Pelle", 171: "Tombereau", 172: "Chariot élévateur"}
FUEL = {"diesel": "Gazole/GNR", "gasoline": "Essence", "electric": "Électrique",
        "hybrid": "Hybride", "cng": "GNC", "lpg": "GPL", "hydrogen": "Hydrogène"}

GRAS = Font(bold=True)
TITRE = Font(bold=True, size=13)
SOUS = Font(size=9, color="64748B")
HEAD_FILL = PatternFill("solid", fgColor="0F172A")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
CAT_FILL = PatternFill("solid", fgColor="E2E8F0")
WARN = Font(color="B45309", size=9)
FINE = Side(style="thin", color="CBD5E1")
BORD = Border(left=FINE, right=FINE, top=FINE, bottom=FINE)
CENTRE = Alignment(horizontal="center")


def _d(s):
    return datetime.date.fromisoformat(str(s)[:10]) if s else None


def _nom(v):
    return v["name"].split("/")[-2] if "/" in v["name"] else v["name"]


def _calcul(rel, ddu, dau):
    """(relevé début, relevé fin, heures, observation) sur [ddu, dau]."""
    avant = [o for o in rel if _d(o["date"]) < ddu]
    dans = [o for o in rel if ddu <= _d(o["date"]) <= dau]
    deb = avant[-1] if avant else (dans[0] if dans else None)
    fin = dans[-1] if dans else None
    if deb and fin and fin is not deb:
        heures = round(fin["value"] - deb["value"], 1)
        if heures < 0:
            return deb, fin, None, "relevés incohérents (compteur en baisse)"
        obs = "" if avant else "pas de relevé avant la période : départ = 1er relevé de la période"
        return deb, fin, heures, obs
    if fin and deb is fin:
        return deb, fin, None, "un seul relevé sur la période — heures non calculables"
    if not rel:
        return None, None, None, "aucun relevé de compteur"
    return deb, None, None, "aucun relevé sur la période"


def _entete_feuille(ws, sous_titre):
    ws["A1"] = "Relevé des heures de fonctionnement — engins de chantier (carrière)"
    ws["A1"].font = TITRE
    ws["A2"] = sous_titre
    ws["A2"].font = SOUS


def _feuille_detail(ws, vehs, par_veh, ddu, dau):
    """Tableau détaillé (une période) : relevés début/fin, heures, observations."""
    _entete_feuille(ws, "Période du %s au %s · établi le %s · heures = dernier relevé de la "
                        "période − dernier relevé avant la période"
                        % (ddu.strftime("%d/%m/%Y"), dau.strftime("%d/%m/%Y"),
                           datetime.date.today().strftime("%d/%m/%Y")))
    entetes = ["Catégorie", "Engin", "N° de châssis / série", "Code parc", "Année",
               "Carburant", "Relevé initial", "Date", "Relevé final", "Date",
               "Heures de fonctionnement", "Observation"]
    ws.append([])
    ws.append(entetes)
    lh = ws.max_row
    for c in range(1, len(entetes) + 1):
        cell = ws.cell(row=lh, column=c)
        cell.fill, cell.font, cell.border, cell.alignment = HEAD_FILL, HEAD_FONT, BORD, CENTRE

    total, cat_tot, cat_cour = 0.0, 0.0, None

    def ligne_total_cat(nom):
        ws.append(["", "TOTAL %s" % nom, "", "", "", "", "", "", "", "", round(cat_tot, 1), ""])
        r = ws.max_row
        for c in range(1, len(entetes) + 1):
            ws.cell(row=r, column=c).font = GRAS
            ws.cell(row=r, column=c).fill = CAT_FILL
            ws.cell(row=r, column=c).border = BORD

    for v in vehs:
        cat = v["category_id"][1]
        if cat != cat_cour:
            if cat_cour is not None:
                ligne_total_cat(cat_cour)
            cat_cour, cat_tot = cat, 0.0
        deb, fin, heures, obs = _calcul(par_veh.get(v["id"], []), ddu, dau)
        ws.append([cat, _nom(v), v["vin_sn"] or "", v["license_plate"] or "",
                   v["model_year"] or "", FUEL.get(v["fuel_type"], v["fuel_type"] or ""),
                   deb["value"] if deb else "", _d(deb["date"]).strftime("%d/%m/%Y") if deb else "",
                   fin["value"] if fin else "", _d(fin["date"]).strftime("%d/%m/%Y") if fin else "",
                   heures if heures is not None else "", obs])
        r = ws.max_row
        for c in range(1, len(entetes) + 1):
            ws.cell(row=r, column=c).border = BORD
            if c in (7, 8, 9, 10, 11):
                ws.cell(row=r, column=c).alignment = CENTRE
        if obs:
            ws.cell(row=r, column=12).font = WARN
        if heures:
            cat_tot += heures
            total += heures
    if cat_cour is not None:
        ligne_total_cat(cat_cour)
    ws.append(["", "TOTAL GÉNÉRAL", "", "", "", "", "", "", "", "", round(total, 1), ""])
    r = ws.max_row
    for c in range(1, len(entetes) + 1):
        ws.cell(row=r, column=c).font = Font(bold=True, size=11)
        ws.cell(row=r, column=c).border = BORD
    for col, w in zip("ABCDEFGHIJKL", (16, 30, 22, 12, 8, 12, 12, 12, 12, 12, 20, 42)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A%d" % (lh + 1)


def _feuille_recap(ws, vehs, par_veh, annees, today):
    """Récap toutes années : engins en lignes, années en colonnes, totaux."""
    _entete_feuille(ws, "Récapitulatif %d-%d (toutes années) · établi le %s · détail par année "
                        "dans les feuilles suivantes"
                        % (annees[0], annees[-1], today.strftime("%d/%m/%Y")))
    entetes = ["Catégorie", "Engin", "N° de châssis / série"] + [str(a) for a in annees] + ["TOTAL"]
    ws.append([])
    ws.append(entetes)
    lh = ws.max_row
    for c in range(1, len(entetes) + 1):
        cell = ws.cell(row=lh, column=c)
        cell.fill, cell.font, cell.border, cell.alignment = HEAD_FILL, HEAD_FONT, BORD, CENTRE

    tot_an = {a: 0.0 for a in annees}
    cat_an, cat_cour = None, None
    nca = len(annees)

    def bornes(a):
        return datetime.date(a, 1, 1), min(datetime.date(a, 12, 31), today)

    def ligne_total_cat(nom):
        vals = [round(cat_an[a], 1) if cat_an[a] else "" for a in annees]
        ws.append(["", "TOTAL %s" % nom, ""] + vals +
                  [round(sum(cat_an.values()), 1) if any(cat_an.values()) else ""])
        r = ws.max_row
        for c in range(1, len(entetes) + 1):
            ws.cell(row=r, column=c).font = GRAS
            ws.cell(row=r, column=c).fill = CAT_FILL
            ws.cell(row=r, column=c).border = BORD

    for v in vehs:
        cat = v["category_id"][1]
        if cat != cat_cour:
            if cat_cour is not None:
                ligne_total_cat(cat_cour)
            cat_cour = cat
            cat_an = {a: 0.0 for a in annees}
        rel = par_veh.get(v["id"], [])
        vals, vt = [], 0.0
        for a in annees:
            ddu, dau = bornes(a)
            _, _, heures, _ = _calcul(rel, ddu, dau)
            vals.append(heures if heures is not None else "")
            if heures:
                cat_an[a] += heures
                tot_an[a] += heures
                vt += heures
        ws.append([cat, _nom(v), v["vin_sn"] or ""] + vals + [round(vt, 1) if vt else ""])
        r = ws.max_row
        for c in range(1, len(entetes) + 1):
            ws.cell(row=r, column=c).border = BORD
            if c > 3:
                ws.cell(row=r, column=c).alignment = CENTRE
    if cat_cour is not None:
        ligne_total_cat(cat_cour)
    ws.append(["", "TOTAL GÉNÉRAL", ""] + [round(tot_an[a], 1) for a in annees] +
              [round(sum(tot_an.values()), 1)])
    r = ws.max_row
    for c in range(1, len(entetes) + 1):
        ws.cell(row=r, column=c).font = Font(bold=True, size=11)
        ws.cell(row=r, column=c).border = BORD
    for i, w in enumerate((16, 30, 22) + (11,) * (nca + 1)):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A%d" % (lh + 1)


def generer(call_kw, du=None, au=None):
    """Classeur multi-années (défaut) ou période unique si du/au fournis."""
    today = datetime.date.today()
    vehs = call_kw("fleet.vehicle", "search_read",
                   [[["active", "=", True], ["category_id", "in", list(CATEGORIES_ENGINS)]]],
                   {"fields": ["name", "license_plate", "vin_sn", "model_year",
                               "category_id", "fuel_type"], "order": "category_id, name",
                    "limit": 500})
    odos = call_kw("fleet.vehicle.odometer", "search_read",
                   [[["vehicle_id", "in", [v["id"] for v in vehs]]]],
                   {"fields": ["vehicle_id", "date", "value"],
                    "order": "date asc, id asc", "limit": 10000})
    par_veh = {}
    for o in odos:
        if o["date"]:
            par_veh.setdefault(o["vehicle_id"][0], []).append(o)

    wb = Workbook()
    if du and au:
        _feuille_detail(wb.active, vehs, par_veh, _d(du), _d(au))
    else:
        premiere = min((_d(o["date"]) for o in odos if o["date"]), default=today)
        annees = list(range(premiere.year, today.year + 1))
        wb.active.title = "Récap %d-%d" % (annees[0], annees[-1])
        _feuille_recap(wb.active, vehs, par_veh, annees, today)
        for a in annees:
            ws = wb.create_sheet(str(a))
            _feuille_detail(ws, vehs, par_veh, datetime.date(a, 1, 1),
                            min(datetime.date(a, 12, 31), today))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
