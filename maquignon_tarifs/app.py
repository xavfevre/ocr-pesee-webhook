# -*- coding: utf-8 -*-
"""Tarifs clients MAQUIGNON — même principe que le module Protec
(`protec_planning`, page /protec/tarifs), adapté au groupe multi-sociétés.

Page bureau ouverte depuis Odoo (action « 💶 Tarifs client » sur la fiche
contact et sur les devis) : pour un client et une société du groupe,
catalogue complet avec prix standard (liste de référence de la société) vs
prix client (sa liste de prix), ajout / modification / suppression de tarifs
spécifiques à prix fixe avec plage de validité, historique facturé par
produit, onglet évolution par année, export Excel.

Différences avec Protec :
- multi-sociétés : sélecteur en tête de page, listes de référence par société
  (`ir.config_parameter maquignon.tarifs_defaut_pl`, JSON {société: liste}) ;
- les clients Maquignon sont souvent sur des listes PARTAGÉES (« Tarif Pro
  2026 »…) : on n'y écrit JAMAIS un tarif client. À la première saisie, une
  liste dédiée au client est créée (nom du client) avec un item de repli
  « formule = 100 % de l'ancienne liste » pour que tous les autres prix
  restent inchangés, puis affectée au client (contexte de la société) ;
- le jeton d'accès vit dans Odoo (`maquignon.tarifs_key`), pas en variable
  d'environnement.
"""
import os
import re
import xmlrpc.client
from datetime import date, datetime, timedelta, timezone

from flask import Blueprint, render_template, request, redirect, url_for, abort, Response

try:
    from zoneinfo import ZoneInfo
    TZ = ZoneInfo("Europe/Paris")
except Exception:  # pragma: no cover
    TZ = timezone(timedelta(hours=2))
UTC = timezone.utc

bp = Blueprint("mtarifs", __name__, template_folder="templates")

ODOO_URL = os.environ.get("ODOO_URL", "")
ODOO_DB = os.environ.get("ODOO_DB", "")
ODOO_USER = os.environ.get("ODOO_USER", "")
ODOO_PASSWORD = os.environ.get("ODOO_PASSWORD", "")

_conn = {}


def _q(model, method, *params, **kw):
    if "uid" not in _conn:
        common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
        _conn["uid"] = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
        _conn["models"] = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    return _conn["models"].execute_kw(ODOO_DB, _conn["uid"], ODOO_PASSWORD,
                                      model, method, list(params), kw)


def _param(key):
    v = _q("ir.config_parameter", "get_param", key)
    return v if isinstance(v, str) else ""


def _check_token():
    import hmac
    tok = request.args.get("token", "")
    ref = _param("maquignon.tarifs_key")
    if not (tok and ref and hmac.compare_digest(tok, ref)):
        abort(403)


def _defaut_pl_map():
    """{société: liste de prix de référence} — paramètre JSON, avec défauts."""
    import json
    base = {1: 299, 3: 284, 2: 293, 4: 281, 13: 1690}
    raw = _param("maquignon.tarifs_defaut_pl")
    if raw:
        try:
            base.update({int(k): int(v) for k, v in json.loads(raw).items()})
        except Exception:
            pass
    return base


def _paris_to_utc(day_str, end=False):
    d = date.fromisoformat(day_str)
    t = datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=TZ) if end else \
        datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=TZ)
    return t.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S")


def _utc_to_local(dt_str):
    if not dt_str:
        return None
    return datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=UTC).astimezone(TZ)


def _fmt_dt(dt_str):
    dt = _utc_to_local(dt_str)
    return dt.strftime("%d/%m/%Y") if dt else ""


def _ctx(comp_id):
    """Contexte multi-sociétés : la société choisie en tête (elle pilote les
    propriétés « company dependent » comme la liste de prix du client), mais
    TOUTES les sociétés autorisées — un contact peut appartenir à une autre
    société du groupe et deviendrait illisible sinon."""
    ids = [c["id"] for c in _companies()]
    return {"allowed_company_ids": [comp_id] + [i for i in ids if i != comp_id],
            "company_id": comp_id}


def _companies():
    return _q("res.company", "search_read", [], fields=["name"], order="id")


def _default_company(pid):
    """Société de la dernière facture du client, sinon SARL MAQUIGNON."""
    mv = _q("account.move", "search_read",
            [["partner_id", "child_of", pid], ["move_type", "=", "out_invoice"],
             ["state", "=", "posted"]],
            fields=["company_id"], order="invoice_date desc", limit=1)
    return mv[0]["company_id"][0] if mv else 1


def _partner_pl(pid, comp_id):
    """Liste de prix du client dans le contexte de la société."""
    p = _q("res.partner", "read", [pid],
           fields=["name", "property_product_pricelist"],
           context=_ctx(comp_id))
    return p[0] if p else None


def _pl_is_dediee(pl_id, pl_name, partner_name):
    """Une liste est « dédiée » (modifiable pour ce client) si son nom recouvre
    celui du client — c'est la convention de création de cette page — ou si
    elle est déclarée dans maquignon.tarifs_pl_dediees (ids séparés par des
    virgules)."""
    extra = {int(x) for x in re.findall(r"\d+", _param("maquignon.tarifs_pl_dediees") or "")}
    if pl_id in extra:
        return True
    a = (pl_name or "").split("(")[0].strip().lower()
    b = (partner_name or "").strip().lower()
    return bool(a and b and (a.startswith(b) or b.startswith(a)))


@bp.route("/", methods=["GET", "POST"])
def tarifs_client():
    _check_token()
    pid = request.args.get("c", type=int)
    if not pid:
        abort(404)
    comp_id = request.args.get("soc", type=int) or _default_company(pid)
    CTX = _ctx(comp_id)

    partner = _partner_pl(pid, comp_id)
    if not partner:
        abort(404)
    defaut_map = _defaut_pl_map()
    default_pl_id = defaut_map.get(comp_id, 0)
    prop = partner.get("property_product_pricelist")
    client_pl_id = prop[0] if prop else 0
    client_pl_name = prop[1] if prop else ""
    if client_pl_id == default_pl_id:
        client_pl_id = 0
    pl_dediee = client_pl_id and _pl_is_dediee(client_pl_id, client_pl_name, partner["name"])

    error, ok = None, request.args.get("ok")
    if request.method == "POST":
        f = request.form
        try:
            tmpl_id = int(f.get("tmpl") or 0)
            prix = float((f.get("prix") or "").replace(",", ".").replace(" ", ""))
            du, au = f.get("du", "").strip(), f.get("au", "").strip()
            if not tmpl_id or prix <= 0:
                raise ValueError("Produit et prix sont obligatoires.")
            if not pl_dediee:
                # liste partagée (ou aucune) : liste dédiée au client, avec un
                # repli « formule 100 % de l'ancienne liste » pour ne rien changer
                # d'autre que les tarifs saisis ici
                vals_pl = {"name": partner["name"], "company_id": comp_id}
                created = _q("product.pricelist", "create", [vals_pl], context=CTX)
                new_pl = created[0] if isinstance(created, list) else created
                base_pl = client_pl_id or default_pl_id
                if base_pl:
                    _q("product.pricelist.item", "create", [{
                        "pricelist_id": new_pl, "applied_on": "3_global",
                        "compute_price": "formula", "base": "pricelist",
                        "base_pricelist_id": base_pl, "price_discount": 0,
                    }], context=CTX)
                _q("res.partner", "write", [pid],
                   {"property_product_pricelist": new_pl}, context=CTX)
                client_pl_id = new_pl
            vals = {"pricelist_id": client_pl_id, "applied_on": "1_product",
                    "product_tmpl_id": tmpl_id, "compute_price": "fixed",
                    "fixed_price": prix}
            if du:
                vals["date_start"] = _paris_to_utc(du)
            if au:
                vals["date_end"] = _paris_to_utc(au, end=True)
            _q("product.pricelist.item", "create", [vals], context=CTX)
            return redirect(url_for(".tarifs_client", token=request.args.get("token"),
                                    c=pid, soc=comp_id, ok=1))
        except ValueError as ex:
            error = str(ex) or "Saisie invalide."
        except Exception:
            error = "Erreur d'enregistrement — réessayez."

    ds = _dataset(pid, comp_id, default_pl_id, client_pl_id)
    edit_map = {r["vid"]: r["items_edit"] for r in ds["rows"] if r["items_edit"]}
    comps = _companies()
    return render_template("tarifs_maquignon.html", partner=partner, pid=pid,
                           comp_id=comp_id, comps=comps,
                           edit_map=edit_map, pl_dediee=pl_dediee,
                           client_pl_name=client_pl_name if client_pl_id else "",
                           token=request.args.get("token"), ok=ok, error=error,
                           today=date.today().isoformat(), **ds)


def _dataset(pid, comp_id, default_pl_id, client_pl_id):
    """Catalogue, tarifs client, historique facturé et évolution par année."""
    products = _q("product.product", "search_read",
                  [["sale_ok", "=", True], ["active", "=", True],
                   "|", ["company_id", "=", False], ["company_id", "=", comp_id]],
                  fields=["display_name", "default_code", "list_price",
                          "product_tmpl_id", "categ_id"],
                  context=_ctx(comp_id))
    std_var, std_tmpl = {}, {}
    if default_pl_id:
        for it in _q("product.pricelist.item", "search_read",
                     [["pricelist_id", "=", default_pl_id], ["compute_price", "=", "fixed"]],
                     fields=["applied_on", "product_id", "product_tmpl_id", "fixed_price"]):
            if it["applied_on"] == "0_product_variant" and it["product_id"]:
                std_var[it["product_id"][0]] = it["fixed_price"]
            elif it["product_tmpl_id"]:
                std_tmpl[it["product_tmpl_id"][0]] = it["fixed_price"]

    cli_var, cli_tmpl = {}, {}
    if client_pl_id:
        for it in _q("product.pricelist.item", "search_read",
                     [["pricelist_id", "=", client_pl_id]],
                     fields=["applied_on", "product_id", "product_tmpl_id",
                             "compute_price", "fixed_price", "percent_price",
                             "date_start", "date_end", "min_quantity"],
                     order="date_start desc, id desc"):
            if it["applied_on"] == "0_product_variant" and it["product_id"]:
                cli_var.setdefault(it["product_id"][0], []).append(it)
            elif it["applied_on"] == "1_product" and it["product_tmpl_id"]:
                cli_tmpl.setdefault(it["product_tmpl_id"][0], []).append(it)

    hist_by_prod = {}
    for l in _q("account.move.line", "search_read",
                [["parent_state", "=", "posted"],
                 ["move_id.move_type", "in", ["out_invoice", "out_refund"]],
                 ["partner_id", "child_of", pid],
                 ["company_id", "=", comp_id],
                 ["product_id", "!=", False]],
                fields=["date", "move_name", "product_id", "quantity",
                        "price_unit", "discount"],
                order="date desc, id desc", limit=3000):
        hist_by_prod.setdefault(l["product_id"][0], []).append(l)

    now_utc = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for p in products:
        vid, tid = p["id"], p["product_tmpl_id"][0]
        std = std_var.get(vid, std_tmpl.get(tid, p["list_price"]))
        items = cli_var.get(vid, []) + cli_tmpl.get(tid, [])
        current = next((i for i in items
                        if (not i["date_start"] or i["date_start"] <= now_utc)
                        and (not i["date_end"] or i["date_end"] >= now_utc)), None)
        future = min((i for i in items if i["date_start"] and i["date_start"] > now_utc),
                     key=lambda i: i["date_start"], default=None)
        shown = current or future
        cli_price, validity = None, ""
        if shown:
            if shown["compute_price"] == "fixed":
                cli_price = shown["fixed_price"]
            elif shown["compute_price"] == "percentage":
                cli_price = round(std * (1 - shown["percent_price"] / 100), 2)
            ds_, de_ = shown["date_start"], shown["date_end"]
            if shown is future:
                validity = f"à partir du {_fmt_dt(ds_)}"
            elif ds_ and de_:
                validity = f"du {_fmt_dt(ds_)} au {_fmt_dt(de_)}"
            elif de_:
                validity = f"jusqu'au {_fmt_dt(de_)}"
            elif ds_:
                validity = f"depuis le {_fmt_dt(ds_)}"
        hist = hist_by_prod.get(vid, [])
        items_edit = []
        for it in items:
            if it["compute_price"] != "fixed":
                continue
            du_l = _utc_to_local(it["date_start"])
            au_l = _utc_to_local(it["date_end"])
            lbl = f"{it['fixed_price']:.2f} €".replace(".", ",")
            if du_l:
                lbl += f" — du {du_l.strftime('%d/%m/%Y')}"
            if au_l:
                lbl += f" au {au_l.strftime('%d/%m/%Y')}"
            items_edit.append({"id": it["id"], "prix": f"{it['fixed_price']:.2f}",
                               "du": du_l.date().isoformat() if du_l else "",
                               "au": au_l.date().isoformat() if au_l else "",
                               "label": lbl})
        rows.append({
            "vid": vid, "tmpl": tid, "name": p["display_name"],
            "code": p["default_code"] or "",
            "categ": p["categ_id"][1] if p.get("categ_id") else "Autre",
            "std": std,
            "items_edit": items_edit,
            "specific": bool(items), "cli_price": cli_price, "validity": validity,
            "nb_items": len(items), "items": items,
            "ecart": round((cli_price - std) / std * 100, 1) if (cli_price is not None and std) else None,
            "hist": [{"date": h["date"], "piece": h["move_name"],
                      "qte": h["quantity"], "pu": h["price_unit"],
                      "remise": h["discount"]} for h in hist[:30]],
            "hist_all": hist,
        })
    rows.sort(key=lambda r: (not r["specific"], not r["hist"], r["name"].lower()))
    nb_spec = sum(1 for r in rows if r["specific"])
    nb_fact = sum(1 for r in rows if r["hist"])
    categories = sorted({r["categ"] for r in rows}, key=str.lower)

    def _year_of(dt_str):
        return int(dt_str[:4]) if dt_str else None
    years_set = set()
    for r in rows:
        for it in r["items"]:
            for k in ("date_start", "date_end"):
                y = _year_of(it[k])
                if y:
                    years_set.add(y)
        for h in r["hist_all"]:
            y = _year_of(h["date"])
            if y:
                years_set.add(y)
    this_year = date.today().year
    years = sorted(y for y in years_set if y <= this_year)[-8:] or [this_year]
    evo_rows = []
    for r in rows:
        if not r["items"] and not r["hist_all"]:
            continue
        cells = []
        has_data = False
        for y in years:
            y_start, y_end = f"{y}-01-01 00:00:00", f"{y}-12-31 23:59:59"
            tarif = None
            applicable = [it for it in r["items"]
                          if it["compute_price"] == "fixed"
                          and (not it["date_start"] or it["date_start"] <= y_end)
                          and (not it["date_end"] or it["date_end"] >= y_start)]
            if applicable:
                applicable.sort(key=lambda it: it["date_start"] or "", reverse=True)
                tarif = applicable[0]["fixed_price"]
            pus = [h["price_unit"] for h in r["hist_all"] if _year_of(h["date"]) == y]
            fact = round(sum(pus) / len(pus), 2) if pus else None
            if tarif is not None or fact is not None:
                has_data = True
            cells.append({"tarif": tarif, "fact": fact})
        prev_val = None
        for cell in cells:
            val = cell["tarif"] if cell["tarif"] is not None else cell["fact"]
            cell["pct"] = None
            if val is not None and prev_val:
                pct = (val - prev_val) / prev_val * 100
                if abs(pct) >= 0.05:
                    cell["pct"] = round(pct, 1)
            if val is not None:
                prev_val = val
        if has_data:
            evo_rows.append({"name": r["name"], "code": r["code"],
                             "categ": r["categ"], "specific": r["specific"],
                             "cells": cells})

    return {"rows": rows, "evo_rows": evo_rows, "years": years,
            "categories": categories, "nb_spec": nb_spec, "nb_fact": nb_fact}


def _item_ok(item_id, pid, comp_id):
    """L'item appartient-il bien à la liste dédiée du client ? (garde-fou :
    on ne modifie jamais un item d'une liste partagée depuis cette page)"""
    it = _q("product.pricelist.item", "read", [item_id], fields=["pricelist_id"])
    if not it:
        return False
    pl_id, pl_name = it[0]["pricelist_id"]
    partner = _partner_pl(pid, comp_id)
    return bool(partner and _pl_is_dediee(pl_id, pl_name, partner["name"]))


@bp.route("/item/modifier", methods=["POST"])
def item_modifier():
    _check_token()
    pid = request.args.get("c", type=int)
    comp_id = request.args.get("soc", type=int) or 1
    f = request.form
    item_id = int(f.get("item_id") or 0)
    ok = "mod"
    try:
        prix = float((f.get("prix") or "").replace(",", ".").replace(" ", ""))
        if not item_id or prix <= 0 or not _item_ok(item_id, pid, comp_id):
            raise ValueError()
        du, au = f.get("du", "").strip(), f.get("au", "").strip()
        _q("product.pricelist.item", "write", [item_id], {
            "fixed_price": prix,
            "date_start": _paris_to_utc(du) if du else False,
            "date_end": _paris_to_utc(au, end=True) if au else False,
        })
    except (ValueError, TypeError):
        ok = None
    args = {"token": request.args.get("token"), "c": pid, "soc": comp_id}
    if ok:
        args["ok"] = ok
    return redirect(url_for(".tarifs_client", **args))


@bp.route("/item/supprimer", methods=["POST"])
def item_supprimer():
    _check_token()
    pid = request.args.get("c", type=int)
    comp_id = request.args.get("soc", type=int) or 1
    item_id = int(request.form.get("item_id") or 0)
    if item_id and _item_ok(item_id, pid, comp_id):
        _q("product.pricelist.item", "unlink", [item_id])
    return redirect(url_for(".tarifs_client", token=request.args.get("token"),
                            c=pid, soc=comp_id, ok="del"))


@bp.route("/export")
def tarifs_export():
    _check_token()
    pid = request.args.get("c", type=int)
    if not pid:
        abort(404)
    comp_id = request.args.get("soc", type=int) or _default_company(pid)
    partner = _partner_pl(pid, comp_id)
    if not partner:
        abort(404)
    defaut_map = _defaut_pl_map()
    default_pl_id = defaut_map.get(comp_id, 0)
    prop = partner.get("property_product_pricelist")
    client_pl_id = prop[0] if prop and prop[0] != default_pl_id else 0
    ds = _dataset(pid, comp_id, default_pl_id, client_pl_id)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    H_FILL = PatternFill("solid", fgColor="1F4E5F")
    H_FONT = Font(bold=True, color="FFFFFF")
    SPEC_FILL = PatternFill("solid", fgColor="FFF7E0")
    TARIF_FONT = Font(bold=True, color="1D7F79")
    FACT_FONT = Font(italic=True, color="6B8A89")

    ws = wb.active
    ws.title = "Tarifs actuels"
    ws.append(["Produit", "Code", "Catégorie", "Prix standard HT",
               "Prix client HT", "Écart %", "Validité", "Tarif spécifique"])
    for c in ws[1]:
        c.fill, c.font = H_FILL, H_FONT
    for r in ds["rows"]:
        ws.append([r["name"], r["code"], r["categ"], r["std"],
                   r["cli_price"] if r["cli_price"] is not None else "",
                   r["ecart"] if r["ecart"] is not None else "",
                   r["validity"], "Oui" if r["specific"] else ""])
        if r["specific"]:
            for c in ws[ws.max_row]:
                c.fill = SPEC_FILL
    for col, w in zip("ABCDEFGH", (52, 14, 18, 16, 14, 10, 26, 14)):
        ws.column_dimensions[col].width = w

    we = wb.create_sheet("Évolution")
    we.append(["Produit", "Code"] + [str(y) for y in ds["years"]])
    for c in we[1]:
        c.fill, c.font = H_FILL, H_FONT
    for r in ds["evo_rows"]:
        vals = [r["name"], r["code"]]
        for cell in r["cells"]:
            vals.append(cell["tarif"] if cell["tarif"] is not None else cell["fact"])
        we.append(vals)
        for j, cell in enumerate(r["cells"], 3):
            c = we.cell(row=we.max_row, column=j)
            if cell["tarif"] is not None:
                c.font = TARIF_FONT
            elif cell["fact"] is not None:
                c.font = FACT_FONT
    we.column_dimensions["A"].width = 52
    we.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    comp_name = next((c["name"] for c in _companies() if c["id"] == comp_id), "")
    fname = re.sub(r"[^A-Za-z0-9_-]+", "_", f"Tarifs_{partner['name']}_{comp_name}")
    return Response(buf.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={fname}.xlsx"})
