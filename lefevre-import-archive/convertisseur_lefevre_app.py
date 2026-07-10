#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║   MAQUIGNON — Convertisseur Commandes Lefevre → Odoo               ║
║   Importation directe via API Odoo ou export Excel                  ║
║   Version 1.0 — 2026                                                ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import json
import os
import sys
import re
import xmlrpc.client
from datetime import datetime
from pathlib import Path
import traceback

# ── Dépendances optionnelles (non bloquantes) ──────────────────────
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ══════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════

APP_NAME = "Lefevre → Odoo"
APP_VERSION = "1.0"
CONFIG_DIR = Path.home() / "AppData" / "Roaming" / "MaquignonLefevre"
CONFIG_FILE = CONFIG_DIR / "config.json"

COLORS = {
    "bg":         "#0F1923",
    "sidebar":    "#1A2535",
    "card":       "#1E2D40",
    "border":     "#2A3F58",
    "accent":     "#2E6DA4",
    "accent2":    "#5B9BD5",
    "success":    "#27AE60",
    "warning":    "#F39C12",
    "error":      "#E74C3C",
    "text":       "#ECF0F1",
    "text_dim":   "#8B9BB4",
    "entry_bg":   "#152030",
    "section_bg": "#243447",
    "white":      "#FFFFFF",
    "btn_hover":  "#3A7EC4",
}

DEFAULT_CONFIG = {
    "odoo": {
        "url": "https://votreinstance.odoo.com",
        "database": "",
        "username": "",
        "password": "",
        "use_apikey": False,
    },
    "commande": {
        "customer": "LEFEVRE",
        "customer_code": "LEF001",
        "date": "",
    },
    "fields": {
        "ref_pierre": "x_studio_ref_pierre",
        "nbr":        "x_studio_nbr",
        "long":       "x_studio_long",
        "larg":       "x_studio_larg",
        "haut":       "x_studio_haut",
        "poids":      "x_studio_poids",
    },
    "nature_mapping": {
        # product_code = code interne Odoo (entre [crochets] dans le nom produit)
        # T-P00X = code Sage (≠ default_code Odoo)
        "Usseau":    {"product_code": "TUF0000-PS",   "product": "[TUF0000-PS] Tuffeau - Pierres Pré-sciées 6 faces (Massif)"},
        "Tuffeau":   {"product_code": "TUF0000-PS",   "product": "[TUF0000-PS] Tuffeau - Pierres Pré-sciées 6 faces (Massif)"},
        "Haims":     {"product_code": "HAIMS0020-PS", "product": "[HAIMS0020-PS] Haims - Pré-sciées 6 faces (Massif)"},
        "Migné":     {"product_code": "MIGNE0000-PS", "product": "[MIGNE0000-PS] Migné - Pierres pré-sciées 6 faces (Massif)"},
        "Richemont": {"product_code": "RICH0000-PS",  "product": "[RICH0000-PS] Richemont - Pierres pré-sciées 6 faces (Massif)"},
        "Sireuil":   {"product_code": "SIRE0000-PS",  "product": "[SIRE0000-PS] Sireuil - Pierres pré-sciées 6 faces (Massif)"},
        "Tervoux":   {"product_code": "TERV0000-PS",  "product": "[TERV0000-PS] Tervoux - Pré-sciées 6 faces (Massif)"},
    }
}


# ══════════════════════════════════════════════════════════════════════
# GESTION CONFIGURATION
# ══════════════════════════════════════════════════════════════════════

class Config:
    def __init__(self):
        self.data = json.loads(json.dumps(DEFAULT_CONFIG))
        self.load()

    def load(self):
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                self._merge(self.data, saved)
                self._migrate_product_codes()
            except Exception:
                pass

    def _migrate_product_codes(self):
        """Corrige automatiquement les vieux codes Sage (T-P007) → codes Odoo (TUF0000-PS)"""
        sage_to_odoo = {
            "T-P007": "TUF0000-PS",  "T-P008": "HAIMS0020-PS",
            "T-P009": "MIGNE0000-PS","T-P010": "RICH0000-PS",
            "T-P011": "SIRE0000-PS", "T-P012": "TERV0000-PS",
        }
        nm = self.data.get("nature_mapping", {})
        changed = False
        for nature, info in nm.items():
            old_code = info.get("product_code", "")
            if old_code in sage_to_odoo:
                info["product_code"] = sage_to_odoo[old_code]
                changed = True
        if changed:
            self.save()

    def save(self):
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def _merge(self, base, override):
        for k, v in override.items():
            if k in base and isinstance(base[k], dict) and isinstance(v, dict):
                self._merge(base[k], v)
            else:
                base[k] = v

    def get(self, *keys, default=None):
        d = self.data
        for k in keys:
            if isinstance(d, dict) and k in d:
                d = d[k]
            else:
                return default
        return d

    def set(self, *keys_and_value):
        keys = keys_and_value[:-1]
        value = keys_and_value[-1]
        d = self.data
        for k in keys[:-1]:
            d = d.setdefault(k, {})
        d[keys[-1]] = value


# ══════════════════════════════════════════════════════════════════════
# PARSER LEFEVRE
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# PARSER LEFEVRE — Version robuste avec auto-détection de structure
# ══════════════════════════════════════════════════════════════════════

def safe_str(val):
    try:
        if val is None or (isinstance(val, float) and val != val):
            return ""
        return str(val).strip()
    except:
        return ""

def safe_num(val):
    try:
        if val is None or (isinstance(val, float) and val != val):
            return None
        f = float(str(val).replace(",", "."))
        return f if f == f else None  # NaN check
    except:
        return None

# ── Mots-clés pour l'auto-détection des colonnes ──────────────────
COL_KEYWORDS = {
    "ref":    ["désignation", "designation", "n°", "ref", "repere", "repère"],
    "nature": ["nature", "matiere", "matière", "pierre", "type"],
    "qte":    ["qte", "qté", "quantite", "quantité", "nb", "nbre", "nombre"],
    "long":   ["long", "longueur", "l."],
    "haut":   ["haut", "hauteur", "h."],
    "prof":   ["prof", "profondeur", "ep", "ép", "epais", "épaisseur", "larg"],
    "cube":   ["cube", "m3", "m³", "vol", "volume"],
    "poids":  ["poids", "kg", "masse"],
    "scie":   ["scié", "scie", "6 faces", "6f"],
    "obs":    ["obs", "observation", "remarque", "note"],
}

def detect_columns(df_raw):
    """
    Scanne le fichier pour trouver la ligne d'en-tête de colonnes
    et retourne un dict {nom_champ: index_colonne}.
    Retourne aussi l'index de la ligne d'en-tête.
    """
    best_score = 0
    best_row_idx = None
    best_col_map = {}

    for i, row in df_raw.iterrows():
        if i > 30:
            break
        row_vals = [safe_str(v).lower() for v in row]
        score = 0
        col_map = {}

        for field, keywords in COL_KEYWORDS.items():
            for j, cell in enumerate(row_vals):
                for kw in keywords:
                    if kw in cell and cell != "":
                        if field not in col_map:
                            col_map[field] = j
                            score += 1
                        break

        if score > best_score and score >= 3:  # au moins 3 colonnes reconnues
            best_score = score
            best_row_idx = i
            best_col_map = col_map

    return best_row_idx, best_col_map

def detect_subsection_pattern(vals, col_ref):
    """
    Détecte une sous-section : texte alphanumérique dans col ref,
    rien dans les colonnes suivantes (qte, long).
    Patterns supportés : DO0-01, DE1-09, F01, FEN-01, BAT-A, etc.
    """
    ref = safe_str(vals[col_ref]) if col_ref < len(vals) else ""
    if not ref:
        return False
    # Vérifier que les colonnes données (qte, long) sont vides
    next_vals = [safe_str(vals[col_ref + k]) for k in range(1, 5) if col_ref + k < len(vals)]
    nums = [v for v in next_vals if safe_num(v) is not None]
    if nums:
        return False
    # Pattern : commence par lettres + chiffres ou tiret (codes Lefevre)
    return bool(re.match(r'^[A-Za-z]{1,5}[-_]?\d', ref) or
                re.match(r'^[A-Za-z]{2,}[-_]\d', ref) or
                re.match(r'^[A-Z]{1,4}\d{1,3}', ref))

def detect_section_pattern(vals, col_ref, col_map):
    """
    Détecte une section principale : texte seul sans données chiffrées.
    Exemples : "Façade Ouest", "Façade Est", "Toiture", etc.
    """
    ref = safe_str(vals[col_ref]) if col_ref < len(vals) else ""
    if not ref:
        return False
    if detect_subsection_pattern(vals, col_ref):
        return False
    # Aucune valeur numérique dans les colonnes de données
    data_cols = [col_map.get("qte"), col_map.get("long"), col_map.get("haut"), col_map.get("prof")]
    for c in data_cols:
        if c and c < len(vals) and safe_num(vals[c]) is not None:
            return False
    return True

def detect_unit(values):
    """
    Détecte si les dimensions sont en cm ou mm en analysant les valeurs.
    Si la moyenne des valeurs > 200 → probablement mm, sinon cm.
    """
    nums = [v for v in values if v is not None and v > 0]
    if not nums:
        return "cm"
    avg = sum(nums) / len(nums)
    return "mm" if avg > 300 else "cm"

def parse_header_flexible(df_raw):
    """Extrait l'en-tête du fichier de façon flexible (positions variables)."""
    info = {"chantier": "", "lieu_taille": "", "code_chantier": "",
            "cubage_total": None, "nb_blocs": None}

    kw_chantier   = ["chantier", "affaire", "projet", "site"]
    kw_lieu       = ["lieu de taille", "lieu taille", "tailleur", "atelier"]
    kw_code       = ["code chantier", "code affaire", "ref chantier", "n° chantier"]
    kw_cubage     = ["cubage", "volume total", "total m3", "total cube"]
    kw_blocs      = ["nombre de blocs", "nb blocs", "total blocs", "nbre blocs"]

    for i, row in df_raw.iterrows():
        if i > 20:
            break
        row_str = [safe_str(v) for v in row]
        full_row = " | ".join(row_str).lower()

        for j, cell in enumerate(row_str):
            cell_l = cell.lower()
            # Chantier
            if any(kw in cell_l for kw in kw_chantier) and not info["chantier"]:
                # Chercher la valeur dans les cellules suivantes
                for k in range(j+1, min(j+8, len(row_str))):
                    if row_str[k] and not any(kw in row_str[k].lower() for kw in kw_chantier):
                        info["chantier"] = row_str[k]
                        break
            # Lieu de taille
            if any(kw in cell_l for kw in kw_lieu) and not info["lieu_taille"]:
                for k in range(j+1, min(j+8, len(row_str))):
                    if row_str[k]:
                        info["lieu_taille"] = row_str[k]
                        break
            # Code chantier
            if any(kw in cell_l for kw in kw_code) and not info["code_chantier"]:
                for k in range(j+1, min(j+8, len(row_str))):
                    if row_str[k]:
                        info["code_chantier"] = row_str[k]
                        break

        # Cubage / blocs (valeur sur la même ligne)
        if any(kw in full_row for kw in kw_cubage) and not info["cubage_total"]:
            for v in row:
                n = safe_num(v)
                if n and n > 0.01:
                    info["cubage_total"] = n
                    break
        if any(kw in full_row for kw in kw_blocs) and not info["nb_blocs"]:
            for v in row:
                n = safe_num(v)
                if n and n > 1:
                    info["nb_blocs"] = int(n)
                    break

    return info

def parse_lefevre(filepath, nature_mapping, sheet_name=None):
    if not HAS_PANDAS:
        raise RuntimeError("Le module 'pandas' est requis.")

    # ── Lecture du fichier ──────────────────────────────────────────
    xl = pd.ExcelFile(filepath)
    sheets = xl.sheet_names

    # Sélection de la feuille : paramètre > 1ère non vide
    if sheet_name and sheet_name in sheets:
        target_sheet = sheet_name
    else:
        target_sheet = sheets[0]
        # Préférer une feuille avec un nom pertinent si plusieurs
        for s in sheets:
            sl = s.lower()
            if any(k in sl for k in ["taille", "pierre", "devis", "commande", "feuil", "sheet"]):
                target_sheet = s
                break

    df_raw = pd.read_excel(filepath, sheet_name=target_sheet, header=None)

    # ── En-tête ─────────────────────────────────────────────────────
    header_info = parse_header_flexible(df_raw)
    header_info["sheet_used"] = target_sheet
    header_info["all_sheets"] = sheets

    # ── Auto-détection des colonnes ─────────────────────────────────
    header_row_idx, col_map = detect_columns(df_raw)

    # Fallback si détection échoue : structure Lefevre standard
    if not col_map or "ref" not in col_map:
        col_map = {"ref": 1, "nature": 2, "qte": 3, "long": 4,
                   "haut": 5, "prof": 6, "cube": 7, "poids": 8}
        header_row_idx = None

    col_ref   = col_map.get("ref",   1)
    col_nat   = col_map.get("nature", col_ref + 1)
    col_qte   = col_map.get("qte",   col_ref + 2)
    col_long  = col_map.get("long",  col_ref + 3)
    col_haut  = col_map.get("haut",  col_ref + 4)
    col_prof  = col_map.get("prof",  col_ref + 5)
    col_cube  = col_map.get("cube",  col_ref + 6)
    col_poids = col_map.get("poids", col_ref + 7)

    # ── Détection unité (cm vs mm) ──────────────────────────────────
    sample_longs = []
    for _, row in df_raw.iterrows():
        vals = list(row) + [None] * 20
        n = safe_num(vals[col_long])
        if n and n > 0:
            sample_longs.append(n)
        if len(sample_longs) >= 20:
            break
    unit = detect_unit(sample_longs)
    divisor = 10.0 if unit == "mm" else 100.0

    # ── Parsing des lignes ──────────────────────────────────────────
    parsed_rows = []
    current_section = ""
    current_subsection = ""
    in_data = False

    # Mots-clés qui déclenchent le début des données
    start_keywords = ["désignation", "designation", "n°", "repere", "repère",
                      "long", "haut", "prof", "cube"]

    for i, row in df_raw.iterrows():
        vals = list(row) + [None] * 20
        ref_cell = safe_str(vals[col_ref]).lower()

        # Démarrage de la zone données
        if not in_data:
            if header_row_idx is not None and i > header_row_idx:
                in_data = True
            elif any(kw in ref_cell for kw in start_keywords):
                in_data = True
                continue

        if not in_data:
            continue

        ref  = safe_str(vals[col_ref])
        nat  = safe_str(vals[col_nat])
        qte  = safe_num(vals[col_qte])
        long_raw  = safe_num(vals[col_long])
        haut_raw  = safe_num(vals[col_haut])
        prof_raw  = safe_num(vals[col_prof])
        cube_raw  = safe_num(vals[col_cube])
        poids_raw = safe_num(vals[col_poids])

        # Ligne vide → skip
        if not ref and not nat and qte is None and long_raw is None:
            continue

        # Ligne de total (chiffre seul en col cube sans ref/qte) → skip
        if not ref and qte is None and cube_raw is not None and long_raw is None:
            continue

        # ── Détection type de ligne ──────────────────────────────
        if qte is not None and qte > 0 and long_raw is not None and long_raw > 0:
            # LIGNE PIERRE
            long_m = round(long_raw / divisor, 5)
            haut_m = round(haut_raw / divisor, 5) if haut_raw else None
            prof_m = round(prof_raw / divisor, 5) if prof_raw else None

            qty_m3 = cube_raw
            if qty_m3 is None and long_m and haut_m and prof_m:
                qty_m3 = round(long_m * haut_m * prof_m * qte, 6)
            elif qty_m3:
                qty_m3 = round(qty_m3, 6)

            # Mapping produit
            prod_info = nature_mapping.get(nat)
            if prod_info is None:
                for k, v in nature_mapping.items():
                    if k.lower() == nat.lower():
                        prod_info = v
                        break

            parsed_rows.append({
                "type":         "line",
                "ref":          ref if ref else "SUP",
                "nature":       nat,
                "qte":          int(qte) if qte == int(qte) else qte,
                "long_m":       long_m,
                "haut_m":       haut_m,
                "prof_m":       prof_m,
                "qty_m3":       qty_m3,
                "poids":        poids_raw,
                "section":      current_section,
                "subsection":   current_subsection,
                "product":      prod_info["product"]      if prod_info else f"[INCONNU] {nat}",
                "product_code": prod_info["product_code"] if prod_info else "???",
                "mapped":       prod_info is not None,
            })

        elif detect_subsection_pattern(vals, col_ref):
            # SOUS-SECTION
            current_subsection = ref
            parsed_rows.append({
                "type": "subsection", "label": ref,
                "section": current_section, "subsection": ref,
            })

        elif detect_section_pattern(vals, col_ref, col_map):
            # SECTION
            current_section = ref
            current_subsection = ""
            parsed_rows.append({
                "type": "section", "label": ref,
                "section": ref, "subsection": "",
            })

    # Recalcul cubage si absent de l'en-tête
    if not header_info.get("cubage_total"):
        header_info["cubage_total"] = round(
            sum(r.get("qty_m3") or 0 for r in parsed_rows if r["type"] == "line"), 6)

    header_info["col_map"]    = col_map
    header_info["unit"]       = unit
    header_info["header_row"] = header_row_idx

    return header_info, parsed_rows


# ══════════════════════════════════════════════════════════════════════
# CONNECTEUR ODOO XML-RPC
# ══════════════════════════════════════════════════════════════════════

class OdooConnector:
    def __init__(self, url, database, username, password):
        self.url = url.rstrip("/")
        self.db = database
        self.username = username
        self.password = password
        self.uid = None
        self._models = None

    def connect(self):
        common = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/common", allow_none=True)
        self.uid = common.authenticate(self.db, self.username, self.password, {})
        if not self.uid:
            raise ConnectionError("Authentification échouée. Vérifiez vos identifiants Odoo.")
        self._models = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/object", allow_none=True)
        version = common.version()
        return version.get("server_version", "?")

    def call(self, model, method, args, kwargs=None):
        if kwargs is None:
            kwargs = {}
        return self._models.execute_kw(self.db, self.uid, self.password,
                                        model, method, args, kwargs)

    def search_partner(self, name=None, ref=None):
        domain = []
        if ref:
            domain = [["ref", "=", ref]]
            res = self.call("res.partner", "search_read", [domain], {"fields": ["id", "name", "ref"], "limit": 1})
            if res:
                return res[0]
        if name:
            domain = [["name", "ilike", name]]
            res = self.call("res.partner", "search_read", [domain], {"fields": ["id", "name", "ref"], "limit": 1})
            if res:
                return res[0]
        return None

    def search_product(self, default_code):
        """Cherche un produit par code interne. Utilise le cache."""
        if not hasattr(self, "_product_cache"):
            self._product_cache = {}
        if default_code in self._product_cache:
            return self._product_cache[default_code]

        result = None

        # 1) product.product exact
        res = self.call("product.product", "search_read",
                        [[["default_code", "=", default_code]]],
                        {"fields": ["id", "name", "default_code", "uom_id"], "limit": 1})
        if res:
            result = res[0]

        # 2) product.template exact → prendre la première variante
        if not result:
            tpl = self.call("product.template", "search_read",
                            [[["default_code", "=", default_code]]],
                            {"fields": ["id", "name", "default_code", "product_variant_ids"], "limit": 1})
            if tpl and tpl[0].get("product_variant_ids"):
                variant_id = tpl[0]["product_variant_ids"][0]
                v = self.call("product.product", "read", [[variant_id]],
                              {"fields": ["id", "name", "default_code", "uom_id"]})
                if v:
                    result = v[0]

        # 3) product.product recherche partielle (code contenu dans le nom ou code)
        if not result:
            res = self.call("product.product", "search_read",
                            [[["default_code", "ilike", default_code]]],
                            {"fields": ["id", "name", "default_code", "uom_id"], "limit": 1})
            if res:
                result = res[0]

        self._product_cache[default_code] = result  # None aussi mis en cache
        return result

    def preload_products(self, codes):
        """Précharge tous les produits nécessaires en 2 appels API (product.product + product.template)."""
        if not hasattr(self, "_product_cache"):
            self._product_cache = {}
        codes = [c for c in set(codes) if c and not c.startswith("???")]

        # Batch sur product.product
        res = self.call("product.product", "search_read",
                        [[["default_code", "in", codes]]],
                        {"fields": ["id", "name", "default_code", "uom_id"]})
        for p in res:
            if p.get("default_code"):
                self._product_cache[p["default_code"]] = p

        # Les codes non trouvés → essayer product.template
        missing = [c for c in codes if c not in self._product_cache]
        if missing:
            tpls = self.call("product.template", "search_read",
                             [[["default_code", "in", missing]]],
                             {"fields": ["id", "name", "default_code", "product_variant_ids"]})
            variant_ids = []
            tpl_map = {}
            for t in tpls:
                if t.get("product_variant_ids"):
                    vid = t["product_variant_ids"][0]
                    variant_ids.append(vid)
                    tpl_map[vid] = t["default_code"]

            if variant_ids:
                variants = self.call("product.product", "read", [variant_ids],
                                     {"fields": ["id", "name", "default_code", "uom_id"]})
                for v in variants:
                    code = tpl_map.get(v["id"])
                    if code:
                        self._product_cache[code] = v

        # Marquer les introuvables
        for c in codes:
            if c not in self._product_cache:
                self._product_cache[c] = None

        found = sum(1 for v in self._product_cache.values() if v)
        return found, len(codes)

    def diagnose_product(self, code):
        """Retourne des infos de diagnostic sur un code produit."""
        lines = []
        # Exact product.product
        r1 = self.call("product.product", "search_read",
                       [[["default_code", "=", code]]],
                       {"fields": ["id", "name", "default_code", "active"], "limit": 5})
        lines.append(f"product.product exact '{code}': {len(r1)} résultat(s)")
        for p in r1:
            lines.append(f"  → id={p['id']} | {p['name']} | actif={p.get('active')}")

        # Exact product.template
        r2 = self.call("product.template", "search_read",
                       [[["default_code", "=", code]]],
                       {"fields": ["id", "name", "default_code", "active"], "limit": 5})
        lines.append(f"product.template exact '{code}': {len(r2)} résultat(s)")
        for p in r2:
            lines.append(f"  → id={p['id']} | {p['name']} | actif={p.get('active')}")

        # Partial product.template
        r3 = self.call("product.template", "search_read",
                       [[["default_code", "ilike", code], ["active", "in", [True, False]]]],
                       {"fields": ["id", "name", "default_code", "active"], "limit": 5})
        lines.append(f"product.template partiel '{code}' (actifs+archivés): {len(r3)} résultat(s)")
        for p in r3:
            lines.append(f"  → id={p['id']} | code={p.get('default_code')} | {p['name']} | actif={p.get('active')}")

        # Partial product.product avec archivés
        r4 = self.call("product.product", "search_read",
                       [[["default_code", "ilike", code], ["active", "in", [True, False]]]],
                       {"fields": ["id", "name", "default_code", "active"], "limit": 5})
        lines.append(f"product.product partiel '{code}' (actifs+archivés): {len(r4)} résultat(s)")
        for p in r4:
            lines.append(f"  → id={p['id']} | code={p.get('default_code')} | {p['name']} | actif={p.get('active')}")

        return lines

    def get_uom_m3(self):
        res = self.call("uom.uom", "search_read",
                        [[["name", "ilike", "m³"]]],
                        {"fields": ["id", "name"], "limit": 1})
        if not res:
            res = self.call("uom.uom", "search_read",
                            [[["name", "ilike", "m3"]]],
                            {"fields": ["id", "name"], "limit": 1})
        return res[0]["id"] if res else False

    def create_sale_order(self, partner_id, order_ref, date_order, chantier=""):
        vals = {
            "partner_id": partner_id,
            "state": "draft",
        }
        if date_order:
            # Odoo attend le format ISO : YYYY-MM-DD HH:MM:SS
            # On convertit depuis DD/MM/YYYY HH:MM:SS si nécessaire
            iso_date = date_order
            for fmt_in in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    iso_date = datetime.strptime(date_order, fmt_in).strftime("%Y-%m-%d %H:%M:%S")
                    break
                except ValueError:
                    continue
            vals["date_order"] = iso_date
        if order_ref:
            vals["client_order_ref"] = order_ref
        return self.call("sale.order", "create", [vals])

    def create_order_line_section(self, order_id, name):
        return self.call("sale.order.line", "create", [{
            "order_id": order_id,
            "display_type": "line_section",
            "name": name,
        }])

    def create_order_line_product(self, order_id, line_data, fields_cfg, uom_id=False):
        product = self.search_product(line_data["product_code"])
        if not product:
            return None, f"Produit introuvable: {line_data['product_code']}"

        vals = {
            "order_id": order_id,
            "product_id": product["id"],
            "product_uom_qty": line_data.get("qty_m3") or 0,
            "name": line_data.get("product", ""),
        }
        if uom_id:
            vals["product_uom"] = uom_id

        # Champs Studio personnalisés
        fld_ref  = fields_cfg.get("ref_pierre", "x_studio_ref_pierre")
        fld_nbr  = fields_cfg.get("nbr",        "x_studio_nbr")
        fld_long = fields_cfg.get("long",        "x_studio_long")
        fld_larg = fields_cfg.get("larg",        "x_studio_larg")
        fld_haut = fields_cfg.get("haut",        "x_studio_haut")
        fld_poids = fields_cfg.get("poids",      "x_studio_poids")

        if line_data.get("ref"):
            vals[fld_ref] = line_data["ref"]
        if line_data.get("qte"):
            vals[fld_nbr] = line_data["qte"]
        if line_data.get("long_m"):
            vals[fld_long] = line_data["long_m"]
        if line_data.get("prof_m"):
            vals[fld_larg] = line_data["prof_m"]
        if line_data.get("haut_m"):
            vals[fld_haut] = line_data["haut_m"]
        if line_data.get("poids"):
            vals[fld_poids] = line_data["poids"]

        try:
            line_id = self.call("sale.order.line", "create", [vals])
            return line_id, None
        except Exception as e:
            # Retry sans les champs Studio si erreur
            for fld in [fld_ref, fld_nbr, fld_long, fld_larg, fld_haut, fld_poids]:
                vals.pop(fld, None)
            try:
                line_id = self.call("sale.order.line", "create", [vals])
                return line_id, f"⚠ Champs Studio ignorés pour {line_data['ref']}: {e}"
            except Exception as e2:
                return None, f"❌ Erreur ligne {line_data['ref']}: {e2}"

    def check_studio_fields(self, fields_cfg):
        """Vérifie que les champs Studio existent dans Odoo"""
        fields_info = self.call("sale.order.line", "fields_get",
                                [list(fields_cfg.values())],
                                {"attributes": ["string", "type"]})
        results = {}
        for key, fname in fields_cfg.items():
            results[key] = fname in fields_info
        return results


# ══════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════

COLS_ODOO = [
    "Order Reference", "Order Date", "Customer", "Customer Code",
    "Order Lines / Display Type", "Order Lines / Name",
    "Order Lines / Ref. Pierre", "Order Lines / Product",
    "Order Lines / Product Code", "Order Lines / Nbr.",
    "Order Lines / Long.", "Order Lines / Larg.", "Order Lines / Haut.",
    "Order Lines / Quantity", "Order Lines / Poids (kg)",
]

def fmt_num(val, d=3):
    return str(round(val, d)).replace(".", ",") if val is not None else ""

def generate_excel(header_info, parsed_rows, order_ref, customer, customer_code, order_date, output_path):
    if not HAS_OPENPYXL:
        raise RuntimeError("Le module 'openpyxl' est requis.")

    def rgb(h): return "FF" + h.lstrip("#").upper()
    def sc(cell, bold=False, bg=None, fg="000000", sz=9, align="left", border=False):
        cell.font = Font(name="Calibri", size=sz, bold=bold, color=fg)
        if bg: cell.fill = PatternFill("solid", start_color=rgb(bg))
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if border:
            t = Side(style="thin", color="BBBBBB")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

    C_H = "1A3A5C"; C_S = "2E6DA4"; C_SS = "5B9BD5"
    C_ALT = "EEF4FB"; C_W = "FFFFFF"; C_WARN = "FFF2CC"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odoo_Import"
    ws.row_dimensions[1].height = 28

    for ci, cn in enumerate(COLS_ODOO, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        sc(c, bold=True, bg=C_H, fg=C_W, sz=10, align="center", border=True)

    row_num = 2
    first_line = True

    def base():
        r = {c: "" for c in COLS_ODOO}
        if first_line:
            r["Order Reference"] = order_ref
            r["Order Date"] = order_date
            r["Customer"] = customer
            r["Customer Code"] = customer_code
        return r

    for i, item in enumerate(parsed_rows):
        ws.row_dimensions[row_num].height = 17
        if item["type"] in ("section", "subsection"):
            r = base()
            r["Order Lines / Display Type"] = "line_section"
            lbl = item["label"] if item["type"] == "section" else f"  {item['label']}"
            r["Order Lines / Name"] = lbl
            bg = C_S if item["type"] == "section" else C_SS
            for ci, cn in enumerate(COLS_ODOO, 1):
                c = ws.cell(row=row_num, column=ci, value=r[cn])
                sc(c, bold=True, bg=bg, fg=C_W, sz=10, border=True)
        else:
            r = base()
            r["Order Lines / Ref. Pierre"] = item["ref"]
            r["Order Lines / Product"] = item["product"]
            r["Order Lines / Product Code"] = item["product_code"]
            r["Order Lines / Nbr."] = item["qte"]
            r["Order Lines / Long."] = fmt_num(item["long_m"])
            r["Order Lines / Larg."] = fmt_num(item["prof_m"])
            r["Order Lines / Haut."] = fmt_num(item["haut_m"])
            r["Order Lines / Quantity"]   = fmt_num(item["qty_m3"], 6)
            r["Order Lines / Poids (kg)"] = fmt_num(item.get("poids"), 3) if item.get("poids") else ""
            is_unk = item["product_code"].startswith("???")
            bg = C_WARN if is_unk else (C_ALT if i % 2 == 0 else C_W)
            for ci, cn in enumerate(COLS_ODOO, 1):
                c = ws.cell(row=row_num, column=ci, value=r[cn])
                sc(c, bg=bg, sz=9, border=True)

        first_line = False
        row_num += 1

    widths = [18, 20, 22, 14, 16, 28, 16, 50, 16, 7, 7, 7, 7, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Récap
    ws2 = wb.create_sheet("Récap")
    for ci, h in enumerate(["Section", "Sous-section", "Blocs", "Volume m³", "Poids kg"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        sc(c, bold=True, bg=C_H, fg=C_W, sz=10, align="center", border=True)
    from collections import defaultdict
    agg = defaultdict(lambda: {"n": 0, "v": 0.0, "p": 0.0})
    for item in parsed_rows:
        if item["type"] == "line":
            key = (item["section"], item["subsection"])
            agg[key]["n"] += item["qte"] if isinstance(item["qte"], int) else 1
            agg[key]["v"] += item["qty_m3"] or 0
            agg[key]["p"] += item["poids"] or 0
    r = 2; prev_sec = None
    for (sec, sub), v in sorted(agg.items()):
        new_sec = sec != prev_sec; prev_sec = sec
        row_vals = [sec if new_sec else "", sub, v["n"], round(v["v"], 6), round(v["p"], 2)]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=r, column=ci, value=val)
            bg = C_S if new_sec and ci == 1 else (C_ALT if r % 2 == 0 else C_W)
            sc(c, bold=new_sec and ci == 1, bg=bg, fg=C_W if bg == C_S else "000000", sz=9, border=True)
        r += 1
    tot = [v for v in agg.values()]
    tv, tp, tn = sum(x["v"] for x in tot), sum(x["p"] for x in tot), sum(x["n"] for x in tot)
    for ci, val in enumerate(["TOTAL", "", tn, round(tv, 6), round(tp, 2)], 1):
        c = ws2.cell(row=r, column=ci, value=val)
        sc(c, bold=True, bg=C_H, fg=C_W, sz=10, align="center", border=True)
    for ci, w in enumerate([22, 18, 10, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    return tv, tn


# ══════════════════════════════════════════════════════════════════════
# INTERFACE GRAPHIQUE PRINCIPALE
# ══════════════════════════════════════════════════════════════════════

class LefevreMaquignonApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Maquignon — {APP_NAME} v{APP_VERSION}")
        self.root.geometry("1100x700")
        self.root.minsize(900, 600)
        self.root.configure(bg=COLORS["bg"])

        self.config = Config()
        self.parsed_rows = []
        self.header_info = {}
        self.selected_file = tk.StringVar()
        self.sheet_var = tk.StringVar(value="")
        self.order_ref = tk.StringVar(value=f"BC{datetime.now().strftime('%Y%m%d')}")
        self.customer = tk.StringVar(value=self.config.get("commande", "customer"))
        self.customer_code = tk.StringVar(value=self.config.get("commande", "customer_code"))
        self.order_date = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y 00:00:00"))
        self.odoo_url = tk.StringVar(value=self.config.get("odoo", "url"))
        self.odoo_db = tk.StringVar(value=self.config.get("odoo", "database"))
        self.odoo_user = tk.StringVar(value=self.config.get("odoo", "username"))
        self.odoo_pass = tk.StringVar(value=self.config.get("odoo", "password"))
        self.status_var = tk.StringVar(value="Prêt")
        self.progress_var = tk.DoubleVar(value=0)
        self._connector = None

        self._apply_style()
        self._build_ui()

    # ── Style ──────────────────────────────────────────────────────────
    def _apply_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        C = COLORS
        style.configure(".", background=C["card"], foreground=C["text"],
                        font=("Segoe UI", 10), fieldbackground=C["entry_bg"], bordercolor=C["border"])
        style.configure("TNotebook", background=C["bg"], borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure("TNotebook.Tab", background=C["sidebar"], foreground=C["text_dim"],
                        padding=[18, 8], font=("Segoe UI", 10))
        style.map("TNotebook.Tab",
                  background=[("selected", C["accent"])],
                  foreground=[("selected", C["white"])])
        style.configure("TLabel", background=C["card"], foreground=C["text"], font=("Segoe UI", 10))
        style.configure("Dim.TLabel", background=C["card"], foreground=C["text_dim"], font=("Segoe UI", 9))
        style.configure("Header.TLabel", background=C["card"], foreground=C["text"],
                        font=("Segoe UI", 12, "bold"))
        style.configure("TEntry", fieldbackground=C["entry_bg"], foreground=C["text"],
                        insertcolor=C["text"], bordercolor=C["border"], relief="flat")
        style.configure("TButton", background=C["accent"], foreground=C["white"],
                        font=("Segoe UI", 10, "bold"), relief="flat", padding=[12, 6])
        style.map("TButton",
                  background=[("active", C["btn_hover"]), ("disabled", C["border"])],
                  foreground=[("disabled", C["text_dim"])])
        style.configure("Success.TButton", background=C["success"])
        style.map("Success.TButton", background=[("active", "#2ECC71")])
        style.configure("Warning.TButton", background=C["warning"])
        style.configure("TProgressbar", background=C["accent"], troughcolor=C["border"],
                        thickness=6, borderwidth=0)
        style.configure("Treeview", background=C["card"], foreground=C["text"],
                        fieldbackground=C["card"], borderwidth=0, rowheight=22,
                        font=("Segoe UI", 9))
        style.configure("Treeview.Heading", background=C["sidebar"], foreground=C["text"],
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", C["accent"])],
                  foreground=[("selected", C["white"])])
        style.configure("TScrollbar", background=C["border"], troughcolor=C["card"],
                        arrowcolor=C["text_dim"])
        style.configure("TFrame", background=C["card"])
        style.configure("Sidebar.TFrame", background=C["sidebar"])
        style.configure("Card.TFrame", background=C["card"])
        style.configure("TSeparator", background=C["border"])

    # ── Construction UI ────────────────────────────────────────────────
    def _build_ui(self):
        C = COLORS

        # Header
        header = tk.Frame(self.root, bg=C["sidebar"], height=56)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)
        tk.Label(header, text="⬡ MAQUIGNON", bg=C["sidebar"], fg=C["accent2"],
                 font=("Segoe UI", 14, "bold")).pack(side="left", padx=20, pady=14)
        tk.Label(header, text=f"Convertisseur Commandes Lefevre → Odoo  v{APP_VERSION}",
                 bg=C["sidebar"], fg=C["text_dim"], font=("Segoe UI", 10)).pack(side="left", padx=4, pady=14)

        # Barre de statut bas
        status_bar = tk.Frame(self.root, bg=C["sidebar"], height=32)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        tk.Label(status_bar, textvariable=self.status_var, bg=C["sidebar"],
                 fg=C["text_dim"], font=("Segoe UI", 9)).pack(side="left", padx=12, pady=6)
        self.progress = ttk.Progressbar(status_bar, variable=self.progress_var,
                                         maximum=100, mode="determinate", length=200)
        self.progress.pack(side="right", padx=12, pady=10)

        # Notebook principal
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=0, pady=0)

        self._tab_fichier()
        self._tab_odoo()
        self._tab_apercu()
        self._tab_log()
        self._tab_mapping()

    # ── Tab 1 : Fichier & Import ───────────────────────────────────────
    def _tab_fichier(self):
        tab = ttk.Frame(self.nb, style="TFrame")
        self.nb.add(tab, text="  📂  Fichier & Import  ")

        # Layout 2 colonnes
        left = ttk.Frame(tab, style="Card.TFrame")
        left.pack(side="left", fill="both", expand=False, padx=(12, 6), pady=12)
        left.configure(width=380)

        right = ttk.Frame(tab, style="Card.TFrame")
        right.pack(side="left", fill="both", expand=True, padx=(6, 12), pady=12)

        # ── Panneau gauche ──
        self._section_title(left, "Fichier Lefevre")

        file_frame = ttk.Frame(left, style="Card.TFrame")
        file_frame.pack(fill="x", padx=12, pady=(0, 4))
        ttk.Label(file_frame, text="Fichier source (.xlsx, .xlsm)", style="Dim.TLabel").pack(anchor="w")
        entry_frame = ttk.Frame(file_frame, style="Card.TFrame")
        entry_frame.pack(fill="x", pady=4)
        self.file_entry = ttk.Entry(entry_frame, textvariable=self.selected_file, state="readonly")
        self.file_entry.pack(side="left", fill="x", expand=True, ipady=5)
        ttk.Button(entry_frame, text="…", command=self._pick_file, width=3).pack(side="left", padx=(4, 0))

        # Sélecteur de feuille
        sheet_frame = ttk.Frame(left, style="Card.TFrame")
        sheet_frame.pack(fill="x", padx=12, pady=(0, 4))
        ttk.Label(sheet_frame, text="Feuille à utiliser", style="Dim.TLabel", width=24, anchor="w").pack(side="left")
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var,
                                         state="readonly", width=20)
        self.sheet_combo.pack(side="left", fill="x", expand=True)
        ttk.Label(sheet_frame, text="(auto)", style="Dim.TLabel").pack(side="left", padx=4)

        # Bandeau détection (rempli après analyse)
        self.detect_frame = tk.Frame(left, bg=COLORS["section_bg"])
        self.detect_frame.pack(fill="x", padx=12, pady=(0, 4))
        self.detect_label = tk.Label(self.detect_frame, text="",
                                      bg=COLORS["section_bg"], fg=COLORS["text_dim"],
                                      font=("Segoe UI", 8), anchor="w", justify="left", wraplength=330)
        self.detect_label.pack(fill="x", padx=8, pady=4)

        ttk.Separator(left, orient="horizontal").pack(fill="x", padx=12, pady=8)
        self._section_title(left, "Paramètres commande")

        fields = [
            ("Référence commande", self.order_ref),
            ("Client (nom Odoo)", self.customer),
            ("Code client Odoo", self.customer_code),
            ("Date commande", self.order_date),
        ]
        for label, var in fields:
            self._field_row(left, label, var)

        ttk.Separator(left, orient="horizontal").pack(fill="x", padx=12, pady=8)

        # Boutons action
        btn_frame = ttk.Frame(left, style="Card.TFrame")
        btn_frame.pack(fill="x", padx=12, pady=4)

        ttk.Button(btn_frame, text="📊  Analyser le fichier",
                   command=self._parse_file).pack(fill="x", pady=3)
        ttk.Button(btn_frame, text="💾  Exporter Excel (Odoo_Import)",
                   command=self._export_excel).pack(fill="x", pady=3)
        ttk.Button(btn_frame, text="🚀  Importer dans Odoo (API)",
                   command=self._import_odoo, style="Success.TButton").pack(fill="x", pady=3)

        # ── Panneau droit : résumé ──
        self._section_title(right, "Résumé du fichier parsé")

        info_frame = tk.Frame(right, bg=COLORS["section_bg"], bd=0)
        info_frame.pack(fill="x", padx=12, pady=(0, 8))

        self.info_labels = {}
        info_fields = [
            ("chantier", "Chantier"),
            ("lieu", "Lieu de taille"),
            ("code", "Code chantier"),
            ("sections", "Sections"),
            ("subsections", "Sous-sections"),
            ("lignes", "Lignes pierres"),
            ("cubage", "Cubage (m³)"),
            ("blocs", "Nb blocs"),
            ("inconnus", "⚠ Produits inconnus"),
        ]
        for key, label in info_fields:
            row = tk.Frame(info_frame, bg=COLORS["section_bg"])
            row.pack(fill="x", padx=10, pady=3)
            tk.Label(row, text=label + " :", bg=COLORS["section_bg"], fg=COLORS["text_dim"],
                     font=("Segoe UI", 9), width=22, anchor="w").pack(side="left")
            lbl = tk.Label(row, text="—", bg=COLORS["section_bg"], fg=COLORS["text"],
                           font=("Segoe UI", 9, "bold"), anchor="w")
            lbl.pack(side="left")
            self.info_labels[key] = lbl

        ttk.Separator(right, orient="horizontal").pack(fill="x", padx=12, pady=8)
        self._section_title(right, "Mapping Natures → Produits Odoo")

        mapping_top = ttk.Frame(right, style="Card.TFrame")
        mapping_top.pack(fill="x", padx=12, pady=(0, 4))
        ttk.Label(mapping_top, text="Aperçu — cliquer sur l'onglet Mapping pour modifier",
                  style="Dim.TLabel").pack(side="left")
        ttk.Button(mapping_top, text="✏️ Éditer",
                   command=lambda: self.nb.select(4)).pack(side="right")

        mapping_frame = tk.Frame(right, bg=COLORS["section_bg"])
        mapping_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        cols = ("Nature Lefevre", "Code produit Odoo")
        self.mapping_tree = ttk.Treeview(mapping_frame, columns=cols, show="headings", height=8)
        for col in cols:
            self.mapping_tree.heading(col, text=col)
        self.mapping_tree.column("Nature Lefevre", width=140)
        self.mapping_tree.column("Code produit Odoo", width=160)
        self.mapping_tree.pack(fill="both", expand=True, side="left")
        sb = ttk.Scrollbar(mapping_frame, orient="vertical", command=self.mapping_tree.yview)
        sb.pack(side="right", fill="y")
        self.mapping_tree.configure(yscrollcommand=sb.set)
        self._refresh_mapping_tree()

    # ── Tab 2 : Connexion Odoo ─────────────────────────────────────────
    def _tab_odoo(self):
        tab = ttk.Frame(self.nb, style="TFrame")
        self.nb.add(tab, text="  🔗  Connexion Odoo  ")

        frame = ttk.Frame(tab, style="Card.TFrame")
        frame.pack(fill="both", expand=True, padx=12, pady=12)

        left = ttk.Frame(frame, style="Card.TFrame")
        left.pack(side="left", fill="both", expand=False, padx=(0, 8))
        left.configure(width=400)

        right = ttk.Frame(frame, style="Card.TFrame")
        right.pack(side="left", fill="both", expand=True)

        # ── Connexion ──
        self._section_title(left, "Paramètres de connexion Odoo")
        for label, var, pw in [
            ("URL instance", self.odoo_url, False),
            ("Base de données", self.odoo_db, False),
            ("Utilisateur (email)", self.odoo_user, False),
            ("Mot de passe / Clé API", self.odoo_pass, True),
        ]:
            self._field_row(left, label, var, password=pw)

        ttk.Button(left, text="🔌  Tester la connexion",
                   command=self._test_connection).pack(fill="x", padx=12, pady=8)

        self.conn_status = tk.Label(left, text="", bg=COLORS["card"], font=("Segoe UI", 9))
        self.conn_status.pack(padx=12, anchor="w")

        ttk.Separator(left, orient="horizontal").pack(fill="x", padx=12, pady=8)
        self._section_title(left, "Champs Studio (sale.order.line)")

        self.field_vars = {}
        field_labels = [
            ("ref_pierre", "Réf. Pierre"),
            ("nbr",        "Nbr. pièces"),
            ("long",       "Longueur (m)"),
            ("larg",       "Largeur/Prof. (m)"),
            ("haut",       "Hauteur (m)"),
            ("poids",      "Poids (kg)"),
        ]
        for key, label in field_labels:
            var = tk.StringVar(value=self.config.get("fields", key, default="x_studio_" + key))
            self._field_row(left, label, var)
            self.field_vars[key] = var

        ttk.Button(left, text="✅  Vérifier les champs Studio",
                   command=self._check_studio_fields).pack(fill="x", padx=12, pady=(4, 0))

        ttk.Separator(left, orient="horizontal").pack(fill="x", padx=12, pady=8)
        self._section_title(left, "Diagnostic produits Odoo")

        diag_frame = ttk.Frame(left, style="Card.TFrame")
        diag_frame.pack(fill="x", padx=12, pady=2)
        ttk.Label(diag_frame, text="Code produit à tester", style="Dim.TLabel", width=24, anchor="w").pack(side="left")
        self.diag_code_var = tk.StringVar(value="T-P007")
        ttk.Entry(diag_frame, textvariable=self.diag_code_var).pack(side="left", fill="x", expand=True, ipady=4)

        ttk.Button(left, text="🔍  Diagnostiquer ce produit dans Odoo",
                   command=self._diagnose_product).pack(fill="x", padx=12, pady=(4, 0))

        ttk.Button(left, text="💾  Enregistrer la configuration",
                   command=self._save_config, style="Success.TButton").pack(fill="x", padx=12, pady=8)

        # ── Droite : Aide ──
        self._section_title(right, "Aide — Configuration Odoo")

        help_text = (
            "CONNEXION\n"
            "──────────────────────────────────────────\n"
            "• URL : https://votre-instance.odoo.com\n"
            "• Utilisateur : votre email de connexion Odoo\n"
            "• Mot de passe : votre mot de passe OU une clé API\n"
            "  (Paramètres → Technique → Clés API)\n\n"
            "CHAMPS STUDIO\n"
            "──────────────────────────────────────────\n"
            "Ces champs sont créés par Odoo Studio sur\n"
            "la ligne de commande (sale.order.line).\n\n"
            "Pour retrouver les noms techniques :\n"
            "  Ventes → Commandes → (ouvrir une commande)\n"
            "  → Activer mode développeur\n"
            "  → Passer la souris sur le champ → info-bulle\n\n"
            "Noms par défaut générés par Studio :\n"
            "  x_studio_ref_pierre\n"
            "  x_studio_nbr\n"
            "  x_studio_long\n"
            "  x_studio_larg\n"
            "  x_studio_haut\n\n"
            "⚠ Les caractères accentués sont supprimés\n"
            "des noms de champs Studio.\n\n"
            "IMPORT DIRECT API\n"
            "──────────────────────────────────────────\n"
            "L'import via API crée directement un devis\n"
            "(état Brouillon) dans Odoo avec les sections\n"
            "et sous-sections comme séparateurs de lignes.\n"
            "Les produits sont recherchés par code interne\n"
            "(ex: T-P007) — ils doivent exister dans Odoo.\n"
        )

        txt = tk.Text(right, bg=COLORS["section_bg"], fg=COLORS["text_dim"],
                      font=("Consolas", 9), relief="flat", padx=12, pady=12,
                      wrap="word", state="normal")
        txt.insert("1.0", help_text)
        txt.configure(state="disabled")
        txt.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    # ── Tab 3 : Aperçu ─────────────────────────────────────────────────
    def _tab_apercu(self):
        tab = ttk.Frame(self.nb, style="TFrame")
        self.nb.add(tab, text="  🔍  Aperçu données  ")

        top = ttk.Frame(tab, style="Card.TFrame")
        top.pack(fill="x", padx=12, pady=(12, 0))
        ttk.Label(top, text="Aperçu des lignes parsées (analysez d'abord un fichier)",
                  style="Dim.TLabel").pack(side="left", padx=12, pady=8)
        ttk.Button(top, text="⬅ Aller à Fichier & Import",
                   command=lambda: self.nb.select(0)).pack(side="right", padx=12)

        tree_frame = ttk.Frame(tab, style="Card.TFrame")
        tree_frame.pack(fill="both", expand=True, padx=12, pady=8)

        cols = ("Type", "Section", "Sous-section", "Réf. Pierre", "Nature",
                "Nbr", "Long (m)", "Larg (m)", "Haut (m)", "Vol. m³", "Produit Odoo", "Mapped")
        self.preview_tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=20)
        widths = [80, 120, 100, 80, 80, 40, 70, 70, 70, 80, 250, 60]
        for col, w in zip(cols, widths):
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=w, minwidth=40)
        self.preview_tree.tag_configure("section", background=COLORS["accent"], foreground=COLORS["white"])
        self.preview_tree.tag_configure("subsection", background=COLORS["section_bg"], foreground=COLORS["text"])
        self.preview_tree.tag_configure("unmapped", background="#4A2020")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.preview_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    # ── Tab 4 : Journal ────────────────────────────────────────────────
    def _tab_log(self):
        tab = ttk.Frame(self.nb, style="TFrame")
        self.nb.add(tab, text="  📋  Journal  ")

        top = ttk.Frame(tab, style="Card.TFrame")
        top.pack(fill="x", padx=12, pady=(12, 0))
        ttk.Label(top, text="Journal des opérations", style="Header.TLabel").pack(side="left", padx=12, pady=8)
        ttk.Button(top, text="🗑 Effacer", command=self._clear_log).pack(side="right", padx=12)

        self.log_text = scrolledtext.ScrolledText(
            tab, bg=COLORS["entry_bg"], fg=COLORS["text"],
            font=("Consolas", 9), relief="flat", padx=12, pady=12,
            insertbackground=COLORS["text"]
        )
        self.log_text.pack(fill="both", expand=True, padx=12, pady=8)
        self.log_text.tag_configure("ok", foreground=COLORS["success"])
        self.log_text.tag_configure("warn", foreground=COLORS["warning"])
        self.log_text.tag_configure("err", foreground=COLORS["error"])
        self.log_text.tag_configure("info", foreground=COLORS["accent2"])

    # ── Tab 5 : Mapping ────────────────────────────────────────────────
    def _tab_mapping(self):
        tab = ttk.Frame(self.nb, style="TFrame")
        self.nb.add(tab, text="  🗂️  Mapping  ")
        C = COLORS

        # En-tête
        top = ttk.Frame(tab, style="Card.TFrame")
        top.pack(fill="x", padx=12, pady=(12, 0))
        ttk.Label(top, text="Correspondance Natures de pierre (Lefevre) → Codes produits Odoo",
                  style="Header.TLabel").pack(side="left", padx=12, pady=8)

        # Layout : tableau à gauche, formulaire à droite
        body = ttk.Frame(tab, style="Card.TFrame")
        body.pack(fill="both", expand=True, padx=12, pady=8)

        left = ttk.Frame(body, style="Card.TFrame")
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        right = ttk.Frame(body, style="Card.TFrame")
        right.pack(side="left", fill="y", expand=False)
        right.configure(width=340)

        # ── Tableau mapping ──
        self._section_title(left, "Lignes de mapping actives")

        tree_frame = tk.Frame(left, bg=C["section_bg"])
        tree_frame.pack(fill="both", expand=True, padx=0, pady=(4, 0))

        edit_cols = ("Nature Lefevre", "Code produit Odoo", "Libellé produit")
        self.edit_mapping_tree = ttk.Treeview(tree_frame, columns=edit_cols,
                                               show="headings", height=16, selectmode="browse")
        self.edit_mapping_tree.heading("Nature Lefevre",    text="Nature Lefevre")
        self.edit_mapping_tree.heading("Code produit Odoo", text="Code produit Odoo")
        self.edit_mapping_tree.heading("Libellé produit",   text="Libellé Odoo (info)")
        self.edit_mapping_tree.column("Nature Lefevre",    width=140, minwidth=100)
        self.edit_mapping_tree.column("Code produit Odoo", width=160, minwidth=120)
        self.edit_mapping_tree.column("Libellé produit",   width=320, minwidth=200)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.edit_mapping_tree.yview)
        self.edit_mapping_tree.configure(yscrollcommand=vsb.set)
        self.edit_mapping_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.edit_mapping_tree.bind("<<TreeviewSelect>>", self._on_mapping_select)
        self.edit_mapping_tree.bind("<Double-1>", lambda e: self._mapping_edit())

        # Boutons sous le tableau
        btn_row = ttk.Frame(left, style="Card.TFrame")
        btn_row.pack(fill="x", pady=(6, 0))
        ttk.Button(btn_row, text="➕  Ajouter",
                   command=self._mapping_add).pack(side="left", padx=(0, 4))
        ttk.Button(btn_row, text="✏️  Modifier",
                   command=self._mapping_edit).pack(side="left", padx=4)
        ttk.Button(btn_row, text="🗑  Supprimer",
                   command=self._mapping_delete).pack(side="left", padx=4)
        ttk.Button(btn_row, text="💾  Enregistrer",
                   command=self._mapping_save, style="Success.TButton").pack(side="right")

        # ── Formulaire d'édition ──
        self._section_title(right, "Ajouter / Modifier une entrée")

        form = tk.Frame(right, bg=C["section_bg"])
        form.pack(fill="x", padx=0, pady=(4, 8))

        self.map_nature_var = tk.StringVar()
        self.map_code_var   = tk.StringVar()
        self.map_label_var  = tk.StringVar()

        fields_map = [
            ("Nature Lefevre *",    self.map_nature_var,
             "Ex : Usseau, Tuffeau, Haims..."),
            ("Code produit Odoo *", self.map_code_var,
             "Ex : TUF0000-PS  (code entre [crochets] dans Odoo)"),
            ("Libellé produit",     self.map_label_var,
             "Optionnel — nom complet du produit"),
        ]
        for label, var, hint in fields_map:
            f = tk.Frame(form, bg=C["section_bg"])
            f.pack(fill="x", padx=10, pady=4)
            tk.Label(f, text=label, bg=C["section_bg"], fg=C["text_dim"],
                     font=("Segoe UI", 9), anchor="w").pack(fill="x")
            ttk.Entry(f, textvariable=var).pack(fill="x", ipady=5)
            tk.Label(f, text=hint, bg=C["section_bg"], fg=C["text_dim"],
                     font=("Segoe UI", 8), anchor="w").pack(fill="x")

        btn_form = tk.Frame(right, bg=C["card"])
        btn_form.pack(fill="x", padx=0, pady=(0, 8))
        ttk.Button(btn_form, text="✅  Valider",
                   command=self._mapping_validate).pack(fill="x", padx=0, pady=2)
        ttk.Button(btn_form, text="✖  Annuler",
                   command=self._mapping_cancel).pack(fill="x", padx=0, pady=2)

        # Aide
        ttk.Separator(right, orient="horizontal").pack(fill="x", pady=8)
        self._section_title(right, "Comment trouver le code produit ?")
        help_txt = tk.Text(right, bg=C["section_bg"], fg=C["text_dim"],
                           font=("Segoe UI", 9), relief="flat", padx=10, pady=8,
                           wrap="word", height=10, state="normal")
        help_txt.insert("1.0",
            "1. Dans Odoo, aller sur un article (ex : Tuffeau Pré-sciées)\n\n"
            "2. Le code entre crochets dans le nom\n"
            "   [TUF0000-PS] = le code interne\n\n"
            "3. En mode développeur : passer la souris\n"
            "   sur le champ 'Référence interne'\n"
            "   → la valeur affichée = default_code\n\n"
            "4. Utiliser le bouton 'Diagnostiquer'\n"
            "   dans l'onglet Connexion Odoo pour\n"
            "   vérifier qu'un code est trouvable."
        )
        help_txt.configure(state="disabled")
        help_txt.pack(fill="x", padx=0)

        self._refresh_edit_mapping_tree()

    # ── Mapping : méthodes CRUD ────────────────────────────────────────
    def _refresh_edit_mapping_tree(self):
        if not hasattr(self, "edit_mapping_tree"):
            return
        self.edit_mapping_tree.delete(*self.edit_mapping_tree.get_children())
        nm = self.config.get("nature_mapping") or {}
        for nature, info in sorted(nm.items()):
            self.edit_mapping_tree.insert("", "end", iid=nature, values=(
                nature,
                info.get("product_code", ""),
                info.get("product", ""),
            ))

    def _on_mapping_select(self, event=None):
        sel = self.edit_mapping_tree.selection()
        if not sel:
            return
        nature = sel[0]
        nm = self.config.get("nature_mapping") or {}
        info = nm.get(nature, {})
        self.map_nature_var.set(nature)
        self.map_code_var.set(info.get("product_code", ""))
        self.map_label_var.set(info.get("product", ""))

    def _mapping_add(self):
        self.map_nature_var.set("")
        self.map_code_var.set("")
        self.map_label_var.set("")
        self.edit_mapping_tree.selection_remove(*self.edit_mapping_tree.selection())

    def _mapping_edit(self):
        sel = self.edit_mapping_tree.selection()
        if not sel:
            messagebox.showinfo("Sélection", "Sélectionnez une ligne à modifier.")
            return
        self._on_mapping_select()

    def _mapping_delete(self):
        sel = self.edit_mapping_tree.selection()
        if not sel:
            messagebox.showinfo("Sélection", "Sélectionnez une ligne à supprimer.")
            return
        nature = sel[0]
        if messagebox.askyesno("Confirmer", f"Supprimer '{nature}' du mapping ?"):
            nm = self.config.get("nature_mapping") or {}
            nm.pop(nature, None)
            self.config.set("nature_mapping", nm)
            self.config.save()
            self._refresh_edit_mapping_tree()
            self._refresh_mapping_tree()

    def _mapping_validate(self):
        nature = self.map_nature_var.get().strip()
        code   = self.map_code_var.get().strip()
        label  = self.map_label_var.get().strip()
        if not nature:
            messagebox.showwarning("Champ manquant", "La nature Lefevre est obligatoire.")
            return
        if not code:
            messagebox.showwarning("Champ manquant", "Le code produit Odoo est obligatoire.")
            return
        # Extraire code entre crochets si l'utilisateur colle le nom complet
        import re as _re
        m = _re.match(r'^\[([^\]]+)\]', code)
        if m:
            code = m.group(1)
        # Si label vide, construire depuis le code
        if not label:
            label = f"[{code}]"
        nm = self.config.get("nature_mapping") or {}
        nm[nature] = {"product_code": code, "product": label}
        self.config.set("nature_mapping", nm)
        self._refresh_edit_mapping_tree()
        self._refresh_mapping_tree()
        self.map_nature_var.set("")
        self.map_code_var.set("")
        self.map_label_var.set("")

    def _mapping_cancel(self):
        self.map_nature_var.set("")
        self.map_code_var.set("")
        self.map_label_var.set("")
        self.edit_mapping_tree.selection_remove(*self.edit_mapping_tree.selection())

    def _mapping_save(self):
        self.config.save()
        messagebox.showinfo("Sauvegardé",
            "Mapping enregistré dans :\n%APPDATA%\\MaquignonLefevre\\config.json")
        self.log("✓ Mapping sauvegardé", "ok")

    # ── Utilitaires UI ─────────────────────────────────────────────────
    def _section_title(self, parent, text):
        f = ttk.Frame(parent, style="Card.TFrame")
        f.pack(fill="x", padx=12, pady=(10, 4))
        tk.Label(f, text=text.upper(), bg=COLORS["card"], fg=COLORS["accent2"],
                 font=("Segoe UI", 8, "bold")).pack(side="left")

    def _field_row(self, parent, label, var, password=False):
        f = ttk.Frame(parent, style="Card.TFrame")
        f.pack(fill="x", padx=12, pady=2)
        ttk.Label(f, text=label, style="Dim.TLabel", width=24, anchor="w").pack(side="left")
        show = "*" if password else ""
        e = ttk.Entry(f, textvariable=var, show=show)
        e.pack(side="left", fill="x", expand=True, ipady=4)

    def _refresh_mapping_tree(self):
        self.mapping_tree.delete(*self.mapping_tree.get_children())
        nm = self.config.get("nature_mapping") or {}
        for nature, info in nm.items():
            self.mapping_tree.insert("", "end", values=(
                nature,
                info.get("product_code", ""),
            ))

    # ── Log ────────────────────────────────────────────────────────────
    def log(self, msg, tag=""):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{ts}] {msg}\n", tag)
        self.log_text.see("end")
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.delete("1.0", "end")

    def _set_status(self, msg, progress=None):
        self.status_var.set(msg)
        if progress is not None:
            self.progress_var.set(progress)
        self.root.update_idletasks()

    # ── Actions ────────────────────────────────────────────────────────
    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="Sélectionner un fichier commande Lefevre",
            filetypes=[("Fichiers Excel", "*.xlsx *.xlsm"), ("Tous les fichiers", "*.*")]
        )
        if not path:
            return
        self.selected_file.set(path)
        # Peupler le sélecteur de feuilles
        try:
            import pandas as pd
            xl = pd.ExcelFile(path)
            sheets = xl.sheet_names
            self.sheet_combo["values"] = sheets
            if sheets:
                self.sheet_var.set(sheets[0])
        except Exception:
            self.sheet_combo["values"] = []
            self.sheet_var.set("")

    def _parse_file(self):
        path = self.selected_file.get()
        if not path:
            messagebox.showwarning("Fichier manquant", "Veuillez sélectionner un fichier Lefevre.")
            return
        if not HAS_PANDAS:
            messagebox.showerror("Module manquant", "pandas est requis:\n\npip install pandas openpyxl")
            return

        self._set_status("Analyse en cours...", 20)
        self.log(f"Analyse : {os.path.basename(path)}", "info")

        try:
            nm = self.config.get("nature_mapping") or {}
            sheet_choice = self.sheet_var.get() or None
            self.header_info, self.parsed_rows = parse_lefevre(path, nm, sheet_name=sheet_choice)

            # Infos de détection automatique
            col_map    = self.header_info.get("col_map", {})
            unit       = self.header_info.get("unit", "cm")
            hdr_row    = self.header_info.get("header_row")
            sheet_used = self.header_info.get("sheet_used", "")
            col_info   = " | ".join(f"{k}→col{v}" for k, v in sorted(col_map.items()) if k in ("ref","nature","qte","long","haut","prof","cube","poids"))
            detect_msg = f"Feuille : {sheet_used}   Unité : {unit}   En-tête ligne : {hdr_row or 'auto'}\nColonnes détectées : {col_info}"
            self.detect_label.config(text=detect_msg)
            self.log(f"  Détection → feuille={sheet_used} | unité={unit} | {col_info}", "info")

            nb_sec = sum(1 for r in self.parsed_rows if r["type"] == "section")
            nb_sub = sum(1 for r in self.parsed_rows if r["type"] == "subsection")
            nb_lines = sum(1 for r in self.parsed_rows if r["type"] == "line")
            nb_unk = sum(1 for r in self.parsed_rows if r["type"] == "line" and not r["mapped"])
            cubage = self.header_info.get("cubage_total") or sum(
                (r.get("qty_m3") or 0) for r in self.parsed_rows if r["type"] == "line")

            # Mise à jour labels info
            self.info_labels["chantier"].config(text=self.header_info.get("chantier", "—")[:50])
            self.info_labels["lieu"].config(text=self.header_info.get("lieu_taille", "—"))
            self.info_labels["code"].config(text=self.header_info.get("code_chantier", "—"))
            self.info_labels["sections"].config(text=str(nb_sec))
            self.info_labels["subsections"].config(text=str(nb_sub))
            self.info_labels["lignes"].config(text=str(nb_lines))
            self.info_labels["cubage"].config(text=f"{cubage:.6f}")
            self.info_labels["blocs"].config(text=str(self.header_info.get("nb_blocs", nb_lines)))
            self.info_labels["inconnus"].config(
                text=str(nb_unk) if nb_unk == 0 else f"⚠ {nb_unk}",
                fg=COLORS["error"] if nb_unk > 0 else COLORS["success"])

            # Mise à jour aperçu
            self.preview_tree.delete(*self.preview_tree.get_children())
            for item in self.parsed_rows:
                if item["type"] == "section":
                    self.preview_tree.insert("", "end", values=(
                        "SECTION", item["label"], "", "", "", "", "", "", "", "", "", ""), tags=("section",))
                elif item["type"] == "subsection":
                    self.preview_tree.insert("", "end", values=(
                        "sous-sec", item["section"], item["label"], "", "", "", "", "", "", "", "", ""), tags=("subsection",))
                else:
                    tag = "" if item["mapped"] else "unmapped"
                    self.preview_tree.insert("", "end", values=(
                        "ligne",
                        item["section"], item["subsection"], item["ref"],
                        item["nature"], item["qte"],
                        item.get("long_m") or "", item.get("prof_m") or "",
                        item.get("haut_m") or "",
                        f'{item["qty_m3"]:.6f}' if item.get("qty_m3") else "",
                        item["product_code"],
                        "✓" if item["mapped"] else "❌",
                    ), tags=(tag,))

            self.log(f"✓ {nb_lines} lignes parsées | {nb_sec} sections | {nb_sub} sous-sections | {cubage:.4f} m³", "ok")
            if nb_unk:
                self.log(f"⚠ {nb_unk} lignes avec nature de pierre non mappée", "warn")
            self._set_status(f"Fichier analysé : {nb_lines} lignes, {cubage:.4f} m³", 100)
            self.nb.select(0)

        except Exception as e:
            self.log(f"❌ Erreur analyse : {e}", "err")
            self.log(traceback.format_exc(), "err")
            self._set_status("Erreur lors de l'analyse", 0)
            messagebox.showerror("Erreur", f"Impossible de lire le fichier :\n{e}")

    def _export_excel(self):
        if not self.parsed_rows:
            messagebox.showwarning("Aucune donnée", "Analysez d'abord un fichier Lefevre.")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("Module manquant", "openpyxl est requis:\n\npip install openpyxl")
            return

        base = os.path.splitext(os.path.basename(self.selected_file.get()))[0]
        ref = self.order_ref.get()
        default_name = f"Odoo_Import_{base}_{ref}.xlsx"
        output = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Odoo_Import",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Fichier Excel", "*.xlsx")]
        )
        if not output:
            return

        self._set_status("Export Excel en cours...", 50)
        try:
            vol, nb = generate_excel(
                self.header_info, self.parsed_rows,
                self.order_ref.get(), self.customer.get(),
                self.customer_code.get(), self.order_date.get(),
                output
            )
            self.log(f"✓ Export Excel : {os.path.basename(output)} ({nb} blocs, {vol:.4f} m³)", "ok")
            self._set_status(f"Export réussi : {os.path.basename(output)}", 100)
            messagebox.showinfo("Export réussi", f"Fichier créé :\n{output}\n\n{nb} blocs | {vol:.6f} m³")
        except Exception as e:
            self.log(f"❌ Erreur export : {e}", "err")
            messagebox.showerror("Erreur export", str(e))
            self._set_status("Erreur export", 0)

    def _import_odoo(self):
        if not self.parsed_rows:
            messagebox.showwarning("Aucune donnée", "Analysez d'abord un fichier Lefevre.")
            return

        nb_lines = sum(1 for r in self.parsed_rows if r["type"] == "line")
        if not messagebox.askyesno("Confirmer l'import",
            f"Importer {nb_lines} lignes de pierres dans Odoo ?\n\n"
            f"Client : {self.customer.get()} ({self.customer_code.get()})\n"
            f"Réf. : {self.order_ref.get()}\n"
            f"Chantier : {self.header_info.get('chantier', '?')}\n\n"
            f"Un devis (brouillon) sera créé dans Odoo."):
            return

        self.nb.select(3)  # Aller au journal
        threading.Thread(target=self._do_import_odoo, daemon=True).start()

    def _do_import_odoo(self):
        try:
            url = self.odoo_url.get()
            db = self.odoo_db.get()
            user = self.odoo_user.get()
            pwd = self.odoo_pass.get()
            if not all([url, db, user, pwd]):
                self.log("❌ Paramètres de connexion Odoo incomplets", "err")
                self._set_status("Connexion Odoo non configurée", 0)
                return

            self.log("Connexion à Odoo...", "info")
            self._set_status("Connexion Odoo...", 5)
            conn = OdooConnector(url, db, user, pwd)
            version = conn.connect()
            self.log(f"✓ Connecté — Odoo {version}", "ok")

            # Recherche partner
            self.log(f"Recherche client : {self.customer_code.get()} / {self.customer.get()}...", "info")
            partner = conn.search_partner(name=self.customer.get(), ref=self.customer_code.get())
            if not partner:
                self.log(f"⚠ Client non trouvé — création avec le nom '{self.customer.get()}'", "warn")
                partner_id = conn.call("res.partner", "create", [{
                    "name": self.customer.get(),
                    "ref": self.customer_code.get(),
                    "is_company": True,
                }])
            else:
                partner_id = partner["id"]
                self.log(f"✓ Client trouvé : [{partner_id}] {partner.get('name', '')}", "ok")

            # UoM m³
            uom_id = conn.get_uom_m3()

            # Création devis
            self.log(f"Création du devis {self.order_ref.get()}...", "info")
            order_id = conn.create_sale_order(
                partner_id,
                self.order_ref.get(),
                self.order_date.get(),
                self.header_info.get("chantier", "")
            )
            self.log(f"✓ Devis créé (ID: {order_id})", "ok")

            # Champs Studio
            fields_cfg = {k: v.get() for k, v in self.field_vars.items()}

            # ── Préchargement produits (1 seul appel API pour tous les codes) ──
            all_codes = list(set(
                r["product_code"] for r in self.parsed_rows
                if r["type"] == "line" and not r["product_code"].startswith("???")
            ))
            self.log(f"Préchargement produits : {all_codes}...", "info")
            self._set_status("Recherche des produits dans Odoo...", 15)
            found, total_codes = conn.preload_products(all_codes)
            self.log(f"✓ Produits trouvés : {found}/{total_codes}", "ok" if found == total_codes else "warn")

            # Diagnostic des produits introuvables
            not_found = [c for c in all_codes if conn._product_cache.get(c) is None]
            if not_found:
                self.log(f"⚠ Produits non trouvés dans Odoo : {not_found}", "warn")
                self.log("  → Lancement du diagnostic (recherche étendue + archivés)...", "warn")
                for code in not_found:
                    diag = conn.diagnose_product(code)
                    for line in diag:
                        self.log(f"  DIAG: {line}", "warn")
                self.log("  → Vérifiez le code interne du produit dans Odoo (mode dev > fiche article)", "warn")

            # Création lignes
            total = len(self.parsed_rows)
            ok_count = 0
            warn_count = 0
            err_count = 0

            for i, item in enumerate(self.parsed_rows):
                progress = 20 + int((i / total) * 75)
                self._set_status(f"Import ligne {i+1}/{total}...", progress)

                if item["type"] in ("section", "subsection"):
                    lbl = item["label"] if item["type"] == "section" else f"  {item['label']}"
                    conn.create_order_line_section(order_id, lbl)
                    ok_count += 1
                elif item["type"] == "line":
                    line_id, err = conn.create_order_line_product(order_id, item, fields_cfg, uom_id)
                    if err and line_id:
                        self.log(f"  {err}", "warn")
                        warn_count += 1
                        ok_count += 1
                    elif err:
                        # Log uniquement le 1er occurrence par produit pour ne pas spammer
                        code = item.get("product_code", "")
                        err_key = f"_err_logged_{code}"
                        if not getattr(self, err_key, False):
                            self.log(f"  {err}", "err")
                            setattr(self, err_key, True)
                        err_count += 1
                    else:
                        ok_count += 1

            # Résumé
            order_url = f"{url.rstrip('/')}/web#action=sale.action_quotations_with_onboarding&id={order_id}&model=sale.order&view_type=form"
            self.log("", "")
            self.log(f"══ IMPORT TERMINÉ ══════════════════════════", "ok")
            self.log(f"✓ Devis Odoo ID : {order_id}", "ok")
            self.log(f"✓ Lignes créées : {ok_count}", "ok")
            if warn_count: self.log(f"⚠ Avertissements : {warn_count}", "warn")
            if err_count:  self.log(f"❌ Erreurs : {err_count}", "err")
            self.log(f"🔗 Ouvrir dans Odoo : {order_url}", "info")
            self._set_status(f"Import terminé — Devis #{order_id} créé dans Odoo", 100)

        except Exception as e:
            self.log(f"❌ Erreur import : {e}", "err")
            self.log(traceback.format_exc(), "err")
            self._set_status("Erreur lors de l'import", 0)

    def _test_connection(self):
        self._save_config()
        self.conn_status.config(text="Connexion en cours...", fg=COLORS["text_dim"])
        threading.Thread(target=self._do_test_connection, daemon=True).start()

    def _do_test_connection(self):
        try:
            conn = OdooConnector(
                self.odoo_url.get(), self.odoo_db.get(),
                self.odoo_user.get(), self.odoo_pass.get()
            )
            version = conn.connect()
            self.conn_status.config(text=f"✓ Connecté — Odoo {version}", fg=COLORS["success"])
            self.log(f"✓ Connexion Odoo réussie — version {version}", "ok")
            self._set_status(f"Connexion Odoo OK — v{version}", 100)
        except Exception as e:
            self.conn_status.config(text=f"❌ {e}", fg=COLORS["error"])
            self.log(f"❌ Connexion Odoo échouée : {e}", "err")
            self._set_status("Connexion Odoo échouée", 0)

    def _diagnose_product(self):
        code = self.diag_code_var.get().strip()
        if not code:
            messagebox.showwarning("Code manquant", "Saisissez un code produit à diagnostiquer.")
            return
        self.nb.select(3)
        threading.Thread(target=self._do_diagnose_product, args=(code,), daemon=True).start()

    def _do_diagnose_product(self, code):
        try:
            conn = OdooConnector(
                self.odoo_url.get(), self.odoo_db.get(),
                self.odoo_user.get(), self.odoo_pass.get()
            )
            conn.connect()
            self.log(f"══ DIAGNOSTIC PRODUIT : '{code}' ══════════════", "info")
            diag_lines = conn.diagnose_product(code)
            for line in diag_lines:
                tag = "ok" if "résultat(s)" not in line and "→" in line else ("warn" if "0 résultat" in line else "info")
                self.log(line, tag)
            self.log("──────────────────────────────────────────────", "info")
            self.log("Si le produit est archivé : Odoo > Articles > Activer les archivés > Désarchiver", "warn")
            self.log("Si code différent : modifier le mapping dans l'onglet Fichier & Import", "warn")
        except Exception as e:
            self.log(f"❌ Erreur diagnostic : {e}", "err")

    def _check_studio_fields(self):
        fields_cfg = {k: v.get() for k, v in self.field_vars.items()}
        threading.Thread(target=self._do_check_fields, args=(fields_cfg,), daemon=True).start()

    def _do_check_fields(self, fields_cfg):
        try:
            conn = OdooConnector(
                self.odoo_url.get(), self.odoo_db.get(),
                self.odoo_user.get(), self.odoo_pass.get()
            )
            conn.connect()
            results = conn.check_studio_fields(fields_cfg)
            self.log("Vérification des champs Studio :", "info")
            for key, exists in results.items():
                fname = fields_cfg[key]
                if exists:
                    self.log(f"  ✓ {key} → {fname}", "ok")
                else:
                    self.log(f"  ❌ {key} → {fname} (NON TROUVÉ dans Odoo)", "err")
        except Exception as e:
            self.log(f"❌ Vérification échouée : {e}", "err")

    def _save_config(self):
        self.config.set("odoo", "url", self.odoo_url.get())
        self.config.set("odoo", "database", self.odoo_db.get())
        self.config.set("odoo", "username", self.odoo_user.get())
        self.config.set("odoo", "password", self.odoo_pass.get())
        self.config.set("commande", "customer", self.customer.get())
        self.config.set("commande", "customer_code", self.customer_code.get())
        for key, var in self.field_vars.items():
            self.config.set("fields", key, var.get())
        self.config.save()
        self.log("✓ Configuration enregistrée", "ok")
        self.conn_status.config(text="Configuration sauvegardée ✓", fg=COLORS["success"])


# ══════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE
# ══════════════════════════════════════════════════════════════════════

def check_dependencies():
    missing = []
    if not HAS_PANDAS:
        missing.append("pandas")
    if not HAS_OPENPYXL:
        missing.append("openpyxl")
    return missing

def main():
    root = tk.Tk()
    root.withdraw()

    missing = check_dependencies()
    if missing:
        import subprocess
        msg = f"Modules manquants : {', '.join(missing)}\n\nInstaller maintenant ?"
        if messagebox.askyesno("Installation requise", msg):
            for mod in missing:
                subprocess.call([sys.executable, "-m", "pip", "install", mod])
            messagebox.showinfo("OK", "Modules installés. Relancez l'application.")
            return

    root.deiconify()
    app = LefevreMaquignonApp(root)

    try:
        root.iconbitmap(default="")
    except:
        pass

    root.mainloop()

if __name__ == "__main__":
    main()
