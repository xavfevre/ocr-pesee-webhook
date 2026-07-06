# -*- coding: utf-8 -*-
"""Livre de police décharge (ISDI) + DAP — CARRIÈRE D'HAIMS / MAQUIGNON.

Usage : python generate_livre_police.py <Factures_MM.YYYY.xlsx>

Lit le relevé mensuel du pont bascule :
  - Feuil2 = pesées RECEPTION taguées (colonne Code = TP86170504…, colonne DAP)
  - Feuil1 = blocs par client avec n° de facture (croisement pesée → facture)

Produit :
  - Livre_de_police_<mois>.xlsx : registre réglementaire + DAP à créer + contrôle
  - DAP_a_signer_<mois>.pdf     : formulaires DAP pré-remplis (un par DAP manquante)
"""
import sys, re, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = sys.argv[1] if len(sys.argv) > 1 else "/root/.claude/uploads/c94bd159-9526-5b32-9229-6ed47d7c55e8/50007919-Factures_03.2026.xlsx"
OUT_DIR = "/tmp/claude-0/-home-user/c94bd159-9526-5b32-9229-6ed47d7c55e8/scratchpad"

SOCIETE = "CARRIÈRE D'HAIMS — SARL MAQUIGNON FRÈRES"
SITE = "Site de remblaiement (décharge inertes) — Haims"

# ── Référentiel codes (grille d'identification du PDF) ──────────────
CLIENT_TYPES = ["TP", "BAT", "COL", "PART"]
DEPTS = ["86", "37"]
DECHETS = {
    "170101": "Bétons",
    "170102": "Briques",
    "170103": "Tuiles et céramiques",
    "170107": "Mélange de béton, briques, tuiles et céramiques",
    "170302": "Mélanges bitumineux (sans goudron)",
    "170504": "Terres et pierres (y compris déblais)",
}
ALL_CODES = [f"{t}{d}{w}" for t in CLIENT_TYPES for d in DEPTS for w in DECHETS]

def euro_code(waste6):
    return f"{waste6[0:2]} {waste6[2:4]} {waste6[4:6]}"

def parse_code(code):
    m = re.match(r"(TP|BAT|COL|PART)(86|37)(\d{6})", str(code or "").replace(" ", "").upper())
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3)

# ── Lecture du classeur source ──────────────────────────────────────
wb = openpyxl.load_workbook(SRC, data_only=True)
f1, f2 = wb["Feuil1"], wb["Feuil2"]

# pesée -> n° facture (Feuil1 : la facture en col A vaut pour le bloc)
fact_by_pesee = {}
cur_fact = None
for r in range(1, f1.max_row + 1):
    a = f1.cell(r, 1).value
    b = f1.cell(r, 2).value
    c = f1.cell(r, 3).value
    if a and str(a).strip():
        cur_fact = str(a).strip()
        if not str(cur_fact).upper().startswith("FA"):
            cur_fact = None  # bloc sans facture (nom client seul)
    if isinstance(b, (int, float)) and c in ("RECEPTION", "EXPEDITION"):
        fact_by_pesee[int(b)] = cur_fact

# pesées Feuil2 (le livre de police)
rows = []
for r in range(2, f2.max_row + 1):
    pesee = f2.cell(r, 1).value
    if not pesee:
        continue
    d = f2.cell(r, 3).value
    rows.append({
        "pesee": int(pesee),
        "date": d,
        "net_kg": f2.cell(r, 6).value or 0,
        "vehicule": str(f2.cell(r, 7).value or "").upper(),
        "transporteur": str(f2.cell(r, 8).value or ""),
        "client": str(f2.cell(r, 9).value or "").strip(),
        "produit": str(f2.cell(r, 10).value or ""),
        "lieu": str(f2.cell(r, 11).value or "").strip(),
        "code": str(f2.cell(r, 12).value or "").replace(" ", "").upper(),
        "dap": f2.cell(r, 13).value,
    })
rows.sort(key=lambda x: (x["date"] or datetime.datetime.min, x["pesee"]))
month_label = rows[0]["date"].strftime("%Y-%m") if rows and rows[0]["date"] else "mois"

# ── DAP : existantes et à créer ─────────────────────────────────────
def dap_key(x):
    return (x["client"].upper(), x["lieu"].upper(), x["code"])

dap_known = {}
max_dap = 0
for x in rows:
    if x["dap"] not in (None, ""):
        dap_known.setdefault(dap_key(x), x["dap"])
        try:
            max_dap = max(max_dap, int(x["dap"]))
        except Exception:
            pass

new_daps = {}
next_num = max_dap + 1
for x in rows:
    k = dap_key(x)
    if x["dap"] in (None, "") and k not in dap_known and k not in new_daps:
        new_daps[k] = next_num
        next_num += 1

def dap_for(x):
    if x["dap"] not in (None, ""):
        return x["dap"], ""
    k = dap_key(x)
    if k in dap_known:
        return dap_known[k], "reprise"
    return new_daps[k], "À CRÉER"

# ── Classeur de sortie ──────────────────────────────────────────────
out = openpyxl.Workbook()
TEAL = "01666B"
th = Font(bold=True, color="FFFFFF", size=10)
fillh = PatternFill("solid", fgColor=TEAL)
thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)
warn = PatternFill("solid", fgColor="FFF3CD")

# — Feuille 1 : Livre de police —
ws = out.active
ws.title = "Livre de police"
ws["A1"] = f"LIVRE DE POLICE — REGISTRE DES DÉCHETS ENTRANTS (remblaiement) — {month_label}"
ws["A1"].font = Font(bold=True, size=14, color=TEAL)
ws["A2"] = f"{SOCIETE} — {SITE}"
ws["A2"].font = Font(italic=True, size=10)
HEAD = ["N° ordre", "Date réception", "N° pesée", "Code déchet", "Désignation du déchet",
        "Code interne", "Tonnage (t)", "Producteur / détenteur", "Origine (chantier)",
        "Transporteur", "Immatriculation", "N° DAP", "N° facture", "Contrôle visuel / obs."]
ws.append([])
ws.append(HEAD)
hr = ws.max_row
for c in range(1, len(HEAD) + 1):
    cell = ws.cell(hr, c)
    cell.font = th; cell.fill = fillh; cell.border = thin
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for i, x in enumerate(rows, 1):
    p = parse_code(x["code"])
    dap, dstat = dap_for(x)
    ws.append([
        i, x["date"].strftime("%d/%m/%Y %H:%M") if x["date"] else "", x["pesee"],
        euro_code(p[2]) if p else "?", DECHETS.get(p[2], x["produit"]) if p else x["produit"],
        x["code"], round((x["net_kg"] or 0) / 1000.0, 3), x["client"], x["lieu"],
        x["transporteur"], x["vehicule"], dap, fact_by_pesee.get(x["pesee"], "") or "",
        "Conforme" if not dstat else ("DAP " + dstat),
    ])
    if dstat == "À CRÉER":
        for c in range(1, len(HEAD) + 1):
            ws.cell(ws.max_row, c).fill = warn
    for c in range(1, len(HEAD) + 1):
        ws.cell(ws.max_row, c).border = thin
        ws.cell(ws.max_row, c).font = Font(size=9)
tot = sum((x["net_kg"] or 0) for x in rows) / 1000.0
ws.append([])
ws.append(["", "", "", "", "", "TOTAL (t)", round(tot, 3)])
ws.cell(ws.max_row, 6).font = Font(bold=True)
ws.cell(ws.max_row, 7).font = Font(bold=True)
widths = [7, 15, 9, 10, 34, 13, 10, 22, 18, 16, 14, 8, 11, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = f"A{hr+1}"
# menu déroulant sur Code interne (colonne F)
dv = DataValidation(type="list", formula1='"' + ",".join(ALL_CODES) + '"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"F{hr+1}:F{hr+len(rows)}")

# — Feuille 2 : DAP à créer —
w2 = out.create_sheet("DAP à créer")
w2["A1"] = "DAP MANQUANTES — à faire signer au producteur (formulaires pré-remplis dans le PDF joint)"
w2["A1"].font = Font(bold=True, size=12, color="B45309")
w2.append([])
H2 = ["N° DAP proposé", "Producteur", "Chantier / origine", "Code interne", "Code déchet",
      "Désignation", "Nb pesées", "Tonnage (t)", "Période"]
w2.append(H2)
for c in range(1, len(H2) + 1):
    w2.cell(3, c).font = th; w2.cell(3, c).fill = fillh
for k, num in sorted(new_daps.items(), key=lambda a: a[1]):
    grp = [x for x in rows if dap_key(x) == k]
    p = parse_code(k[2])
    dates = [x["date"] for x in grp if x["date"]]
    w2.append([num, grp[0]["client"], grp[0]["lieu"], k[2],
               euro_code(p[2]) if p else "?", DECHETS.get(p[2], "") if p else "",
               len(grp), round(sum(x["net_kg"] for x in grp) / 1000.0, 3),
               (min(dates).strftime("%d/%m") + " → " + max(dates).strftime("%d/%m/%Y")) if dates else ""])
for i, w in enumerate([14, 24, 20, 13, 10, 32, 9, 11, 20], 1):
    w2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# — Feuille 3 : Contrôle (réceptions Feuil1 absentes de Feuil2) —
w3 = out.create_sheet("Contrôle")
w3["A1"] = "RÉCEPTIONS du relevé (Feuil1) ABSENTES du livre (Feuil2) — vérifier si à intégrer"
w3["A1"].font = Font(bold=True, size=11, color="B45309")
w3.append([])
w3.append(["N° pesée", "Date", "Net (kg)", "Client", "Produit", "Lieu", "N° facture"])
seen = {x["pesee"] for x in rows}
nmiss = 0
for r in range(1, f1.max_row + 1):
    b = f1.cell(r, 2).value; c = f1.cell(r, 3).value
    if isinstance(b, (int, float)) and c == "RECEPTION" and int(b) not in seen:
        d = f1.cell(r, 4).value
        w3.append([int(b), d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d or ""),
                   f1.cell(r, 7).value, f1.cell(r, 10).value, f1.cell(r, 11).value,
                   f1.cell(r, 12).value, fact_by_pesee.get(int(b), "")])
        nmiss += 1
for i, w in enumerate([9, 12, 10, 22, 34, 18, 11], 1):
    w3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# — Feuille 4 : Référentiel codes —
w4 = out.create_sheet("Référentiel codes")
w4.append(["Code interne", "Type client", "Dépt", "Code déchet", "Désignation"])
for c in range(1, 6):
    w4.cell(1, c).font = th; w4.cell(1, c).fill = fillh
for code in ALL_CODES:
    t, d, wst = parse_code(code)
    w4.append([code, {"TP": "Entreprise TP", "BAT": "Entreprise bâtiment", "COL": "Collectivité", "PART": "Particulier"}[t],
               d, euro_code(wst), DECHETS[wst]])
for i, w in enumerate([14, 20, 6, 10, 40], 1):
    w4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

xlsx_path = f"{OUT_DIR}/Livre_de_police_{month_label}.xlsx"
out.save(xlsx_path)
print("XLSX:", xlsx_path)
print(f"  {len(rows)} entrées, {round(tot,1)} t | DAP à créer: {len(new_daps)} | réceptions hors livre: {nmiss}")

# ── PDF : formulaires DAP pré-remplis ───────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as C
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

TEALC = C.HexColor("#01666B")
sT = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=15, alignment=1, textColor=C.white)
sS = ParagraphStyle("s", fontName="Helvetica", fontSize=9, textColor=C.HexColor("#555555"))
sL = ParagraphStyle("l", fontName="Helvetica-Bold", fontSize=10)
sV = ParagraphStyle("v", fontName="Helvetica", fontSize=10)

def dap_page(story, num, client, lieu, code, waste6, grp):
    t = Table([[Paragraph("DEMANDE D'ACCEPTATION PRÉALABLE (DAP)<br/>Déchets inertes — remblaiement", sT)]], colWidths=[17*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), TEALC), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10)]))
    story.append(t); story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"<b>DAP n° {num}</b> — {SOCIETE}", sL))
    story.append(Paragraph(SITE, sS)); story.append(Spacer(1, 0.35*cm))
    tons = round(sum(x['net_kg'] for x in grp)/1000.0, 2)
    dates = [x['date'] for x in grp if x['date']]
    per = (min(dates).strftime('%d/%m/%Y') + " au " + max(dates).strftime('%d/%m/%Y')) if dates else ""
    body = [
        ["Producteur / détenteur du déchet", client],
        ["Adresse du producteur", ""],
        ["Chantier d'origine", lieu],
        ["Code déchet (déchet inerte)", f"{euro_code(waste6)} — {DECHETS.get(waste6,'')}"],
        ["Code interne site", code],
        ["Quantité livrée / estimée", f"{tons} t  ({len(grp)} livraison(s), {per})"],
        ["Transporteur(s)", ", ".join(sorted({x['transporteur'] for x in grp if x['transporteur']})) or ""],
        ["Immatriculation(s)", ", ".join(sorted({x['vehicule'] for x in grp if x['vehicule']}))],
    ]
    tb = Table([[Paragraph(a, sL), Paragraph(b, sV)] for a, b in body], colWidths=[6.5*cm, 10.5*cm])
    tb.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, C.HexColor("#BBBBBB")),
                            ("BACKGROUND", (0, 0), (0, -1), C.HexColor("#EAF4F4")),
                            ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                            ("LEFTPADDING", (0, 0), (-1, -1), 8)]))
    story.append(tb); story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("Le producteur atteste que les déchets sont <b>inertes</b>, non contaminés (pas d'amiante, "
                           "de goudron, de plâtre, de bois, de plastiques, de terre végétale polluée…), et qu'ils "
                           "proviennent exclusivement du chantier indiqué ci-dessus.", sV))
    story.append(Spacer(1, 0.7*cm))
    sig = Table([[Paragraph("Date :", sL), Paragraph("Signature et cachet du producteur :", sL),
                  Paragraph("Visa exploitant (Maquignon) :", sL)], ["", "", ""]],
                colWidths=[4*cm, 7*cm, 6*cm], rowHeights=[0.7*cm, 2.6*cm])
    sig.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, C.HexColor("#BBBBBB"))]))
    story.append(sig)
    story.append(PageBreak())

pdf_path = f"{OUT_DIR}/DAP_a_signer_{month_label}.pdf"
doc = SimpleDocTemplate(pdf_path, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=1.6*cm, bottomMargin=1.6*cm)
story = []
for k, num in sorted(new_daps.items(), key=lambda a: a[1]):
    grp = [x for x in rows if dap_key(x) == k]
    p = parse_code(k[2])
    dap_page(story, num, grp[0]["client"], grp[0]["lieu"], k[2], p[2] if p else "170504", grp)
if story:
    doc.build(story)
    print("PDF :", pdf_path, f"({len(new_daps)} formulaires)")
else:
    print("PDF : aucune DAP manquante")
