# -*- coding: utf-8 -*-
import json, urllib.request, ssl, http.cookiejar, datetime
import sys; sys.path.insert(0, '/tmp/venv/lib/python3.11/site-packages')
import openpyxl
URL="https://maquignon.odoo.com"; DB="maquignon"; U="isabelle@maquignon.com"; P="Isamaq71*"
ctx=ssl.create_default_context(); cj=http.cookiejar.CookieJar()
op=urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj), urllib.request.HTTPSHandler(context=ctx))
def call(p,pl):
    r=urllib.request.Request(URL+p,data=json.dumps(pl).encode(),headers={"Content-Type":"application/json"})
    return json.loads(op.open(r,timeout=180).read())
def kw(m,meth,args,kwargs=None):
    r=call("/web/dataset/call_kw",{"jsonrpc":"2.0","method":"call","params":{"model":m,"method":meth,"args":args,"kwargs":kwargs or {}}})
    if r.get("error"): raise Exception(str(r["error"].get("data",{}).get("message",r["error"]))[:300])
    return r["result"]
call("/web/session/authenticate",{"jsonrpc":"2.0","params":{"db":DB,"login":U,"password":P}})

# éviter double import
n_exist = kw("x_livre_police","search_count",[[]])
if n_exist:
    print("déjà", n_exist, "entrées — abandon"); raise SystemExit

wb = openpyxl.load_workbook('/root/.claude/uploads/c94bd159-9526-5b32-9229-6ed47d7c55e8/50007919-Factures_03.2026.xlsx', data_only=True)
f1, f2 = wb["Feuil1"], wb["Feuil2"]
fact_by_pesee={}; cur=None
for r in range(1, f1.max_row+1):
    a=f1.cell(r,1).value; b=f1.cell(r,2).value; c=f1.cell(r,3).value
    if a and str(a).strip():
        cur=str(a).strip()
        if not cur.upper().startswith("FA"): cur=None
    if isinstance(b,(int,float)) and c in ("RECEPTION","EXPEDITION"): fact_by_pesee[int(b)]=cur

rows=[]
for r in range(2, f2.max_row+1):
    if not f2.cell(r,1).value: continue
    rows.append(dict(pesee=int(f2.cell(r,1).value), date=f2.cell(r,3).value,
        p1=f2.cell(r,4).value or 0, p2=f2.cell(r,5).value or 0, net=f2.cell(r,6).value or 0,
        veh=str(f2.cell(r,7).value or "").upper(), tra=str(f2.cell(r,8).value or ""),
        cli=str(f2.cell(r,9).value or "").strip(), pro=str(f2.cell(r,10).value or ""),
        lieu=str(f2.cell(r,11).value or "").strip(),
        code=str(f2.cell(r,12).value or "").replace(" ","").upper(), dap=f2.cell(r,13).value))

# DAP : existantes (numérotées) + nouvelles pour les groupes sans n°
def key(x): return (x["cli"].upper(), x["lieu"].upper(), x["code"])
dapnum={}; mx=0
for x in rows:
    if x["dap"] not in (None,""):
        dapnum.setdefault(key(x), int(x["dap"])); mx=max(mx,int(x["dap"]))
nxt=mx+1
for x in rows:
    if key(x) not in dapnum:
        dapnum[key(x)]=nxt; nxt+=1

dap_ids={}
for k,num in sorted(dapnum.items(), key=lambda a:a[1]):
    grp=[x for x in rows if key(x)==k]
    dates=[x["date"] for x in grp if x["date"]]
    did = kw("x_dap","create",[{ "x_name":str(num), "x_studio_client":grp[0]["cli"],
        "x_studio_chantier":grp[0]["lieu"], "x_studio_code":(k[2] or False),
        "x_studio_date": min(dates).strftime("%Y-%m-%d") if dates else False,
        "x_studio_tonnage_estime": round(sum(x["net"] for x in grp)/1000.0,3) }])
    dap_ids[k]=did
print("DAP créées:", len(dap_ids))

vals_list=[]
for x in rows:
    vals_list.append({ "x_name":str(x["pesee"]),
        "x_studio_date": x["date"].strftime("%Y-%m-%d %H:%M:%S") if x["date"] else False,
        "x_studio_poids1":x["p1"], "x_studio_poids2":x["p2"], "x_studio_net":x["net"],
        "x_studio_vehicule":x["veh"], "x_studio_transporteur":x["tra"],
        "x_studio_client":x["cli"], "x_studio_produit":x["pro"], "x_studio_chantier":x["lieu"],
        "x_studio_code": x["code"] or False, "x_studio_dap_id": dap_ids[key(x)],
        "x_studio_facture": fact_by_pesee.get(x["pesee"]) or "", "x_studio_controle":"conforme" })
ids = kw("x_livre_police","create",[vals_list])
print("entrées créées:", len(ids) if isinstance(ids,list) else ids)
