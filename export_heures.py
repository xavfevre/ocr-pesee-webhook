# -*- coding: utf-8 -*-
"""Export paie des heures salariés — un classeur Excel, un onglet par salarié.

Format inspiré du fichier de travail de la comptable (« Feuille de temps
hebdomadaire ») : blocs par semaine avec Arrivée/Départ matin et après-midi,
heures effectuées, écart vs horaire contractuel, mentions CP / MALADIE /
FERIE / ABSENCE / RECUP dans les cases, totaux hebdo + récap mensuel.

Les données viennent du modèle Odoo `x_heures_jour` (saisie web /mes-heures
et /heures-admin).
"""
import io
import re
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TYPES = {
    'cp': 'CP', 'maladie': 'MALADIE', 'ferie': 'FERIE',
    'absence': 'ABSENT', 'recup': 'RECUP', 'repos': '',
}
JOURS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']


def _fmt_h(h):
    """7.5 -> '07:30' ; 0 -> ''"""
    if not h:
        return ''
    return '%02d:%02d' % (int(h), round((h - int(h)) * 60))


def _theo_day(cal, d):
    """Heures théoriques + plages types pour un jour selon le calendrier."""
    if not cal:
        return 0.0
    wt = str(((d - date(1970, 1, 5)).days // 7) % 2)
    tot = 0.0
    for a in cal['attendance_ids']:
        if a['dayofweek'] != str(d.weekday()) or a.get('display_type'):
            continue
        if cal['two_weeks_calendar'] and a.get('week_type') and a['week_type'] != wt:
            continue
        tot += a['hour_to'] - a['hour_from']
    return tot


def build(call, mois, comp='all', du=None, au=None, exc=None):
    """Génère le classeur ; `call(model, method, *args, **kw)` = execute_kw.
    mois = 'YYYY-MM' ; comp = 'all' ou id de société."""
    if du and au:
        d1 = datetime.strptime(du, '%Y-%m-%d').date()
        d2 = datetime.strptime(au, '%Y-%m-%d').date()
    else:
        y, m = int(mois[:4]), int(mois[5:7])
        d1 = date(y, m, 1)
        d2 = (date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)) - timedelta(days=1)
    y, m = d1.year, d1.month
    # exceptions par salarie : {id: (du, au)} — periode propre a l'onglet
    exc = {str(k): v for k, v in (exc or {}).items()}
    exc_dates = [datetime.strptime(x, '%Y-%m-%d').date() for v in exc.values() for x in v]
    bmin = min([d1] + exc_dates)
    bmax = max([d2] + exc_dates)
    # semaines couvrant le mois (lundi -> dimanche)
    monday0 = d1 - timedelta(days=d1.weekday())
    dom = [('active', '=', True)]
    if comp != 'all':
        dom.append(('company_id', '=', int(comp)))
    emps = call('hr.employee', 'search_read', dom,
                fields=['name', 'company_id', 'resource_calendar_id', 'x_matricule_paie',
                        'x_recup_solde', 'x_cp_ref_date', 'x_contrat_mensuel'],
                order='company_id, name')
    cal_ids = sorted(set(e['resource_calendar_id'][0] for e in emps if e['resource_calendar_id']))
    cals = {c['id']: c for c in call('resource.calendar', 'read', cal_ids,
            fields=['name', 'two_weeks_calendar', 'attendance_ids'])} if cal_ids else {}
    att_ids = sorted(set(a for c in cals.values() for a in c['attendance_ids']))
    atts = {a['id']: a for a in call('resource.calendar.attendance', 'read', att_ids,
            fields=['dayofweek', 'hour_from', 'hour_to', 'day_period', 'week_type', 'display_type'])} if att_ids else {}
    for c in cals.values():
        c['attendance_ids'] = [atts[a] for a in c['attendance_ids'] if a in atts]
    rows = call('x_heures_jour', 'search_read',
                [('x_employee_id', 'in', [e['id'] for e in emps]),
                 ('x_date', '>=', (bmin - timedelta(days=bmin.weekday())).strftime('%Y-%m-%d')),
                 ('x_date', '<=', (bmax + timedelta(days=6)).strftime('%Y-%m-%d'))],
                fields=['x_employee_id', 'x_date', 'x_type', 'x_m_deb', 'x_m_fin',
                        'x_am_deb', 'x_am_fin', 'x_heures', 'x_theo', 'x_hs', 'x_note', 'x_decouchage'])
    by_emp = {}
    for r in rows:
        by_emp.setdefault(r['x_employee_id'][0], {})[r['x_date']] = r
    # récup : heures mises (x_recup_ligne) et jours/demi-jours récupérés depuis le
    # début de période (01/06), pour le solde « arrêté bureau + mises − récupérées »
    rl_by_emp = {}
    for l in call('x_recup_ligne', 'search_read',
                  [('x_employee_id', 'in', [e['id'] for e in emps])],
                  fields=['x_employee_id', 'x_date', 'x_heures']):
        rl_by_emp.setdefault(l['x_employee_id'][0], []).append(l)
    pstart = date(y if m >= 6 else y - 1, 6, 1)
    pris_by_emp = {}
    for r in call('x_heures_jour', 'search_read',
                  ['&', ('x_employee_id', 'in', [e['id'] for e in emps]),
                   ('x_date', '>=', pstart.strftime('%Y-%m-%d')),
                   '|', ('x_type', '=', 'recup'), ('x_note', 'like', 'Récupération —')],
                  fields=['x_employee_id', 'x_date', 'x_type', 'x_theo', 'x_note']):
        pris_by_emp.setdefault(r['x_employee_id'][0], []).append(r)

    def _recup_pris_h(r, cal):
        """Heures récupérées portées par un jour : jour entier -> théo figé ;
        demi-jour (type travail, note Récupération) -> partie non travaillée."""
        if r['x_type'] == 'recup':
            return r['x_theo'] or 0.0
        d = datetime.strptime(r['x_date'], '%Y-%m-%d').date()
        return max(_theo_day(cal, d) - (r['x_theo'] or 0.0), 0.0)

    wb = Workbook()
    wb.remove(wb.active)
    F10 = Font(name='Arial', size=10)
    FB = Font(name='Arial', size=10, bold=True)
    FT = Font(name='Arial', size=13, bold=True)
    HDR = Font(name='Arial', size=9.5, bold=True, color='FFFFFF')
    FILL = PatternFill('solid', start_color='1F4E5F')
    WFILL = PatternFill('solid', start_color='E8EEF1')
    thin = Border(bottom=Side(style='thin', color='CCCCCC'))
    CTR = Alignment(horizontal='center')

    for e in emps:
        title = re.sub(r'[\[\]:*?/\\]', ' ', e['name'])[:31]
        ws = wb.create_sheet(title)
        cal = cals.get(e['resource_calendar_id'][0]) if e['resource_calendar_id'] else None
        saisies = by_emp.get(e['id'], {})
        mat = e.get('x_matricule_paie') or ''
        exc_e = exc.get(str(e['id']))
        if exc_e:
            e1 = datetime.strptime(exc_e[0], '%Y-%m-%d').date()
            e2 = datetime.strptime(exc_e[1], '%Y-%m-%d').date()
        else:
            e1, e2 = d1, d2
        emonday0 = e1 - timedelta(days=e1.weekday())
        ws['A1'] = f"HEURES — {e['name']}" + (f" — matricule {mat}" if mat else '')
        ws['A1'].font = FT
        ws['A2'] = f"{e['company_id'][1]} · {cal['name'] if cal else 'sans horaire'} · {e1.strftime('%d/%m/%Y')} → {e2.strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        # récap mensuel (calculé sur les jours DU mois uniquement)
        tot_h = tot_theo = 0.0
        n_cp = n_mal = n_abs = n_fer = n_rec = n_dec = 0
        d = e1
        while d <= e2:
            s = saisies.get(d.strftime('%Y-%m-%d'))
            # théorique de l'écart : jour travail = théo figé (demi-journées gérées),
            # jour vide = calendrier ; les jours posés en absence ne comptent pas
            if s and s['x_type'] == 'travail':
                tot_theo += s['x_theo']
            elif not s:
                tot_theo += _theo_day(cal, d)
            if s:
                t = s['x_type']
                if t == 'travail':
                    tot_h += s['x_heures']
                elif t == 'cp':
                    n_cp += 1
                elif t == 'maladie':
                    n_mal += 1
                elif t == 'absence':
                    n_abs += 1
                elif t == 'ferie':
                    n_fer += 1
                elif t == 'recup':
                    n_rec += 1
                if s.get('x_decouchage'):
                    n_dec += 1
            d += timedelta(days=1)
        # horaire mensuel contractuel = hebdo (moyenne A/B si cycle) x 52/12 ;
        # base légale française 151,67 h/mois (35 h), l'écart = h. sup structurelles
        BASE_LEGALE = 35 * 52 / 12
        heb_a = heb_b = 0.0
        if cal:
            for a in cal['attendance_ids']:
                if a.get('display_type'):
                    continue
                dur = a['hour_to'] - a['hour_from']
                if cal['two_weeks_calendar'] and a.get('week_type') == '1':
                    heb_b += dur
                else:
                    heb_a += dur
        # le contrat mensuel saisi sur la fiche (chauffeurs : 190 h) prime sur hebdo x 52/12
        contrat_mensuel = e.get('x_contrat_mensuel') or \
            ((((heb_a + heb_b) / 2) if (cal and cal['two_weeks_calendar']) else heb_a) * 52 / 12)
        hs_struct = max(0.0, contrat_mensuel - BASE_LEGALE)
        # récup en heures : mises dans le mois, récupérées dans le mois, solde courant
        ref = e.get('x_cp_ref_date') or ''
        lignes_rl = rl_by_emp.get(e['id'], [])
        pris_rows = pris_by_emp.get(e['id'], [])
        m1, m2 = e1.strftime('%Y-%m-%d'), e2.strftime('%Y-%m-%d')
        mises_mois = sum(l['x_heures'] for l in lignes_rl if m1 <= l['x_date'] <= m2)
        recup_mois = sum(_recup_pris_h(r, cal) for r in pris_rows if m1 <= r['x_date'] <= m2)
        # solde arrêté à la fin du mois exporté (les mouvements postérieurs n'y entrent pas)
        solde = ((e.get('x_recup_solde') or 0.0)
                 + sum(l['x_heures'] for l in lignes_rl if (not ref or l['x_date'] > ref) and l['x_date'] <= m2)
                 - sum(_recup_pris_h(r, cal) for r in pris_rows if (not ref or r['x_date'] > ref) and r['x_date'] <= m2))
        heads = ['Matricule', 'Heures effectuées', 'Heures théoriques', 'Écart',
                 'Contrat mensuel (h)', 'Base légale (h)', 'H. sup structurelles/mois',
                 'Jours CP', 'Jours maladie', 'Jours absence', 'Fériés', 'Jours récup',
                 'Découchages',
                 'H. mises en récup', 'H. récupérées', 'Solde récup (h)']
        vals = [mat or '—', round(tot_h, 2), round(tot_theo, 2), round(tot_h - tot_theo, 2),
                round(contrat_mensuel, 2), round(BASE_LEGALE, 2), round(hs_struct, 2),
                n_cp, n_mal, n_abs, n_fer, n_rec, n_dec,
                round(mises_mois, 2), round(recup_mois, 2), round(solde, 2)]
        for j, (h, v) in enumerate(zip(heads, vals), 1):
            c = ws.cell(row=4, column=j, value=h); c.font = HDR; c.fill = FILL; c.alignment = CTR
            c2 = ws.cell(row=5, column=j, value=v); c2.font = FB; c2.alignment = CTR
        # blocs hebdomadaires
        r = 7
        monday = emonday0
        while monday <= e2:
            sunday = monday + timedelta(days=6)
            c = ws.cell(row=r, column=1, value=f"Semaine {monday.isocalendar()[1]} — du {monday.strftime('%d/%m')} au {sunday.strftime('%d/%m')}")
            c.font = FB
            for j in range(1, 10):
                ws.cell(row=r, column=j).fill = WFILL
            r += 1
            for j, h in enumerate(['Jour', 'Date', 'Arrivée', 'Départ', 'Arrivée', 'Départ',
                                   'Heures', 'Écart', 'Note'], 1):
                c = ws.cell(row=r, column=j, value=h); c.font = HDR; c.fill = FILL; c.alignment = CTR
            r += 1
            wtot = wtheo = 0.0
            for i in range(7):
                d = monday + timedelta(days=i)
                in_month = e1 <= d <= e2
                s = saisies.get(d.strftime('%Y-%m-%d'))
                theo = _theo_day(cal, d)
                if in_month:
                    wtheo += theo
                ws.cell(row=r, column=1, value=JOURS[i]).font = F10
                ws.cell(row=r, column=2, value=d.strftime('%d/%m/%Y')).font = F10
                if s and s['x_type'] == 'travail':
                    for j, v in enumerate([_fmt_h(s['x_m_deb']), _fmt_h(s['x_m_fin']),
                                           _fmt_h(s['x_am_deb']), _fmt_h(s['x_am_fin'])], 3):
                        c = ws.cell(row=r, column=j, value=v); c.font = F10; c.alignment = CTR
                    c = ws.cell(row=r, column=7, value=round(s['x_heures'], 2)); c.font = FB; c.alignment = CTR
                    c = ws.cell(row=r, column=8, value=round(s['x_hs'], 2)); c.font = F10; c.alignment = CTR
                    if in_month:
                        wtot += s['x_heures']
                elif s:
                    lab = TYPES.get(s['x_type'], s['x_type'].upper())
                    for j in range(3, 7):
                        c = ws.cell(row=r, column=j, value=lab); c.font = FB; c.alignment = CTR
                if s and s.get('x_note'):
                    ws.cell(row=r, column=9, value=s['x_note']).font = F10
                if not in_month:
                    for j in range(1, 10):
                        ws.cell(row=r, column=j).font = Font(name='Arial', size=10, color='AAAAAA')
                for j in range(1, 10):
                    ws.cell(row=r, column=j).border = thin
                r += 1
            ws.cell(row=r, column=6, value='Total semaine (part du mois)').font = FB
            c = ws.cell(row=r, column=7, value=round(wtot, 2)); c.font = FB; c.alignment = CTR
            c = ws.cell(row=r, column=8, value=round(wtot - wtheo, 2)); c.font = FB; c.alignment = CTR
            r += 2
            monday += timedelta(days=7)
        for col, w in zip('ABCDEFGHI', [11, 12, 9, 9, 9, 9, 9, 9, 30]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A6'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── EXPORT SILAE (éléments variables de paie) ───────────────────────────────
# Format « standard » en attendant le modèle d'import exact du dossier Silae
# du cabinet : une ligne par élément (Matricule / Code rubrique / Valeur) +
# un onglet Absences par périodes datées. Les codes rubriques par défaut
# ci-dessous sont MODIFIABLES SANS REDÉPLOIEMENT via l'ir.config_parameter
# `maquignon.silae_codes` (JSON, mêmes clés).
SILAE_CODES_DEFAUT = {
    'hs': 'HS',              # heures d'écart du mois (effectué − théorique)
    'cp': 'ABCP',            # congés payés
    'maladie': 'ABMA',       # maladie
    'absence': 'ABNJ',       # absence injustifiée / autre
    'recup': 'ABRC',         # jour de récupération
    'sans_solde': 'ABSS',    # congé sans solde
    'maternite': 'ABMT',     # congé maternité
    'paternite': 'ABPT',     # congé paternité
    'evt_familial': 'ABEF',  # événement familial
    'decouchage': 'DECOU',   # nombre de découchages (prime par nuit)
    'enfant_malade': 'ABEM', # enfant malade
}
SILAE_LIBELLES = {
    'hs': 'Heures écart (+/−)', 'cp': 'Congés payés', 'maladie': 'Maladie',
    'absence': 'Absence injustifiée', 'recup': 'Récupération',
    'decouchage': 'Découchages',
    'sans_solde': 'Congé sans solde', 'maternite': 'Congé maternité',
    'paternite': 'Congé paternité', 'evt_familial': 'Événement familial',
    'enfant_malade': 'Enfant malade',
}
_LBL_TO_KEY = {
    'Congés payés': 'cp', 'Récupération': 'recup', 'Maladie': 'maladie',
    'Sans solde': 'sans_solde', 'Congé maternité': 'maternite',
    'Congé paternité': 'paternite', 'Événement familial': 'evt_familial',
    'Enfant malade': 'enfant_malade',
}


def _abs_key(r):
    """Clé d'absence d'un jour posé (None si jour de travail plein)."""
    t, note = r['x_type'], r.get('x_note') or ''
    if t in ('cp', 'maladie', 'recup'):
        if t == 'cp':
            return 'cp', False
        return t, False
    if t == 'absence':
        lbl = note.split(' approuvé')[0].split(' — ')[0].strip()
        return _LBL_TO_KEY.get(lbl, 'absence'), False
    if t == 'travail' and ' — ' in note and ('matin en congé' in note or 'après-midi en congé' in note):
        lbl = note.split(' — ')[0].strip()
        demi = 'Matin' if 'matin en congé' in note else 'Après-midi'
        return _LBL_TO_KEY.get(lbl, 'absence'), demi
    return None, False


def build_silae(call, mois, comp='all', du=None, au=None, exc=None):
    """Classeur EVP Silae : onglets EVP (matricule/code/valeur), Absences
    (périodes datées) et Lisez-moi (codes + points de contrôle)."""
    import json as _json
    if du and au:
        d1 = datetime.strptime(du, '%Y-%m-%d').date()
        d2 = datetime.strptime(au, '%Y-%m-%d').date()
    else:
        y, m = int(mois[:4]), int(mois[5:7])
        d1 = date(y, m, 1)
        d2 = (date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)) - timedelta(days=1)
    y, m = d1.year, d1.month
    codes = dict(SILAE_CODES_DEFAUT)
    brut = call('ir.config_parameter', 'get_param', 'maquignon.silae_codes')
    if brut:
        try:
            codes.update({k: str(v) for k, v in _json.loads(brut).items()})
        except Exception:
            pass
    exc = {str(k): v for k, v in (exc or {}).items()}
    dom = [('active', '=', True)]
    if comp != 'all':
        dom.append(('company_id', '=', int(comp)))
    emps = call('hr.employee', 'search_read', dom,
                fields=['name', 'company_id', 'resource_calendar_id', 'x_matricule_paie'],
                order='company_id, name')
    cal_ids = sorted(set(e['resource_calendar_id'][0] for e in emps if e['resource_calendar_id']))
    cals = {c['id']: c for c in call('resource.calendar', 'read', cal_ids,
            fields=['name', 'two_weeks_calendar', 'attendance_ids'])} if cal_ids else {}
    att_ids = sorted(set(a for c in cals.values() for a in c['attendance_ids']))
    atts = {a['id']: a for a in call('resource.calendar.attendance', 'read', att_ids,
            fields=['dayofweek', 'hour_from', 'hour_to', 'day_period', 'week_type', 'display_type'])} if att_ids else {}
    for c in cals.values():
        c['attendance_ids'] = [atts[a] for a in c['attendance_ids'] if a in atts]
    rows = call('x_heures_jour', 'search_read',
                [('x_employee_id', 'in', [e['id'] for e in emps]),
                 ('x_date', '>=', min([d1.strftime('%Y-%m-%d')] + [v[0] for v in exc.values()])),
                 ('x_date', '<=', max([d2.strftime('%Y-%m-%d')] + [v[1] for v in exc.values()])),],
                fields=['x_employee_id', 'x_date', 'x_type', 'x_heures', 'x_theo', 'x_note', 'x_decouchage'])
    by_emp = {}
    for r in rows:
        by_emp.setdefault(r['x_employee_id'][0], {})[r['x_date']] = r

    wb = Workbook()
    wb.remove(wb.active)
    FB = Font(name='Arial', size=10, bold=True)
    F10 = Font(name='Arial', size=10)
    HDR = Font(name='Arial', size=9.5, bold=True, color='FFFFFF')
    FILL = PatternFill('solid', start_color='1F4E5F')
    CTR = Alignment(horizontal='center')

    ws = wb.create_sheet('EVP')
    wa = wb.create_sheet('Absences')
    for sheet, heads in ((ws, ['Matricule', 'Salarié', 'Code rubrique', 'Libellé', 'Valeur']),
                         (wa, ['Matricule', 'Salarié', 'Code rubrique', 'Libellé', 'Du', 'Au', 'Jours', 'Demi-journée'])):
        for j, h in enumerate(heads, 1):
            c = sheet.cell(row=1, column=j, value=h); c.font = HDR; c.fill = FILL; c.alignment = CTR
    sans_mat = []
    re_ws, ra = 2, 2
    for e in emps:
        cal = cals.get(e['resource_calendar_id'][0]) if e['resource_calendar_id'] else None
        exc_e = exc.get(str(e['id']))
        e1s = exc_e[0] if exc_e else d1.strftime('%Y-%m-%d')
        e2s = exc_e[1] if exc_e else d2.strftime('%Y-%m-%d')
        saisies = {k: v for k, v in by_emp.get(e['id'], {}).items() if e1s <= k <= e2s}
        mat = e.get('x_matricule_paie') or ''
        # heures d'écart du mois (jours travaillés uniquement)
        hs = sum((s['x_heures'] - s['x_theo']) for s in saisies.values() if s['x_type'] == 'travail')
        # jours d'absence (clé, demi) posés dans le mois
        jours = {}
        for dstr, s in saisies.items():
            key, demi = _abs_key(s)
            if key:
                jours[dstr] = (key, demi)
        n_dec = sum(1 for s2 in saisies.values() if s2.get('x_decouchage'))
        if not jours and abs(hs) < 0.005 and not n_dec:
            continue
        if not mat:
            sans_mat.append(e['name'])
        evp = {}
        if abs(hs) >= 0.005:
            evp['hs'] = round(hs, 2)
        if n_dec:
            evp['decouchage'] = n_dec
        for key, demi in jours.values():
            evp[key] = evp.get(key, 0) + (0.5 if demi else 1)
        for key, val in sorted(evp.items()):
            for j, v in enumerate([mat, e['name'], codes.get(key, key), SILAE_LIBELLES.get(key, key), val], 1):
                c = ws.cell(row=re_ws, column=j, value=v); c.font = F10
            re_ws += 1
        # périodes d'absences : jours consécutifs de même clé (week-ends/jours
        # sans horaire enjambés) ; les demi-journées restent des lignes seules
        pleins = sorted(d for d, (k, demi) in jours.items() if not demi)
        runs = []
        for dstr in pleins:
            key = jours[dstr][0]
            d = datetime.strptime(dstr, '%Y-%m-%d').date()
            if runs and runs[-1][0] == key:
                prev_end = runs[-1][2]
                gap_ok = True
                g = prev_end + timedelta(days=1)
                while g < d:
                    if g.strftime('%Y-%m-%d') in jours or _theo_day(cal, g) > 0:
                        gap_ok = False
                        break
                    g += timedelta(days=1)
                if gap_ok:
                    runs[-1] = (key, runs[-1][1], d, runs[-1][3] + 1)
                    continue
            runs.append((key, d, d, 1))
        for dstr in sorted(d for d, (k, demi) in jours.items() if demi):
            key, demi = jours[dstr]
            d = datetime.strptime(dstr, '%Y-%m-%d').date()
            runs.append((key, d, d, 0.5, demi))
        for run in sorted(runs, key=lambda x: x[1]):
            key, du, au, nb = run[0], run[1], run[2], run[3]
            demi = run[4] if len(run) > 4 else ''
            for j, v in enumerate([mat, e['name'], codes.get(key, key), SILAE_LIBELLES.get(key, key),
                                   du.strftime('%d/%m/%Y'), au.strftime('%d/%m/%Y'), nb, demi], 1):
                c = wa.cell(row=ra, column=j, value=v); c.font = F10
            ra += 1
    for sheet, widths in ((ws, [12, 26, 14, 22, 10]), (wa, [12, 26, 14, 22, 12, 12, 8, 12])):
        for j, w in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + j)].width = w
    # lisez-moi
    wl = wb.create_sheet('Lisez-moi')
    lignes = [
        f"Export Silae (éléments variables de paie) — {mois}",
        "",
        "Format STANDARD provisoire, à caler sur le modèle d'import EVP du",
        "dossier Silae du cabinet (codes rubriques et disposition).",
        "Les codes rubriques se changent SANS redéploiement : paramètre Odoo",
        "« maquignon.silae_codes » (JSON, clés : " + ', '.join(sorted(SILAE_CODES_DEFAUT)) + ").",
        "",
        "Onglet EVP : une ligne par élément (matricule / code / valeur).",
        "  - Heures écart = effectué − théorique des jours travaillés du mois",
        "    (contrôle Charlotte avant import : peut être négatif).",
        "  - Absences en jours (0,5 pour les demi-journées).",
        "Onglet Absences : les mêmes absences par périodes datées, si le",
        "dossier Silae importe les absences par dates plutôt qu'en compteurs.",
        "",
        "Codes utilisés dans ce fichier :",
    ] + [f"  {codes.get(k, k)} = {SILAE_LIBELLES[k]}" for k in sorted(SILAE_CODES_DEFAUT)] + ([
        "",
        "⚠ Salariés SANS matricule (à renseigner sur ⏰ Horaires par défaut) :",
    ] + [f"  - {n}" for n in sans_mat] if sans_mat else [])
    for i, t in enumerate(lignes, 1):
        c = wl.cell(row=i, column=1, value=t)
        c.font = FB if i == 1 or t.startswith('⚠') else F10
    wl.column_dimensions['A'].width = 78
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Feuille de temps hebdomadaire (gabarit Excel fourni par le client) ──────
# Reproduit exactement le classeur d'exemple « HEURES__<salarié> » :
# une feuille par semaine du mois, formules et mise en page du gabarit
# `feuille_temps_template.xlsx` conservées, cellules remplies depuis
# x_heures_jour (CP / FERIE / MALADIE écrits dans les colonnes d'heures).
def _ft_time(v):
    import datetime as _dt
    if not v:
        return None
    h = int(v)
    m = int(round((v - h) * 60))
    if m >= 60:
        h, m = h + 1, m - 60
    return _dt.time(h % 24, m)


def build_feuille(call, mois, emp_id, du=None, au=None):
    import calendar
    import copy as _copy
    import io
    import os
    import re as _re
    import openpyxl
    import datetime as _dt
    from openpyxl.styles import Font

    if du and au:
        d1 = _dt.date.fromisoformat(du)
        d2 = _dt.date.fromisoformat(au)
    else:
        y, mo = int(mois[:4]), int(mois[5:7])
        d1 = _dt.date(y, mo, 1)
        d2 = _dt.date(y, mo, calendar.monthrange(y, mo)[1])
    emp = call('hr.employee', 'read', [emp_id],
               ['name', 'parent_id', 'company_id', 'x_recup_solde',
                'x_cp_ref_date', 'resource_calendar_id'])[0]
    lundi = d1 - _dt.timedelta(days=d1.weekday())
    lundis = []
    while lundi <= d2:
        lundis.append(lundi)
        lundi += _dt.timedelta(days=7)
    jours = {}
    for r in call('x_heures_jour', 'search_read',
                  [('x_employee_id', '=', emp_id),
                   ('x_date', '>=', lundis[0].isoformat()),
                   ('x_date', '<=', (lundis[-1] + _dt.timedelta(days=6)).isoformat())],
                  fields=['x_date', 'x_type', 'x_m_deb', 'x_m_fin', 'x_am_deb',
                          'x_am_fin', 'x_theo', 'x_heures']):
        jours[r['x_date']] = r

    MENTION = {'cp': 'CP', 'ferie': 'FERIE', 'maladie': 'MALADIE',
               'absence': 'ABSENCE', 'recup': 'RECUP'}
    # couleurs du document papier : texte colore Century Gothic, pas de fond
    COULEURS = {'cp': 'FF00B050', 'ferie': 'FFFF0000', 'maladie': 'FFFFC000',
                'absence': 'FFFF0000', 'recup': 'FF0070C0', 'repos': 'FF808080'}
    tpl = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'feuille_temps_template.xlsx')
    wb = openpyxl.load_workbook(tpl)
    master = wb['Semaine']
    table_src = master.tables['TimeSheet']
    theos = [r['x_theo'] for r in jours.values() if r.get('x_theo')]
    theo_semaine = max(theos) if theos else 0
    theo_vendredi = 0
    for k, r in jours.items():
        if _dt.date.fromisoformat(k).weekday() == 4 and r.get('x_theo'):
            theo_vendredi = r['x_theo']
            break

    def entete(ws, fin):
        ws['B2'] = emp['name']
        ws['E2'] = _ft_time(theo_semaine)
        ws['F2'] = _ft_time(theo_vendredi)
        ws['K3'] = fin
        ws['K3'].number_format = 'DD/MM/YYYY'
        ws['K5'] = emp['name']
        ws['K6'] = emp['parent_id'][1] if emp.get('parent_id') else ''

    def table(ws, nom, base, tid):
        t = _copy.deepcopy(table_src)
        t.displayName = nom
        t.name = nom
        t.id = tid
        t.ref = 'B%d:K%d' % (base, base + 7)
        ws.add_table(t)
        ws.cell(base + 8, 11).value = '=SUM(%s[Total])' % nom

    def remplir(ws, base, lundi):
        """Remplit un bloc semaine dont la ligne d'en-tete est `base`.

        Le gabarit d'origine (fichier papier du client) a des formules bancales
        héritées du modèle Microsoft : G/H en #VALEUR! dès qu'un jour porte une
        mention texte (CP...), H parasite sur les jours vides, total G limité à
        lundi-vendredi. On ne garde les formules que sur les jours réellement
        horodatés et on refait la ligne de totaux sur les 7 jours (comme le
        client le faisait lui-même sur ses blocs suivants).
        """
        for i in range(7):
            d = lundi + _dt.timedelta(days=i)
            row = base + 1 + i
            r = jours.get(d.isoformat()) if (d1 <= d <= d2) else None
            mention = MENTION.get((r or {}).get('x_type') or '')
            horaires = bool(r) and not mention and any(
                _ft_time(r.get(ch)) is not None
                for ch in ('x_m_deb', 'x_m_fin', 'x_am_deb', 'x_am_fin'))
            if not horaires:
                # pas d'horaires ce jour : on retire les formules G/H héritées
                # (sinon #VALEUR! sur les mentions, et -8 h sur les jours vides)
                ws.cell(row, 7).value = None
                ws.cell(row, 8).value = None
            k = ws.cell(row, 11)
            if isinstance(k.value, str) and k.value.startswith('='):
                k.value = None    # K reste la colonne d'annotations libres du papier
            if not (d1 <= d <= d2):
                continue          # on n'edite que la periode selectionnee
            c = ws.cell(row, 2)
            c.value = d
            c.number_format = '[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy'
            if not r:
                continue
            if mention:
                for col in (3, 4, 5, 6):
                    cc = ws.cell(row, col)
                    cc.value = mention
                    cc.font = Font(name='Century Gothic', size=12,
                                   color=COULEURS[r['x_type']])
                if r['x_type'] == 'cp':
                    ws.cell(row, 10).value = 1
                if r['x_type'] == 'maladie' and r.get('x_theo'):
                    ws.cell(row, 9).value = r['x_theo']
            else:
                for col, champ in ((3, 'x_m_deb'), (4, 'x_m_fin'),
                                   (5, 'x_am_deb'), (6, 'x_am_fin')):
                    t = _ft_time(r.get(champ))
                    if t is not None:
                        cc = ws.cell(row, col)
                        cc.value = t
                        cc.number_format = 'HH:MM'
        # ligne « Nombre total d'heures » : les 7 jours du bloc
        tot = base + 8
        ws.cell(tot, 7).value = '=SUM(G%d:G%d)*24' % (base + 1, base + 7)
        ws.cell(tot, 8).value = '=SUM(H%d:H%d)' % (base + 1, base + 7)
        ws.cell(tot, 9).value = '=SUM(I%d:I%d)' % (base + 1, base + 7)
        ws.cell(tot, 10).value = '=SUM(J%d:J%d)' % (base + 1, base + 7)

    # ── feuille Recap : toutes les semaines empilees (disposition du papier) ──
    recap = wb.copy_worksheet(master)
    recap.title = 'Récap'
    entete(recap, lundis[-1] + _dt.timedelta(days=6))
    for k, lu in enumerate(lundis):
        hb = 11 + 9 * k          # ligne d'en-tete du bloc (pas de 9 comme l'original)
        if k > 0:
            for sr in range(11, 20):
                dr = hb + (sr - 11)
                for cc in range(2, 12):
                    src = master.cell(sr, cc)
                    dst = recap.cell(dr, cc)
                    v = src.value
                    if isinstance(v, str) and v.startswith('='):
                        if sr == 19:
                            v = (v.replace('G12:G16', 'G%d:G%d' % (hb + 1, hb + 5))
                                  .replace('12:', '%d:' % (hb + 1))
                                  .replace(':18', ':%d' % (hb + 7)))
                        else:
                            v = _re.sub(r'(?<=[A-Z])%d(?![0-9])' % sr,
                                        str(dr), v)
                    dst.value = v
                    dst._style = _copy.copy(src._style)
                if master.row_dimensions[sr].height:
                    recap.row_dimensions[dr].height = master.row_dimensions[sr].height
        table(recap, 'TR_%d' % (k + 1), hb, 100 + k)
        remplir(recap, hb, lu)

    # ── totaux globaux du mois (comme le bas de page du fichier papier) ──
    # mise en forme reprise de la ligne « Nombre total d'heures » du gabarit
    # (libellé gris/gras, cellules de valeurs bordées), « Heures M-1 » en
    # cellule de saisie jaune
    from openpyxl.styles import PatternFill
    tot_rows = [11 + 9 * k + 8 for k in range(len(lundis))]
    rg = tot_rows[-1] + 2
    st_lbl = master.cell(19, 6)._style
    st_val = master.cell(19, 7)._style
    # « Heures M-1 » = compteur d'heures à récupérer (badge 🔄 de la page RH),
    # arrêté à la veille du début de période — prérempli, ajustable à la main
    cal_f = None
    if emp.get('resource_calendar_id'):
        cal_f = call('resource.calendar', 'read', [emp['resource_calendar_id'][0]],
                     ['name', 'two_weeks_calendar', 'attendance_ids'])[0]
        att_f = call('resource.calendar.attendance', 'read', cal_f['attendance_ids'],
                     ['dayofweek', 'week_type', 'display_type', 'hour_from', 'hour_to']) \
            if cal_f['attendance_ids'] else []
        cal_f['attendance_ids'] = att_f
    ref_f = emp.get('x_cp_ref_date') or ''
    m0 = (d1 - _dt.timedelta(days=1)).isoformat()

    def _pris_h(r):
        if r['x_type'] == 'recup':
            return r['x_theo'] or 0.0
        dj = _dt.date.fromisoformat(str(r['x_date'])[:10])
        return max(_theo_day(cal_f, dj) - (r['x_theo'] or 0.0), 0.0)

    m_1 = (emp.get('x_recup_solde') or 0.0)
    m_1 += sum(l['x_heures'] for l in call(
        'x_recup_ligne', 'search_read',
        [('x_employee_id', '=', emp_id), ('x_date', '<=', m0)],
        fields=['x_date', 'x_heures']) if not ref_f or str(l['x_date'])[:10] > ref_f)
    pstart_f = _dt.date(d1.year if d1.month >= 6 else d1.year - 1, 6, 1)
    m_1 -= sum(_pris_h(r) for r in call(
        'x_heures_jour', 'search_read',
        ['&', ('x_employee_id', '=', emp_id),
         ('x_date', '>=', pstart_f.isoformat()), ('x_date', '<=', m0),
         '|', ('x_type', '=', 'recup'), ('x_note', 'like', 'Récupération —')],
        fields=['x_date', 'x_type', 'x_theo', 'x_note'])
        if not ref_f or str(r['x_date'])[:10] > ref_f)

    lignes_tot = [
        ('Nombre total d’heures', {7: '=' + '+'.join('G%d' % r for r in tot_rows),
                                   8: '=' + '+'.join('H%d' % r for r in tot_rows),
                                   9: '=' + '+'.join('I%d' % r for r in tot_rows),
                                   10: '=' + '+'.join('J%d' % r for r in tot_rows)}),
        ('Heures M-1', {7: round(m_1, 2)}),       # prérempli (récup), ajustable
        ('Reste heures', {7: '=G%d+G%d' % (rg, rg + 1)}),
    ]
    for off, (libelle, cols) in enumerate(lignes_tot):
        r = rg + off
        c = recap.cell(r, 6)
        c.value = libelle
        c._style = _copy.copy(st_lbl)
        for col in (7, 8, 9, 10):
            cc = recap.cell(r, col)
            cc._style = _copy.copy(st_val)
            if col in cols:
                cc.value = cols[col]
                cc.number_format = '0.00'
        if libelle == 'Heures M-1':
            cc = recap.cell(r, 7)
            cc.fill = PatternFill('solid', fgColor='FFFFF2CC')
            cc.number_format = '0.00'
        if master.row_dimensions[19].height:
            recap.row_dimensions[r].height = master.row_dimensions[19].height

    # ── une feuille par semaine ──
    for idx, lu in enumerate(lundis):
        ws = wb.copy_worksheet(master)
        ws.title = 'S%02d' % lu.isocalendar()[1]
        entete(ws, lu + _dt.timedelta(days=6))
        table(ws, 'TS_%s' % ws.title, 11, idx + 2)
        remplir(ws, 11, lu)
    del wb['Semaine']
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
