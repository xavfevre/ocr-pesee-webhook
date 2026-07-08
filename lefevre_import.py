# -*- coding: utf-8 -*-
"""
Import commandes LEFEVRE → Odoo — version WEB du convertisseur desktop.
Page : GET /import-lefevre?token=<WEBHOOK_SECRET>
Le code métier (parser, connecteur, export Excel) est repris de
convertisseur_lefevre_app.py (desktop), adapté : openpyxl au lieu de pandas,
product_uom_id (Odoo v19), identifiants via variables d'environnement.
"""
import io
import os
import re
import json
import uuid
import threading
import xmlrpc.client
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, send_file

lefevre_bp = Blueprint("lefevre", __name__)

ODOO_URL = os.environ.get("ODOO_URL", "")
ODOO_DB = os.environ.get("ODOO_DB", "")
ODOO_USER = os.environ.get("ODOO_USER", "")
ODOO_PASSWORD = os.environ.get("ODOO_PASSWORD", "")
SECRET = os.environ.get("WEBHOOK_SECRET", "")

CONFIG_PARAM_KEY = "lefevre.import.config"

DEFAULT_FIELDS = {
    "ref_pierre": "x_studio_ref_pierre",
    "nbr":        "x_studio_nbr",
    "long":       "x_studio_long",
    "larg":       "x_studio_larg",
    "haut":       "x_studio_epais",  # champ « Haut. (m) » (float) sur sale.order.line
    "poids":      "x_studio_poids",
}
DEFAULT_MAPPING = {
    "Usseau":    {"product_code": "TUF0000-PS",   "product": "[TUF0000-PS] Tuffeau - Pierres Pré-sciées 6 faces (Massif)"},
    "Tuffeau":   {"product_code": "TUF0000-PS",   "product": "[TUF0000-PS] Tuffeau - Pierres Pré-sciées 6 faces (Massif)"},
    "Haims":     {"product_code": "HAIMS0020-PS", "product": "[HAIMS0020-PS] Haims - Pré-sciées 6 faces (Massif)"},
    "Migné":     {"product_code": "MIGNE0000-PS", "product": "[MIGNE0000-PS] Migné - Pierres pré-sciées 6 faces (Massif)"},
    "Richemont": {"product_code": "RICH0000-PS",  "product": "[RICH0000-PS] Richemont - Pierres pré-sciées 6 faces (Massif)"},
    "Sireuil":   {"product_code": "SIRE0000-PS",  "product": "[SIRE0000-PS] Sireuil - Pierres pré-sciées 6 faces (Massif)"},
    "Tervoux":   {"product_code": "TERV0000-PS",  "product": "[TERV0000-PS] Tervoux - Pré-sciées 6 faces (Massif)"},
}


def _check_token():
    tok = request.values.get("token") or request.headers.get("X-Token") or ""
    if request.is_json:
        tok = tok or (request.get_json(silent=True) or {}).get("token", "")
    return bool(SECRET) and tok == SECRET


# ══════════════════════════════════════════════════════════════════════
# PARSER LEFEVRE (porté depuis le desktop — openpyxl, sans pandas)
# ══════════════════════════════════════════════════════════════════════

def safe_str(val):
    try:
        if val is None or (isinstance(val, float) and val != val):
            return ""
        return str(val).strip()
    except Exception:
        return ""


def safe_num(val):
    try:
        if val is None or (isinstance(val, float) and val != val):
            return None
        f = float(str(val).replace(",", "."))
        return f if f == f else None
    except Exception:
        return None


COL_KEYWORDS = {
    "ref":    ["désignation", "designation", "n°", "ref", "repere", "repère"],
    "nature": ["nature", "matiere", "matière", "pierre", "type"],
    "qte":    ["qte", "qté", "quantite", "quantité", "nb", "nbre", "nombre"],
    "long":   ["long", "longueur", "l."],
    "haut":   ["haut", "hauteur", "h."],
    "prof":   ["prof", "profondeur", "ep", "ép", "epais", "épaisseur", "larg"],
    "cube":   ["cube", "m3", "m³", "vol", "volume"],
    "poids":  ["poids", "kg", "masse"],
    "scie":   ["scié", "scie", "6 faces", "6f"],
    "obs":    ["obs", "observation", "remarque", "note"],
}


def detect_columns(rows):
    best_score, best_row_idx, best_col_map = 0, None, {}
    for i, row in enumerate(rows):
        if i > 30:
            break
        row_vals = [safe_str(v).lower() for v in row]
        score, col_map = 0, {}
        for field, keywords in COL_KEYWORDS.items():
            for j, cell in enumerate(row_vals):
                matched = False
                for kw in keywords:
                    if kw in cell and cell != "":
                        if field not in col_map:
                            col_map[field] = j
                            score += 1
                        matched = True
                        break
                if matched:
                    break
        if score > best_score and score >= 3:
            best_score, best_row_idx, best_col_map = score, i, col_map
    return best_row_idx, best_col_map


def detect_subsection_pattern(vals, col_ref):
    ref = safe_str(vals[col_ref]) if col_ref < len(vals) else ""
    if not ref:
        return False
    next_vals = [safe_str(vals[col_ref + k]) for k in range(1, 5) if col_ref + k < len(vals)]
    if [v for v in next_vals if safe_num(v) is not None]:
        return False
    return bool(re.match(r'^[A-Za-z]{1,5}[-_]?\d', ref) or
                re.match(r'^[A-Za-z]{2,}[-_]\d', ref) or
                re.match(r'^[A-Z]{1,4}\d{1,3}', ref))


def detect_section_pattern(vals, col_ref, col_map):
    ref = safe_str(vals[col_ref]) if col_ref < len(vals) else ""
    if not ref:
        return False
    if detect_subsection_pattern(vals, col_ref):
        return False
    for c in [col_map.get("qte"), col_map.get("long"), col_map.get("haut"), col_map.get("prof")]:
        if c and c < len(vals) and safe_num(vals[c]) is not None:
            return False
    return True


def detect_unit(values):
    nums = [v for v in values if v is not None and v > 0]
    if not nums:
        return "cm"
    return "mm" if sum(nums) / len(nums) > 300 else "cm"


def parse_header_flexible(rows):
    info = {"chantier": "", "lieu_taille": "", "code_chantier": "",
            "cubage_total": None, "nb_blocs": None}
    kw_chantier = ["chantier", "affaire", "projet", "site"]
    kw_lieu = ["lieu de taille", "lieu taille", "tailleur", "atelier"]
    kw_code = ["code chantier", "code affaire", "ref chantier", "n° chantier"]
    kw_cubage = ["cubage", "volume total", "total m3", "total cube"]
    kw_blocs = ["nombre de blocs", "nb blocs", "total blocs", "nbre blocs"]

    for i, row in enumerate(rows):
        if i > 20:
            break
        row_str = [safe_str(v) for v in row]
        full_row = " | ".join(row_str).lower()
        for j, cell in enumerate(row_str):
            cell_l = cell.lower()
            if any(kw in cell_l for kw in kw_chantier) and not info["chantier"]:
                for k in range(j + 1, min(j + 8, len(row_str))):
                    if row_str[k] and not any(kw in row_str[k].lower() for kw in kw_chantier):
                        info["chantier"] = row_str[k]
                        break
            if any(kw in cell_l for kw in kw_lieu) and not info["lieu_taille"]:
                for k in range(j + 1, min(j + 8, len(row_str))):
                    if row_str[k]:
                        info["lieu_taille"] = row_str[k]
                        break
            if any(kw in cell_l for kw in kw_code) and not info["code_chantier"]:
                for k in range(j + 1, min(j + 8, len(row_str))):
                    if row_str[k]:
                        info["code_chantier"] = row_str[k]
                        break
        if any(kw in full_row for kw in kw_cubage) and not info["cubage_total"]:
            for v in row:
                n = safe_num(v)
                if n and n > 0.01:
                    info["cubage_total"] = n
                    break
        if any(kw in full_row for kw in kw_blocs) and not info["nb_blocs"]:
            for v in row:
                n = safe_num(v)
                if n and n > 1:
                    info["nb_blocs"] = int(n)
                    break
    return info


def load_sheet_rows(file_bytes, sheet_name=None):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = wb.sheetnames
    if sheet_name and sheet_name in sheets:
        target = sheet_name
    else:
        target = sheets[0]
        for s in sheets:
            sl = s.lower()
            if any(k in sl for k in ["taille", "pierre", "devis", "commande", "feuil", "sheet"]):
                target = s
                break
    ws = wb[target]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return sheets, target, rows


def parse_lefevre(file_bytes, nature_mapping, sheet_name=None):
    sheets, target_sheet, rows = load_sheet_rows(file_bytes, sheet_name)

    header_info = parse_header_flexible(rows)
    header_info["sheet_used"] = target_sheet
    header_info["all_sheets"] = sheets

    header_row_idx, col_map = detect_columns(rows)
    if not col_map or "ref" not in col_map:
        col_map = {"ref": 1, "nature": 2, "qte": 3, "long": 4,
                   "haut": 5, "prof": 6, "cube": 7, "poids": 8}
        header_row_idx = None

    col_ref = col_map.get("ref", 1)
    col_nat = col_map.get("nature", col_ref + 1)
    col_qte = col_map.get("qte", col_ref + 2)
    col_long = col_map.get("long", col_ref + 3)
    col_haut = col_map.get("haut", col_ref + 4)
    col_prof = col_map.get("prof", col_ref + 5)
    col_cube = col_map.get("cube", col_ref + 6)
    col_poids = col_map.get("poids", col_ref + 7)

    sample_longs = []
    for row in rows:
        vals = list(row) + [None] * 20
        n = safe_num(vals[col_long])
        if n and n > 0:
            sample_longs.append(n)
        if len(sample_longs) >= 20:
            break
    unit = detect_unit(sample_longs)
    divisor = 10.0 if unit == "mm" else 100.0

    parsed_rows = []
    current_section = ""
    current_subsection = ""
    in_data = False
    start_keywords = ["désignation", "designation", "n°", "repere", "repère",
                      "long", "haut", "prof", "cube"]

    for i, row in enumerate(rows):
        vals = list(row) + [None] * 20
        ref_cell = safe_str(vals[col_ref]).lower()

        if not in_data:
            if header_row_idx is not None and i > header_row_idx:
                in_data = True
            elif any(kw in ref_cell for kw in start_keywords):
                in_data = True
                continue
        if not in_data:
            continue

        ref = safe_str(vals[col_ref])
        nat = safe_str(vals[col_nat])
        qte = safe_num(vals[col_qte])
        long_raw = safe_num(vals[col_long])
        haut_raw = safe_num(vals[col_haut])
        prof_raw = safe_num(vals[col_prof])
        cube_raw = safe_num(vals[col_cube])
        poids_raw = safe_num(vals[col_poids])

        if not ref and not nat and qte is None and long_raw is None:
            continue
        if not ref and qte is None and cube_raw is not None and long_raw is None:
            continue

        if qte is not None and qte > 0 and long_raw is not None and long_raw > 0:
            long_m = round(long_raw / divisor, 5)
            haut_m = round(haut_raw / divisor, 5) if haut_raw else None
            prof_m = round(prof_raw / divisor, 5) if prof_raw else None
            qty_m3 = cube_raw
            if qty_m3 is None and long_m and haut_m and prof_m:
                qty_m3 = round(long_m * haut_m * prof_m * qte, 6)
            elif qty_m3:
                qty_m3 = round(qty_m3, 6)

            prod_info = nature_mapping.get(nat)
            if prod_info is None:
                for k, v in nature_mapping.items():
                    if k.lower() == nat.lower():
                        prod_info = v
                        break

            parsed_rows.append({
                "type": "line",
                "ref": ref if ref else "SUP",
                "nature": nat,
                "qte": int(qte) if qte == int(qte) else qte,
                "long_m": long_m, "haut_m": haut_m, "prof_m": prof_m,
                "qty_m3": qty_m3, "poids": poids_raw,
                "section": current_section, "subsection": current_subsection,
                "product": prod_info["product"] if prod_info else f"[INCONNU] {nat}",
                "product_code": prod_info["product_code"] if prod_info else "???",
                "mapped": prod_info is not None,
            })
        elif detect_subsection_pattern(vals, col_ref):
            current_subsection = ref
            parsed_rows.append({"type": "subsection", "label": ref,
                                "section": current_section, "subsection": ref})
        elif detect_section_pattern(vals, col_ref, col_map):
            # ignorer les résidus d'en-tête pris pour des sections
            if ref.lower().strip(" .:") in ("n°", "no", "n", "ref", "réf",
                                            "désignation", "designation", "repère", "repere"):
                continue
            current_section = ref
            current_subsection = ""
            parsed_rows.append({"type": "section", "label": ref,
                                "section": ref, "subsection": ""})

    if not header_info.get("cubage_total"):
        header_info["cubage_total"] = round(
            sum(r.get("qty_m3") or 0 for r in parsed_rows if r["type"] == "line"), 6)
    header_info["col_map"] = col_map
    header_info["unit"] = unit
    header_info["header_row"] = header_row_idx
    return header_info, parsed_rows


# ══════════════════════════════════════════════════════════════════════
# PARSER v2 — mappage explicite des colonnes + validation manuelle
# (gère les « fiches de débit » multi-onglets : nature en en-tête,
#  N° décalé, sections issues de colonnes dédiées « inline »)
# ══════════════════════════════════════════════════════════════════════

FIELDS_V2 = ["ref", "nature", "qte", "long", "prof", "haut", "cube", "poids",
             "section", "subsection"]


def _sheet_is_recap(name):
    nl = (name or "").lower()
    return any(k in nl for k in ["total", "totaux", "tota", "recap", "récap",
                                 "synth", "somme", "résumé", "resume"])


def list_data_sheets(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    data = [s for s in sheets if not _sheet_is_recap(s)]
    return sheets, (data or sheets)


def load_rows_for(file_bytes, sheet_name):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def build_grid(rows, max_rows=16, max_cols=20):
    ncols = 0
    for r in rows[:60]:
        ncols = max(ncols, len(list(r)))
    ncols = min(ncols, max_cols)
    grid = []
    for r in rows[:max_rows]:
        vals = list(r)
        grid.append([safe_str(vals[c]) if c < len(vals) else "" for c in range(ncols)])
    return grid, ncols


def detect_nature_header(rows):
    for row in rows[:15]:
        rs = [safe_str(v) for v in row]
        for j, cell in enumerate(rs):
            if "nature" in cell.lower():
                for k in range(j + 1, min(j + 8, len(rs))):
                    if rs[k] and "nature" not in rs[k].lower():
                        return rs[k]
    return ""


_DITTO = ('"', "“", "”", "''", "«", "»", "”", "″", "same", "idem")


def _ditto(v):
    v = (v or "").strip()
    return "" if v in _DITTO else v


def _emit_line(out, ref, nat, qte, long_raw, haut_raw, prof_raw, cube_raw,
               poids_raw, divisor, nature_mapping, sec, sub):
    long_m = round(long_raw / divisor, 5) if long_raw else None
    haut_m = round(haut_raw / divisor, 5) if haut_raw else None
    prof_m = round(prof_raw / divisor, 5) if prof_raw else None
    qty_m3 = cube_raw
    if qty_m3 is None and long_m and haut_m and prof_m:
        qty_m3 = round(long_m * haut_m * prof_m * (qte or 1), 6)
    elif qty_m3:
        qty_m3 = round(qty_m3, 6)
    prod = nature_mapping.get(nat)
    if prod is None:
        for k, v in nature_mapping.items():
            if k.lower() == (nat or "").lower():
                prod = v
                break
    out.append({
        "type": "line",
        "ref": ref if ref else "SUP",
        "nature": nat,
        "qte": int(qte) if (qte is not None and qte == int(qte)) else qte,
        "long_m": long_m, "haut_m": haut_m, "prof_m": prof_m,
        "qty_m3": qty_m3, "poids": poids_raw,
        "section": sec, "subsection": sub,
        "product": prod["product"] if prod else f"[INCONNU] {nat or ''}",
        "product_code": prod["product_code"] if prod else "???",
        "mapped": prod is not None,
    })


def _parse_one_sheet(rows, mp, nature_mapping, sheet_nature):
    cols = mp.get("cols", {})
    c_ref = cols.get("ref")
    c_nat = cols.get("nature")
    c_qte = cols.get("qte")
    c_long = cols.get("long")
    c_haut = cols.get("haut")
    c_prof = cols.get("prof")
    c_cube = cols.get("cube")
    c_poids = cols.get("poids")
    c_sec = cols.get("section")
    c_sub = cols.get("subsection")
    header_row = mp.get("header_row")
    unit = mp.get("unit") or "auto"

    def gv(vals, c):
        return vals[c] if (c is not None and c < len(vals)) else None

    if unit not in ("cm", "mm"):
        sample = []
        for row in rows:
            vals = list(row) + [None] * 30
            n = safe_num(gv(vals, c_long))
            if n and n > 0:
                sample.append(n)
            if len(sample) >= 20:
                break
        unit = detect_unit(sample)
    divisor = 10.0 if unit == "mm" else 100.0

    inline = (c_sec is not None) or (c_sub is not None)
    out = []
    cur_sec, cur_sub = "", ""
    start = (header_row + 1) if header_row is not None else 0

    for i, row in enumerate(rows):
        if i < start:
            continue
        vals = list(row) + [None] * 30
        ref = safe_str(gv(vals, c_ref))
        qte = safe_num(gv(vals, c_qte))
        long_raw = safe_num(gv(vals, c_long))
        haut_raw = safe_num(gv(vals, c_haut))
        prof_raw = safe_num(gv(vals, c_prof))
        cube_raw = safe_num(gv(vals, c_cube))
        poids_raw = safe_num(gv(vals, c_poids))
        nat = safe_str(gv(vals, c_nat)) if c_nat is not None else ""

        if inline:
            sec_v = _ditto(safe_str(gv(vals, c_sec))) if c_sec is not None else ""
            sub_v = _ditto(safe_str(gv(vals, c_sub))) if c_sub is not None else ""
            # met à jour section/sous-section même sur une ligne de donnée
            if c_sec is not None and sec_v and sec_v != cur_sec:
                cur_sec, cur_sub = sec_v, ""
                out.append({"type": "section", "label": cur_sec,
                            "section": cur_sec, "subsection": ""})
            if c_sub is not None and sub_v and sub_v != cur_sub:
                cur_sub = sub_v
                out.append({"type": "subsection", "label": cur_sub,
                            "section": cur_sec, "subsection": cur_sub})
            # ligne réelle : réf présente + une dimension/cube + qté
            is_line = (bool(ref) and (qte is not None and qte > 0)
                       and ((long_raw and long_raw > 0) or (cube_raw and cube_raw > 0)))
            if is_line:
                _emit_line(out, ref, nat or sheet_nature, qte, long_raw, haut_raw,
                           prof_raw, cube_raw, poids_raw, divisor, nature_mapping,
                           cur_sec, cur_sub)
            continue

        # ── mode « motif » (ancien format : sections détectées) ──
        if not ref and not nat and qte is None and long_raw is None:
            continue
        if qte is not None and qte > 0 and long_raw is not None and long_raw > 0:
            _emit_line(out, ref, nat or sheet_nature, qte, long_raw, haut_raw,
                       prof_raw, cube_raw, poids_raw, divisor, nature_mapping,
                       cur_sec, cur_sub)
        elif detect_subsection_pattern(vals, c_ref if c_ref is not None else 1):
            cur_sub = ref
            out.append({"type": "subsection", "label": ref,
                        "section": cur_sec, "subsection": ref})
        elif detect_section_pattern(vals, c_ref if c_ref is not None else 1,
                                    {"qte": c_qte, "long": c_long,
                                     "haut": c_haut, "prof": c_prof}):
            if ref.lower().strip(" .:") in ("n°", "no", "n", "ref", "réf",
                                            "désignation", "designation",
                                            "repère", "repere"):
                continue
            cur_sec, cur_sub = ref, ""
            out.append({"type": "section", "label": ref,
                        "section": ref, "subsection": ""})
    return out


def parse_lefevre_explicit(file_bytes, mp, nature_mapping, sheet):
    all_sheets, data_sheets = list_data_sheets(file_bytes)
    if sheet == "__ALL__":
        targets = data_sheets
        merge = True
    else:
        if sheet not in all_sheets:
            sheet = data_sheets[0]
        targets = [sheet]
        merge = False

    hrows = load_rows_for(file_bytes, targets[0])
    header_info = parse_header_flexible(hrows)
    header_info["nature_header"] = detect_nature_header(hrows)
    header_info["all_sheets"] = all_sheets
    header_info["data_sheets"] = data_sheets
    header_info["sheet_used"] = sheet
    header_info["unit"] = mp.get("unit") or "auto"
    header_info["header_row"] = mp.get("header_row")

    has_section_col = mp.get("cols", {}).get("section") is not None
    parsed_rows = []
    for sname in targets:
        rows = load_rows_for(file_bytes, sname)
        sheet_nature = mp.get("global_nature") or detect_nature_header(rows) \
            or header_info.get("nature_header") or ""
        sub = _parse_one_sheet(rows, mp, nature_mapping, sheet_nature)
        if merge and not has_section_col and sub:
            parsed_rows.append({"type": "section", "label": sname,
                                "section": sname, "subsection": ""})
        parsed_rows.extend(sub)

    if not header_info.get("cubage_total"):
        header_info["cubage_total"] = round(
            sum(r.get("qty_m3") or 0 for r in parsed_rows if r["type"] == "line"), 6)
    return header_info, parsed_rows


def guess_mapping(rows):
    """Devine un mappage de colonnes (pré-remplissage de l'étape de validation).
    Repère la ligne d'en-tête via les mots-clés de dimensions, trouve le début
    réel des données, décale la réf. quand la colonne « désignation » ne contient
    que des libellés de groupe, et route les sections vers « Observations »."""
    dim_fields = ["long", "prof", "haut", "qte", "cube", "poids"]

    def row_dim_score(i):
        rs = [safe_str(v).lower() for v in rows[i]]
        s = 0
        for fld in dim_fields:
            if any(txt and any(kw in txt for kw in COL_KEYWORDS[fld]) for txt in rs):
                s += 1
        return s

    base, bs = -1, 0
    for i in range(min(len(rows), 25)):
        s = row_dim_score(i)
        if s > bs:
            base, bs = i, s

    if base < 0 or bs < 3:
        return {"header_row": None, "global_nature": "", "unit": "auto",
                "cols": {"ref": 1, "nature": None, "qte": 3, "long": 4,
                         "haut": 5, "prof": 6, "cube": 7, "poids": 8,
                         "section": None, "subsection": None}}

    # colonnes : mots-clés de la ligne d'en-tête + la ligne juste en dessous
    r0 = [safe_str(v).lower() for v in rows[base]]
    r1 = [safe_str(v).lower() for v in (rows[base + 1] if base + 1 < len(rows) else [])]
    ncol = max(len(r0), len(r1))
    ctext = [((r0[c] if c < len(r0) else "") + " " +
              (r1[c] if c < len(r1) else "")).strip() for c in range(ncol)]

    def find(field):
        for j, txt in enumerate(ctext):
            if txt and any(kw in txt for kw in COL_KEYWORDS[field]):
                return j
        return None

    cols = {f: find(f) for f in ["ref", "nature", "qte", "long",
                                 "prof", "haut", "cube", "poids"]}
    cols["section"] = None
    cols["subsection"] = None
    obs = find("obs")

    # début réel des données : 1re ligne après l'en-tête avec un nombre en qté/long/cube
    probe = [cols["qte"], cols["long"], cols["cube"]]
    data_start = base + 1
    for i in range(base + 1, min(len(rows), base + 6)):
        vals = list(rows[i]) + [None] * 30
        if any((safe_num(vals[c]) or 0) > 0 for c in probe if c is not None):
            data_start = i
            break
    header_row = data_start - 1

    used = lambda: [c for c in cols.values() if isinstance(c, int)]
    data = rows[data_start: data_start + 12]

    def col_vals(c):
        vv = []
        for r in data:
            vals = list(r) + [None] * 30
            vv.append(safe_str(vals[c]) if (c is not None and c < len(vals)) else "")
        return [v for v in vv if v]

    # nature collision avec réf. → on retombera sur la nature par défaut
    if cols["nature"] is not None and cols["nature"] == cols["ref"]:
        cols["nature"] = None

    # décalage réf. : colonne « désignation » = libellés de groupe, N° en colonne suivante
    # (déclenché seulement en présence d'une colonne « Observations » = fiche de débit)
    if cols["ref"] is not None and obs is not None:
        rv = col_vals(cols["ref"])
        looks_group = rv and sum(1 for v in rv if (" " in v or not re.search(r"\d", v))) / len(rv) > 0.5
        nxt = cols["ref"] + 1
        nv = col_vals(nxt)
        nxt_refs = nv and sum(1 for v in nv if re.search(r"\d", v)) / len(nv) > 0.5
        if looks_group and nxt not in used() and nxt_refs:
            cols["subsection"] = cols["ref"]
            cols["ref"] = nxt

    if obs is not None and obs not in used():
        cols["section"] = obs

    gn = detect_nature_header(rows) if cols["nature"] is None else ""
    return {"header_row": header_row, "global_nature": gn, "unit": "auto", "cols": cols}


# ══════════════════════════════════════════════════════════════════════
# CONNECTEUR ODOO (repris du desktop ; correctif v19 product_uom_id)
# ══════════════════════════════════════════════════════════════════════

class OdooConnector:
    def __init__(self, url=ODOO_URL, database=ODOO_DB, username=ODOO_USER, password=ODOO_PASSWORD):
        self.url = url.rstrip("/")
        self.db = database
        self.username = username
        self.password = password
        self.uid = None
        self._models = None
        self._product_cache = {}

    def connect(self):
        common = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/common", allow_none=True)
        self.uid = common.authenticate(self.db, self.username, self.password, {})
        if not self.uid:
            raise ConnectionError("Authentification Odoo échouée (variables d'environnement).")
        self._models = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/object", allow_none=True)
        return common.version().get("server_version", "?")

    def call(self, model, method, args, kwargs=None):
        return self._models.execute_kw(self.db, self.uid, self.password,
                                       model, method, args, kwargs or {})

    def search_partner(self, name=None, ref=None):
        if ref:
            res = self.call("res.partner", "search_read", [[["ref", "=", ref]]],
                            {"fields": ["id", "name", "ref"], "limit": 1})
            if res:
                return res[0]
        if name:
            res = self.call("res.partner", "search_read", [[["name", "ilike", name]]],
                            {"fields": ["id", "name", "ref"], "limit": 1})
            if res:
                return res[0]
        return None

    def preload_products(self, codes):
        codes = [c for c in set(codes) if c and not c.startswith("???")]
        res = self.call("product.product", "search_read",
                        [[["default_code", "in", codes]]],
                        {"fields": ["id", "name", "default_code", "uom_id"]})
        for p in res:
            if p.get("default_code"):
                self._product_cache[p["default_code"]] = p
        missing = [c for c in codes if c not in self._product_cache]
        if missing:
            tpls = self.call("product.template", "search_read",
                             [[["default_code", "in", missing]]],
                             {"fields": ["id", "name", "default_code", "product_variant_ids"]})
            variant_ids, tpl_map = [], {}
            for t in tpls:
                if t.get("product_variant_ids"):
                    vid = t["product_variant_ids"][0]
                    variant_ids.append(vid)
                    tpl_map[vid] = t["default_code"]
            if variant_ids:
                for v in self.call("product.product", "read", [variant_ids],
                                   {"fields": ["id", "name", "default_code", "uom_id"]}):
                    code = tpl_map.get(v["id"])
                    if code:
                        self._product_cache[code] = v
        for c in codes:
            self._product_cache.setdefault(c, None)
        found = sum(1 for v in self._product_cache.values() if v)
        return found, len(codes)

    def get_uom_m3(self):
        res = self.call("uom.uom", "search_read", [[["name", "ilike", "m³"]]],
                        {"fields": ["id"], "limit": 1})
        if not res:
            res = self.call("uom.uom", "search_read", [[["name", "ilike", "m3"]]],
                            {"fields": ["id"], "limit": 1})
        return res[0]["id"] if res else False

    def create_sale_order(self, partner_id, order_ref, date_order):
        vals = {"partner_id": partner_id, "state": "draft"}
        if date_order:
            iso_date = date_order
            for fmt_in in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    iso_date = datetime.strptime(date_order, fmt_in).strftime("%Y-%m-%d %H:%M:%S")
                    break
                except ValueError:
                    continue
            vals["date_order"] = iso_date
        if order_ref:
            vals["client_order_ref"] = order_ref
        return self.call("sale.order", "create", [vals])

    def create_order_line_section(self, order_id, name):
        return self.call("sale.order.line", "create", [{
            "order_id": order_id, "display_type": "line_section", "name": name}])

    def create_order_line_product(self, order_id, line_data, fields_cfg, uom_id=False, taille=False):
        product = self._product_cache.get(line_data["product_code"])
        if not product:
            return None, f"Produit introuvable: {line_data['product_code']}"
        vals = {
            "order_id": order_id,
            "product_id": product["id"],
            "product_uom_qty": line_data.get("qty_m3") or 0,
            "name": line_data.get("product", ""),
        }
        if uom_id:
            vals["product_uom_id"] = uom_id  # v19 (ex product_uom)
        studio = {
            fields_cfg.get("ref_pierre", "x_studio_ref_pierre"): line_data.get("ref"),
            fields_cfg.get("nbr", "x_studio_nbr"): line_data.get("qte"),
            fields_cfg.get("long", "x_studio_long"): line_data.get("long_m"),
            fields_cfg.get("larg", "x_studio_larg"): line_data.get("prof_m"),
            fields_cfg.get("haut", "x_studio_epais"): line_data.get("haut_m"),
            fields_cfg.get("poids", "x_studio_poids"): line_data.get("poids"),
        }
        for f, v in studio.items():
            if v:
                vals[f] = v
        # « Taille de pierre » : cochée → flag booléen + mot « taille » ; sinon champ vide
        if taille:
            vals["x_studio_taille_de_pierre"] = True
            vals["x_studio_taille_de_p"] = "taille"
        core = {"order_id", "product_id", "product_uom_qty", "name", "product_uom_id"}
        try:
            return self.call("sale.order.line", "create", [vals]), None
        except Exception as e:
            for f in list(vals):
                if f not in core:
                    vals.pop(f, None)
            try:
                line_id = self.call("sale.order.line", "create", [vals])
                return line_id, f"⚠ Champs Studio ignorés pour {line_data.get('ref')}: {e}"
            except Exception as e2:
                return None, f"❌ Erreur ligne {line_data.get('ref')}: {e2}"

    def check_studio_fields(self, fields_cfg):
        info = self.call("sale.order.line", "fields_get",
                         [list(fields_cfg.values())], {"attributes": ["string"]})
        return {k: (f in info) for k, f in fields_cfg.items()}

    # config persistée dans Odoo (survit aux redéploiements Render)
    def load_config(self):
        raw = self.call("ir.config_parameter", "get_param", [CONFIG_PARAM_KEY])
        cfg = {"fields": dict(DEFAULT_FIELDS), "nature_mapping": dict(DEFAULT_MAPPING)}
        if raw:
            try:
                saved = json.loads(raw)
                cfg["fields"].update(saved.get("fields", {}))
                if saved.get("nature_mapping"):
                    cfg["nature_mapping"] = saved["nature_mapping"]
            except Exception:
                pass
        return cfg

    def save_config(self, cfg):
        self.call("ir.config_parameter", "set_param",
                  [CONFIG_PARAM_KEY, json.dumps(cfg, ensure_ascii=False)])


# ══════════════════════════════════════════════════════════════════════
# EXPORT EXCEL (repris du desktop)
# ══════════════════════════════════════════════════════════════════════

COLS_ODOO = [
    "Order Reference", "Order Date", "Customer", "Customer Code",
    "Order Lines / Display Type", "Order Lines / Name",
    "Order Lines / Ref. Pierre", "Order Lines / Product",
    "Order Lines / Product Code", "Order Lines / Nbr.",
    "Order Lines / Long.", "Order Lines / Larg.", "Order Lines / Haut.",
    "Order Lines / Quantity", "Order Lines / Poids (kg)",
]


def fmt_num(val, d=3):
    return str(round(val, d)).replace(".", ",") if val is not None else ""


def generate_excel(parsed_rows, order_ref, customer, customer_code, order_date):
    def rgb(h):
        return "FF" + h.lstrip("#").upper()

    def sc(cell, bold=False, bg=None, fg="000000", sz=9, align="left", border=False):
        cell.font = Font(name="Calibri", size=sz, bold=bold, color=fg)
        if bg:
            cell.fill = PatternFill("solid", start_color=rgb(bg))
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if border:
            t = Side(style="thin", color="BBBBBB")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

    C_H, C_S, C_SS = "1A3A5C", "2E6DA4", "5B9BD5"
    C_ALT, C_W, C_WARN = "EEF4FB", "FFFFFF", "FFF2CC"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odoo_Import"
    ws.row_dimensions[1].height = 28
    for ci, cn in enumerate(COLS_ODOO, 1):
        sc(ws.cell(row=1, column=ci, value=cn), bold=True, bg=C_H, fg="FFFFFF",
           sz=10, align="center", border=True)

    row_num, first_line = 2, True
    for i, item in enumerate(parsed_rows):
        ws.row_dimensions[row_num].height = 17
        r = {c: "" for c in COLS_ODOO}
        if first_line:
            r["Order Reference"], r["Order Date"] = order_ref, order_date
            r["Customer"], r["Customer Code"] = customer, customer_code
        if item["type"] in ("section", "subsection"):
            r["Order Lines / Display Type"] = "line_section"
            r["Order Lines / Name"] = item["label"] if item["type"] == "section" else f"  {item['label']}"
            bg = C_S if item["type"] == "section" else C_SS
            for ci, cn in enumerate(COLS_ODOO, 1):
                sc(ws.cell(row=row_num, column=ci, value=r[cn]),
                   bold=True, bg=bg, fg="FFFFFF", sz=10, border=True)
        else:
            r["Order Lines / Ref. Pierre"] = item["ref"]
            r["Order Lines / Product"] = item["product"]
            r["Order Lines / Product Code"] = item["product_code"]
            r["Order Lines / Nbr."] = item["qte"]
            r["Order Lines / Long."] = fmt_num(item["long_m"])
            r["Order Lines / Larg."] = fmt_num(item["prof_m"])
            r["Order Lines / Haut."] = fmt_num(item["haut_m"])
            r["Order Lines / Quantity"] = fmt_num(item["qty_m3"], 6)
            r["Order Lines / Poids (kg)"] = fmt_num(item.get("poids"), 3) if item.get("poids") else ""
            bg = C_WARN if item["product_code"].startswith("???") else (C_ALT if i % 2 == 0 else C_W)
            for ci, cn in enumerate(COLS_ODOO, 1):
                sc(ws.cell(row=row_num, column=ci, value=r[cn]), bg=bg, sz=9, border=True)
        first_line = False
        row_num += 1

    for i, w in enumerate([18, 20, 22, 14, 16, 28, 16, 50, 16, 7, 7, 7, 7, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════
# SESSIONS D'IMPORT (thread + polling)
# ══════════════════════════════════════════════════════════════════════

_SESSIONS = {}
_LOCK = threading.Lock()


def _slog(sid, msg, tag="info"):
    with _LOCK:
        s = _SESSIONS.get(sid)
        if s is not None:
            s["logs"].append({"msg": msg, "tag": tag})


def _sprog(sid, val):
    with _LOCK:
        s = _SESSIONS.get(sid)
        if s is not None:
            s["progress"] = val


def _run_import(sid, payload):
    try:
        parsed_rows = payload["rows"]
        order_ref = payload.get("order_ref") or ""
        customer = payload.get("customer") or "LEFEVRE"
        customer_code = payload.get("customer_code") or ""
        order_date = payload.get("order_date") or ""
        taille = bool(payload.get("taille"))

        _slog(sid, "Connexion à Odoo…")
        conn = OdooConnector()
        version = conn.connect()
        _slog(sid, f"✓ Connecté — Odoo {version}", "ok")
        cfg = conn.load_config()
        fields_cfg = cfg["fields"]

        _slog(sid, f"Recherche client : {customer_code} / {customer}…")
        partner = conn.search_partner(name=customer, ref=customer_code)
        if not partner:
            _slog(sid, f"⚠ Client non trouvé — création de '{customer}'", "warn")
            partner_id = conn.call("res.partner", "create", [{
                "name": customer, "ref": customer_code, "is_company": True}])
        else:
            partner_id = partner["id"]
            _slog(sid, f"✓ Client trouvé : [{partner_id}] {partner.get('name', '')}", "ok")

        uom_id = conn.get_uom_m3()
        _slog(sid, f"Création du devis {order_ref}…")
        order_id = conn.create_sale_order(partner_id, order_ref, order_date)
        try:
            order_name = conn.call("sale.order", "read", [[order_id]], {"fields": ["name"]})[0]["name"]
        except Exception:
            order_name = str(order_id)
        _slog(sid, f"✓ Devis créé : {order_name} (ID: {order_id})", "ok")
        _sprog(sid, 10)

        all_codes = list({r["product_code"] for r in parsed_rows
                          if r["type"] == "line" and not r["product_code"].startswith("???")})
        found, total_codes = conn.preload_products(all_codes)
        _slog(sid, f"✓ Produits trouvés : {found}/{total_codes}",
              "ok" if found == total_codes else "warn")
        not_found = [c for c in all_codes if conn._product_cache.get(c) is None]
        if not_found:
            _slog(sid, f"⚠ Produits non trouvés dans Odoo : {not_found}", "warn")

        total = len(parsed_rows)
        ok_count = warn_count = err_count = 0
        logged_codes = set()
        for i, item in enumerate(parsed_rows):
            _sprog(sid, 15 + int((i / max(total, 1)) * 83))
            if item["type"] in ("section", "subsection"):
                lbl = item["label"] if item["type"] == "section" else f"  {item['label']}"
                conn.create_order_line_section(order_id, lbl)
                ok_count += 1
            elif item["type"] == "line":
                line_id, err = conn.create_order_line_product(order_id, item, fields_cfg, uom_id, taille)
                if err and line_id:
                    _slog(sid, f"  {err}", "warn")
                    warn_count += 1
                    ok_count += 1
                elif err:
                    code = item.get("product_code", "")
                    if code not in logged_codes:
                        _slog(sid, f"  {err}", "err")
                        logged_codes.add(code)
                    err_count += 1
                else:
                    ok_count += 1

        order_url = (f"{ODOO_URL.rstrip('/')}/odoo/sales/{order_id}")
        _slog(sid, "══ IMPORT TERMINÉ ══════════════════", "ok")
        _slog(sid, f"✓ Devis Odoo : {order_name}", "ok")
        _slog(sid, f"✓ Lignes créées : {ok_count}", "ok")
        if warn_count:
            _slog(sid, f"⚠ Avertissements : {warn_count}", "warn")
        if err_count:
            _slog(sid, f"❌ Erreurs : {err_count}", "err")
        with _LOCK:
            _SESSIONS[sid].update({"done": True, "progress": 100,
                                   "order_id": order_id, "order_name": order_name,
                                   "order_url": order_url})
    except Exception as e:
        _slog(sid, f"❌ Erreur : {e}", "err")
        with _LOCK:
            _SESSIONS[sid].update({"done": True, "error": str(e)[:300]})


# ══════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════

@lefevre_bp.route("/import-lefevre/parse", methods=["POST"])
def lef_parse():
    if not _check_token():
        return jsonify({"error": "token invalide"}), 401
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "fichier manquant"}), 400
    file_bytes = f.read()
    sheet = request.form.get("sheet") or None
    mapping_json = request.form.get("mapping")
    try:
        conn = OdooConnector()
        conn.connect()
        nature_mapping = conn.load_config()["nature_mapping"]
    except Exception:
        nature_mapping = dict(DEFAULT_MAPPING)

    try:
        all_sheets, data_sheets = list_data_sheets(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Fichier illisible : {e}"}), 422

    # feuille de référence pour la grille / la devinette
    if sheet == "__ALL__" or not sheet or sheet not in all_sheets:
        ref_sheet = data_sheets[0]
    else:
        ref_sheet = sheet
    # par défaut : fusionner toutes les feuilles de données s'il y en a plusieurs
    use_sheet = sheet or ("__ALL__" if len(data_sheets) > 1 else data_sheets[0])

    if mapping_json:
        try:
            mp = json.loads(mapping_json)
        except Exception:
            mp = guess_mapping(load_rows_for(file_bytes, ref_sheet))
    else:
        mp = guess_mapping(load_rows_for(file_bytes, ref_sheet))

    try:
        header_info, parsed_rows = parse_lefevre_explicit(file_bytes, mp, nature_mapping, use_sheet)
    except Exception as e:
        return jsonify({"error": f"Parsing impossible : {e}"}), 422

    grid, ncols = build_grid(load_rows_for(file_bytes, ref_sheet))
    header_info["grid"] = grid
    header_info["ncols"] = ncols
    header_info["mapping"] = mp
    header_info["sheet_used"] = use_sheet

    lines = [r for r in parsed_rows if r["type"] == "line"]
    stats = {
        "sections": sum(1 for r in parsed_rows if r["type"] == "section"),
        "subsections": sum(1 for r in parsed_rows if r["type"] == "subsection"),
        "lines": len(lines),
        "pieces": sum((r["qte"] or 0) for r in lines),
        "cubage": round(sum((r["qty_m3"] or 0) for r in lines), 6),
        "unknown": sum(1 for r in lines if not r["mapped"]),
    }
    return jsonify({"header_info": {k: v for k, v in header_info.items() if k != "col_map"},
                    "stats": stats, "rows": parsed_rows})


@lefevre_bp.route("/import-lefevre/export", methods=["POST"])
def lef_export():
    if not _check_token():
        return jsonify({"error": "token invalide"}), 401
    d = request.get_json(force=True)
    buf = generate_excel(d["rows"], d.get("order_ref", ""), d.get("customer", ""),
                         d.get("customer_code", ""), d.get("order_date", ""))
    name = f"Import_Odoo_{(d.get('order_ref') or 'lefevre').replace('/', '-')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@lefevre_bp.route("/import-lefevre/import", methods=["POST"])
def lef_import():
    if not _check_token():
        return jsonify({"error": "token invalide"}), 401
    payload = request.get_json(force=True)
    if not payload.get("rows"):
        return jsonify({"error": "aucune ligne"}), 400
    sid = uuid.uuid4().hex[:12]
    with _LOCK:
        _SESSIONS[sid] = {"logs": [], "progress": 0, "done": False}
        # purge des vieilles sessions
        if len(_SESSIONS) > 20:
            for k in list(_SESSIONS.keys())[:-10]:
                _SESSIONS.pop(k, None)
    threading.Thread(target=_run_import, args=(sid, payload), daemon=True).start()
    return jsonify({"sid": sid})


@lefevre_bp.route("/import-lefevre/status", methods=["GET"])
def lef_status():
    if not _check_token():
        return jsonify({"error": "token invalide"}), 401
    sid = request.args.get("sid", "")
    after = int(request.args.get("after", 0))
    with _LOCK:
        s = _SESSIONS.get(sid)
        if not s:
            return jsonify({"error": "session inconnue"}), 404
        return jsonify({"logs": s["logs"][after:], "next": len(s["logs"]),
                        "progress": s["progress"], "done": s["done"],
                        "order_id": s.get("order_id"), "order_name": s.get("order_name"),
                        "order_url": s.get("order_url"),
                        "error": s.get("error")})


@lefevre_bp.route("/import-lefevre/config", methods=["GET", "POST"])
def lef_config():
    if not _check_token():
        return jsonify({"error": "token invalide"}), 401
    conn = OdooConnector()
    conn.connect()
    if request.method == "POST":
        d = request.get_json(force=True)
        cfg = conn.load_config()
        if "nature_mapping" in d:
            cfg["nature_mapping"] = d["nature_mapping"]
        if "fields" in d:
            cfg["fields"].update(d["fields"])
        conn.save_config(cfg)
        return jsonify({"ok": True})
    cfg = conn.load_config()
    try:
        cfg["fields_check"] = conn.check_studio_fields(cfg["fields"])
    except Exception:
        cfg["fields_check"] = {}
    return jsonify(cfg)


@lefevre_bp.route("/import-lefevre", methods=["GET"])
def lef_page():
    if not _check_token():
        return "<h3 style='font-family:sans-serif'>⛔ Accès refusé — token manquant ou invalide.</h3>", 401
    today = datetime.now().strftime("%d/%m/%Y") + " 00:00:00"
    return PAGE_HTML.replace("__TODAY__", today)


PAGE_HTML = r"""<!DOCTYPE html><html lang="fr"><head><meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Import Lefevre → Odoo</title><style>
:root{--bg:#0F1923;--sidebar:#1A2535;--card:#1E2D40;--border:#2A3F58;--accent:#2E6DA4;
--accent2:#5B9BD5;--success:#27AE60;--warning:#F39C12;--error:#E74C3C;--text:#ECF0F1;
--dim:#8B9BB4;--entry:#152030;}
*{box-sizing:border-box;} body{margin:0;background:var(--bg);color:var(--text);
font-family:'Segoe UI',system-ui,sans-serif;}
.hdr{background:var(--sidebar);padding:14px 22px;display:flex;align-items:center;gap:14px;border-bottom:2px solid var(--border);}
.hdr h2{margin:0;font-size:19px;} .hdr .sub{color:var(--dim);font-size:12px;}
.tabs{display:flex;gap:2px;background:var(--sidebar);padding:0 14px;}
.tab{padding:11px 20px;cursor:pointer;font-weight:700;font-size:14px;color:var(--dim);border-bottom:3px solid transparent;}
.tab.on{color:var(--text);border-color:var(--accent2);}
.wrap{max-width:1150px;margin:0 auto;padding:18px 14px 60px;}
.card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:14px;}
.card h4{margin:0 0 12px;color:var(--accent2);font-size:13px;text-transform:uppercase;letter-spacing:1px;}
.drop{border:2px dashed var(--accent);border-radius:10px;padding:34px 16px;text-align:center;color:var(--dim);cursor:pointer;font-size:15px;}
.drop.hov{background:var(--entry);border-color:var(--accent2);color:var(--text);}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:10px 18px;}
label{display:block;font-size:12px;color:var(--dim);margin-bottom:3px;}
input[type=text],select{width:100%;background:var(--entry);border:1px solid var(--border);border-radius:7px;color:var(--text);padding:9px 10px;font-size:14px;}
.btns{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px;}
.btn{border:none;border-radius:8px;padding:12px 22px;font-weight:800;font-size:15px;cursor:pointer;}
.btn.p{background:var(--accent);color:#fff;} .btn.g{background:var(--success);color:#fff;}
.btn.s{background:var(--entry);color:var(--text);border:1px solid var(--border);}
.btn:disabled{opacity:.4;cursor:not-allowed;}
.kpis{display:flex;gap:12px;flex-wrap:wrap;}
.kpi{background:var(--entry);border:1px solid var(--border);border-radius:9px;padding:9px 16px;font-size:13px;}
.kpi b{display:block;font-size:19px;color:var(--accent2);}
.kpi.warn b{color:var(--warning);}
table{width:100%;border-collapse:collapse;font-size:12.5px;}
th{background:var(--sidebar);padding:7px 8px;text-align:left;position:sticky;top:0;}
td{padding:5px 8px;border-bottom:1px solid var(--border);}
tr.sec td{background:var(--accent);color:#fff;font-weight:800;}
tr.sub td{background:var(--accent2);color:#fff;font-weight:700;}
tr.unmapped td{background:#4a1f1f;}
.tblbox{max-height:65vh;overflow:auto;border:1px solid var(--border);border-radius:8px;}
#log{background:var(--entry);border:1px solid var(--border);border-radius:8px;padding:12px;height:55vh;overflow:auto;font-family:Consolas,monospace;font-size:13px;white-space:pre-wrap;}
.l-ok{color:var(--success);} .l-warn{color:var(--warning);} .l-err{color:var(--error);} .l-info{color:var(--accent2);}
.pbar{height:12px;background:var(--entry);border-radius:7px;overflow:hidden;margin-top:12px;border:1px solid var(--border);}
.pbar>div{height:100%;background:var(--accent2);width:0%;transition:width .3s;}
.hide{display:none;}
a{color:var(--accent2);}
.maprow input{margin:0;}
#gridtbl td,#gridtbl th{white-space:nowrap;font-size:11.5px;padding:3px 7px;border:1px solid var(--border);}
#gridtbl .colhdr th{background:var(--sidebar);color:var(--accent2);position:sticky;top:0;text-align:center;}
#gridtbl tr.hrow td{background:#123048;color:var(--accent2);font-weight:700;}
#mapcols select,#mapcols input{padding:7px 8px;font-size:13px;}
</style></head><body>
<div class="hdr"><div style="font-size:26px;">🪨</div>
 <div><h2>Import Lefevre → Odoo</h2><div class="sub">Carrières Maquignon — convertisseur de commandes (web)</div></div></div>
<div class="tabs">
 <div class="tab on" data-t="fichier">📄 Fichier & Import</div>
 <div class="tab" data-t="apercu">👁 Aperçu</div>
 <div class="tab" data-t="journal">📜 Journal</div>
 <div class="tab" data-t="mapping">🗂 Mapping</div>
</div>
<div class="wrap">

<div id="t-fichier">
 <div class="card"><h4>Fichier Lefevre</h4>
  <div class="drop" id="drop">📂 Glissez le fichier Excel ici — ou cliquez pour choisir<input type="file" id="file" accept=".xlsx,.xlsm" style="display:none"/></div>
  <div id="finfo" class="sub" style="margin-top:8px;color:var(--dim);"></div>
 </div>
 <div class="card"><h4>Paramètres commande</h4>
  <div class="grid2">
   <div><label>Référence commande</label><input type="text" id="oref" placeholder="Bon Lefèvre n° …"/></div>
   <div><label>Client (nom Odoo)</label><input type="text" id="ocust" value="LEFEVRE"/></div>
   <div><label>Code client Odoo</label><input type="text" id="ocode" value="LEF001"/></div>
   <div><label>Date commande</label><input type="text" id="odate" value="__TODAY__"/></div>
  </div>
  <div style="margin-top:12px;"><label style="display:inline-flex;align-items:center;gap:8px;cursor:pointer;color:var(--text);">
   <input type="checkbox" id="otaille" style="width:auto;transform:scale(1.3);"/> Taille de pierre (cocher si la commande comporte de la taille → coche le champ « Taille de pierre » et inscrit « taille »)
  </label></div>
  <div class="btns">
   <button class="btn p" id="b-analyse" disabled onclick="analyser()">🔍 Analyser</button>
   <button class="btn s" id="b-export" disabled onclick="exporter()">📥 Exporter Excel</button>
   <button class="btn g" id="b-import" disabled onclick="importer()">🚀 Importer dans Odoo</button>
  </div>
 </div>
 <div class="card hide" id="mapcols"><h4>① Validation du mappage des colonnes</h4>
  <div class="sub" style="margin-bottom:10px;color:var(--dim);">Vérifiez que chaque colonne du fichier pointe vers le bon champ, puis appliquez. Les valeurs devinées sont pré-remplies.</div>
  <div class="grid2" style="grid-template-columns:repeat(3,1fr);">
   <div><label>Feuille</label><select id="m_sheet"></select></div>
   <div><label>Ligne d'en-tête (n° Excel)</label><input type="text" id="m_hrow" placeholder="auto"/></div>
   <div><label>Unité dimensions</label><select id="m_unit"><option value="auto">Auto</option><option value="cm">cm</option><option value="mm">mm</option></select></div>
   <div><label>Réf. pièce (N°)</label><select id="m_ref"></select></div>
   <div><label>Nature (colonne)</label><select id="m_nature"></select></div>
   <div><label>Nature par défaut</label><input type="text" id="m_gnat" placeholder="ex : Tuffeau"/></div>
   <div><label>Quantité</label><select id="m_qte"></select></div>
   <div><label>Longueur</label><select id="m_long"></select></div>
   <div><label>Largeur / Prof.</label><select id="m_prof"></select></div>
   <div><label>Hauteur</label><select id="m_haut"></select></div>
   <div><label>Cube (m³)</label><select id="m_cube"></select></div>
   <div><label>Poids</label><select id="m_poids"></select></div>
   <div><label>Section (colonne)</label><select id="m_section"></select></div>
   <div><label>Sous-section (colonne)</label><select id="m_subsection"></select></div>
  </div>
  <div class="tblbox" style="max-height:34vh;margin-top:12px;"><table id="gridtbl"></table></div>
  <div class="btns"><button class="btn p" onclick="applyMapping()">✅ Appliquer le mappage &amp; ré-analyser</button></div>
 </div>
 <div class="card hide" id="resume"><h4>② Résumé</h4>
  <div id="chantier" style="font-weight:800;font-size:16px;margin-bottom:10px;"></div>
  <div class="kpis" id="kpis"></div>
 </div>
</div>

<div id="t-apercu" class="hide">
 <div class="card"><h4>Aperçu des lignes analysées</h4>
  <div class="tblbox"><table id="tbl"><thead><tr>
   <th>Réf.</th><th>Nature</th><th>Nbr</th><th>Long (m)</th><th>Larg (m)</th><th>Haut (m)</th><th>Vol. m³</th><th>Produit Odoo</th>
  </tr></thead><tbody></tbody></table></div>
 </div>
</div>

<div id="t-journal" class="hide">
 <div class="card"><h4>Journal d'import</h4>
  <div id="log"></div>
  <div class="pbar"><div id="pfill"></div></div>
  <div id="done" style="margin-top:12px;font-size:15px;"></div>
 </div>
</div>

<div id="t-mapping" class="hide">
 <div class="card"><h4>Mapping natures → produits Odoo</h4>
  <div class="tblbox" style="max-height:50vh;"><table id="maptbl"><thead><tr>
   <th>Nature Lefevre</th><th>Code produit Odoo</th><th>Libellé</th><th></th>
  </tr></thead><tbody></tbody></table></div>
  <div class="btns">
   <button class="btn s" onclick="mapAdd()">➕ Ajouter</button>
   <button class="btn p" onclick="mapSave()">💾 Enregistrer le mapping</button>
  </div>
  <div id="mapmsg" class="sub" style="margin-top:8px;color:var(--dim);"></div>
 </div>
</div>

</div>
<script>
var TOKEN = new URLSearchParams(location.search).get('token') || '';
var FILE = null, ROWS = null, HDR = null, CFG = null;

document.querySelectorAll('.tab').forEach(function(t){
  t.onclick = function(){
    document.querySelectorAll('.tab').forEach(function(x){ x.classList.remove('on'); });
    t.classList.add('on');
    ['fichier','apercu','journal','mapping'].forEach(function(n){
      document.getElementById('t-'+n).classList.toggle('hide', n !== t.dataset.t);
    });
    if(t.dataset.t === 'mapping' && !CFG){ mapLoad(); }
  };
});

var drop = document.getElementById('drop'), finput = document.getElementById('file');
drop.onclick = function(){ finput.click(); };
['dragover','dragenter'].forEach(function(e){ drop.addEventListener(e, function(ev){ ev.preventDefault(); drop.classList.add('hov'); }); });
['dragleave','drop'].forEach(function(e){ drop.addEventListener(e, function(ev){ ev.preventDefault(); drop.classList.remove('hov'); }); });
drop.addEventListener('drop', function(ev){ if(ev.dataTransfer.files.length){ setFile(ev.dataTransfer.files[0]); } });
finput.onchange = function(){ if(finput.files.length){ setFile(finput.files[0]); } };
function setFile(f){
  FILE = f;
  document.getElementById('finfo').textContent = '📄 ' + f.name + ' (' + Math.round(f.size/1024) + ' Ko)';
  document.getElementById('b-analyse').disabled = false;
  var m = f.name.match(/n[°o]?\s*(\d{4,6})/i);
  if(m && !document.getElementById('oref').value){ document.getElementById('oref').value = 'Bon Lefèvre n° ' + m[1]; }
  analyser();
}

var GRID = null, NCOLS = 0, GRID_HEADERS = [];
var MAPFIELDS = ['ref','nature','qte','long','prof','haut','cube','poids','section','subsection'];

function analyser(mappingOverride){
  if(!FILE){ return; }
  var fd = new FormData();
  fd.append('file', FILE); fd.append('token', TOKEN);
  var ms = document.getElementById('m_sheet');
  var sheet = (ms && ms.value) ? ms.value : '';
  if(sheet){ fd.append('sheet', sheet); }
  if(mappingOverride){ fd.append('mapping', JSON.stringify(mappingOverride)); }
  document.getElementById('finfo').textContent = '⏳ Analyse en cours…';
  fetch('/import-lefevre/parse', {method:'POST', body: fd}).then(function(r){ return r.json(); }).then(function(d){
    if(d.error){ document.getElementById('finfo').textContent = '❌ ' + d.error; return; }
    ROWS = d.rows; HDR = d.header_info;
    GRID = HDR.grid || []; NCOLS = HDR.ncols || 0;
    document.getElementById('finfo').textContent =
      '✓ ' + FILE.name + ' — feuille "' + HDR.sheet_used + '"';
    renderMapper();
    document.getElementById('mapcols').classList.remove('hide');
    document.getElementById('resume').classList.remove('hide');
    document.getElementById('chantier').textContent = '🏗 ' + (HDR.chantier || 'Chantier non détecté') +
      (HDR.code_chantier ? '  ·  code ' + HDR.code_chantier : '');
    var k = d.stats;
    document.getElementById('kpis').innerHTML =
      kpi('Sections', k.sections) + kpi('Sous-sections', k.subsections) + kpi('Lignes', k.lines) +
      kpi('Pièces', k.pieces) + kpi('Cubage m³', k.cubage) +
      kpi('Non mappés', k.unknown, k.unknown > 0);
    document.getElementById('b-export').disabled = false;
    document.getElementById('b-import').disabled = (k.lines === 0);
    fillPreview();
  }).catch(function(e){ document.getElementById('finfo').textContent = '❌ ' + e; });
}

function colLabel(c){
  var letter = String.fromCharCode(65 + c);
  var t = GRID_HEADERS[c] || '';
  return t ? (letter + ' · ' + t) : ('Col ' + letter);
}
function computeHeaders(){
  GRID_HEADERS = [];
  var hr = HDR.mapping.header_row; if(hr === null || hr === undefined){ hr = 0; }
  for(var c=0; c<NCOLS; c++){
    var t = '';
    for(var rr=Math.max(0,hr-1); rr<=hr && rr<GRID.length; rr++){
      var v = (GRID[rr] && GRID[rr][c]) ? GRID[rr][c] : '';
      if(v){ t = (t ? t + ' ' : '') + v; }
    }
    GRID_HEADERS.push(t.length > 22 ? t.slice(0,22) : t);
  }
}
function fillColSel(id, val){
  var sel = document.getElementById('m_' + id);
  var h = '<option value="">— aucune —</option>';
  for(var c=0; c<NCOLS; c++){ h += '<option value="' + c + '">' + esc(colLabel(c)) + '</option>'; }
  sel.innerHTML = h;
  sel.value = (val === null || val === undefined) ? '' : String(val);
}
function renderMapper(){
  computeHeaders();
  var mp = HDR.mapping || {cols:{}};
  var cols = mp.cols || {};
  // feuilles
  var ms = document.getElementById('m_sheet'), opts = '';
  (HDR.all_sheets || []).forEach(function(s){
    opts += '<option value="' + esc(s) + '"' + (s===HDR.sheet_used?' selected':'') + '>' + esc(s) + '</option>';
  });
  if((HDR.data_sheets || []).length > 1){
    opts += '<option value="__ALL__"' + (HDR.sheet_used==='__ALL__'?' selected':'') + '>➕ Toutes les feuilles (fusionner)</option>';
  }
  ms.innerHTML = opts;
  document.getElementById('m_hrow').value = (mp.header_row === null || mp.header_row === undefined) ? '' : (mp.header_row + 1);
  document.getElementById('m_unit').value = mp.unit || 'auto';
  document.getElementById('m_gnat').value = mp.global_nature || '';
  MAPFIELDS.forEach(function(f){ fillColSel(f, cols[f]); });
  renderGrid();
}
function renderGrid(){
  var hr = HDR.mapping.header_row;
  var t = document.getElementById('gridtbl');
  var h = '<tr class="colhdr"><th>#</th>';
  for(var c=0; c<NCOLS; c++){ h += '<th>' + String.fromCharCode(65+c) + '</th>'; }
  h += '</tr>';
  for(var r=0; r<GRID.length; r++){
    h += '<tr' + (r===hr?' class="hrow"':'') + '><td style="color:var(--dim)">' + (r+1) + '</td>';
    for(var c2=0; c2<NCOLS; c2++){ h += '<td>' + esc((GRID[r] && GRID[r][c2]) || '') + '</td>'; }
    h += '</tr>';
  }
  t.innerHTML = h;
}
function applyMapping(){
  var cols = {};
  MAPFIELDS.forEach(function(f){
    var v = document.getElementById('m_' + f).value;
    cols[f] = (v === '') ? null : parseInt(v, 10);
  });
  var hrow = document.getElementById('m_hrow').value.trim();
  var mp = {
    cols: cols,
    header_row: hrow === '' ? null : (parseInt(hrow, 10) - 1),
    unit: document.getElementById('m_unit').value,
    global_nature: document.getElementById('m_gnat').value.trim()
  };
  analyser(mp);
}
function kpi(lbl, val, warn){
  return '<div class="kpi' + (warn?' warn':'') + '"><b>' + val + '</b>' + lbl + '</div>';
}
function fillPreview(){
  var tb = document.querySelector('#tbl tbody'); tb.innerHTML = '';
  ROWS.forEach(function(r){
    var tr = document.createElement('tr');
    if(r.type === 'section'){ tr.className = 'sec'; tr.innerHTML = '<td colspan="8">' + esc(r.label) + '</td>'; }
    else if(r.type === 'subsection'){ tr.className = 'sub'; tr.innerHTML = '<td colspan="8">&nbsp;&nbsp;' + esc(r.label) + '</td>'; }
    else {
      if(!r.mapped){ tr.className = 'unmapped'; }
      tr.innerHTML = '<td>' + esc(r.ref) + '</td><td>' + esc(r.nature) + '</td><td>' + r.qte +
        '</td><td>' + n3(r.long_m) + '</td><td>' + n3(r.prof_m) + '</td><td>' + n3(r.haut_m) +
        '</td><td>' + (r.qty_m3 || '') + '</td><td>' + esc(r.product) + '</td>';
    }
    tb.appendChild(tr);
  });
}
function n3(v){ return (v === null || v === undefined) ? '' : v.toFixed(3); }
function esc(s){ return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;'); }

function payload(){
  return { token: TOKEN, rows: ROWS,
    order_ref: document.getElementById('oref').value,
    customer: document.getElementById('ocust').value,
    customer_code: document.getElementById('ocode').value,
    order_date: document.getElementById('odate').value,
    taille: document.getElementById('otaille').checked };
}
function exporter(){
  fetch('/import-lefevre/export', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(payload())}).then(function(r){ return r.blob(); }).then(function(b){
      var a = document.createElement('a');
      a.href = URL.createObjectURL(b);
      a.download = 'Import_Odoo.xlsx'; a.click();
    });
}
function importer(){
  var nb = ROWS.filter(function(r){ return r.type === 'line'; }).length;
  if(!confirm('Importer ' + nb + ' lignes de pierres dans Odoo ?\n\nUn devis (brouillon) sera créé.')){ return; }
  document.querySelector('[data-t=journal]').click();
  document.getElementById('log').innerHTML = '';
  document.getElementById('done').innerHTML = '';
  document.getElementById('b-import').disabled = true;
  fetch('/import-lefevre/import', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(payload())}).then(function(r){ return r.json(); }).then(function(d){
      if(d.error){ addLog('❌ ' + d.error, 'err'); return; }
      poll(d.sid, 0);
    });
}
function poll(sid, after){
  fetch('/import-lefevre/status?token=' + encodeURIComponent(TOKEN) + '&sid=' + sid + '&after=' + after)
    .then(function(r){ return r.json(); }).then(function(d){
      (d.logs || []).forEach(function(l){ addLog(l.msg, l.tag); });
      document.getElementById('pfill').style.width = (d.progress || 0) + '%';
      if(d.done){
        document.getElementById('b-import').disabled = false;
        if(d.order_url){
          document.getElementById('done').innerHTML =
            '✅ Devis <b>' + (d.order_name || d.order_id) + '</b> créé — <a href="' + d.order_url + '" target="_blank">ouvrir dans Odoo</a>';
        }
        return;
      }
      setTimeout(function(){ poll(sid, d.next); }, 900);
    }).catch(function(){ setTimeout(function(){ poll(sid, after); }, 2000); });
}
function addLog(msg, tag){
  var d = document.getElementById('log');
  var s = document.createElement('div');
  s.className = 'l-' + (tag || 'info'); s.textContent = msg;
  d.appendChild(s); d.scrollTop = d.scrollHeight;
}

function mapLoad(){
  fetch('/import-lefevre/config?token=' + encodeURIComponent(TOKEN))
    .then(function(r){ return r.json(); }).then(function(d){
      CFG = d; mapRender();
      var fc = d.fields_check || {};
      var bad = Object.keys(fc).filter(function(k){ return fc[k] === false; });
      document.getElementById('mapmsg').textContent = bad.length
        ? '⚠ Champs Studio absents sur sale.order.line : ' + bad.join(', ')
        : '✓ Champs Studio vérifiés — tous présents.';
    });
}
function mapRender(){
  var tb = document.querySelector('#maptbl tbody'); tb.innerHTML = '';
  Object.keys(CFG.nature_mapping).forEach(function(nat){
    var m = CFG.nature_mapping[nat];
    var tr = document.createElement('tr'); tr.className = 'maprow';
    tr.innerHTML = '<td><input type="text" value="' + esc(nat) + '"/></td>' +
      '<td><input type="text" value="' + esc(m.product_code) + '"/></td>' +
      '<td><input type="text" value="' + esc(m.product) + '"/></td>' +
      '<td><button class="btn s" onclick="this.closest(\'tr\').remove()">🗑</button></td>';
    tb.appendChild(tr);
  });
}
function mapAdd(){
  var tb = document.querySelector('#maptbl tbody');
  var tr = document.createElement('tr'); tr.className = 'maprow';
  tr.innerHTML = '<td><input type="text" placeholder="Nature"/></td>' +
    '<td><input type="text" placeholder="CODE-PS"/></td>' +
    '<td><input type="text" placeholder="[CODE] Libellé produit"/></td>' +
    '<td><button class="btn s" onclick="this.closest(\'tr\').remove()">🗑</button></td>';
  tb.appendChild(tr);
}
function mapSave(){
  var nm = {};
  document.querySelectorAll('#maptbl tbody tr').forEach(function(tr){
    var ins = tr.querySelectorAll('input');
    var nat = ins[0].value.trim();
    if(!nat){ return; }
    var code = ins[1].value.trim().replace(/^\[|\]$/g, '');
    var mcode = code.match(/\[([^\]]+)\]/); if(mcode){ code = mcode[1]; }
    nm[nat] = { product_code: code, product: ins[2].value.trim() || ('[' + code + ']') };
  });
  fetch('/import-lefevre/config', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({token: TOKEN, nature_mapping: nm})})
    .then(function(r){ return r.json(); }).then(function(d){
      document.getElementById('mapmsg').textContent = d.ok ? '✓ Mapping enregistré (dans Odoo).' : '❌ ' + (d.error || 'erreur');
      CFG.nature_mapping = nm;
    });
}
</script></body></html>
"""
