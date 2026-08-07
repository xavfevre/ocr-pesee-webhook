"""
Planning chauffeur PROTEC + app mobile « Ma tournée » + Fiche de fin de travaux
Monté comme Blueprint sous /protec dans l'app principale (même service Render
que le webhook Maquignon) — ou exécutable seul pour les tests.

- /protec/                 : planning semaine (grille jours x chauffeurs)
- /protec/mois             : planning mois (bandeaux par semaine)
- /protec/ma-tournee       : page mobile chauffeur (lien signé HMAC)
- /protec/fiche/<slot_id>  : fiche de fin de travaux (formulaire mobile)
- /protec/liens            : page admin des liens chauffeurs

Données : Odoo v19 SaaS PROTEC via XML-RPC (planning.slot).
Env vars : PROTEC_ODOO_USER, PROTEC_ODOO_PASSWORD, PROTEC_PLANNING_SECRET
"""
import os, json, time, hmac, hashlib, socket, re
import urllib.request, http.cookiejar
import xmlrpc.client
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo
from flask import Blueprint, Flask, render_template, request, redirect, url_for, abort, Response

bp = Blueprint("protec", __name__, template_folder="templates")
socket.setdefaulttimeout(25)

ODOO_URL      = os.environ.get("PROTEC_ODOO_URL",      "https://protec-s3t.odoo.com")
ODOO_DB       = os.environ.get("PROTEC_ODOO_DB",       "protec-s3t")
ODOO_USER     = os.environ.get("PROTEC_ODOO_USER",     "")
ODOO_PASSWORD = os.environ.get("PROTEC_ODOO_PASSWORD", "")
SECRET        = os.environ.get("PROTEC_PLANNING_SECRET", "")

TZ  = ZoneInfo("Europe/Paris")
UTC = ZoneInfo("UTC")

# Employés à exclure (comptes techniques)
EXCLUDED_EMPLOYEE_IDS = {1, 3, 8, 9}  # compte technique + Manon + Natacha (hors terrain)
CHAUFFEUR_TAG_ID = 1  # étiquette RH « CHAUFFEUR OPÉRATEUR » — seuls ces employés sont listés

# Palette couleurs vives par chauffeur (stable sur l'id employé)
PALETTE = ["#2563eb", "#16a34a", "#ea580c", "#9333ea", "#0d9488",
           "#db2777", "#ca8a04", "#4f46e5", "#dc2626", "#0891b2"]
# Couleurs et ordre fixes, identiques aux pages planning du site
CMAP = {4: "#2563eb", 5: "#eab308", 6: "#dc2626", 7: "#111827",
        10: "#7c3aed", 2: "#7c3aed", 8: "#0d9488", 1: "#db2777", 9: "#ea580c"}
OMAP = {4: 1, 5: 2, 6: 3, 7: 4, 10: 5, 2: 6, 8: 7, 9: 8, 1: 9}

# ─── Fiche de fin de travaux : natures de travaux par onglet ────────────────
# (les codes existants sont conservés → compatibilité des fiches déjà remplies)
TRAVAUX_TABS = [
    ("deb", "Débouchage", [
        ("deb_brancht_eu", "Débouchage branchement EU", "U"),
        ("deb_brancht_ep", "Débouchage branchement EP", "U"),
        ("deb_reseau_eu",  "Débouchage réseau EU",      "U"),
        ("deb_reseau_ep",  "Débouchage réseau EP",      "U"),
    ]),
    ("curage", "Curage", [
        ("hydro_reu",       "Hydrocurage REU",         "ML"),
        ("hydro_rep",       "Hydrocurage REP",         "ML"),
        ("hydro_reu_amiante", "Hydrocurage REU amianté", "ML"),
        ("pomp_siphon",     "Curage siphon",           "U"),
    ]),
    ("poste", "Pompage", [
        ("pomp_poste",      "Poste de relevage",   "U"),
        ("pomp_puisard",    "Puisard",             "U"),
        ("pomp_avaloir",    "Avaloir",             "U"),
        ("pomp_regard",     "Regard de visite",    "U"),
        ("travaux_pompage", "Travaux de pompage",  "H"),
        ("pomp_bac_graisse", "Bac à graisses",        "U"),
        ("pomp_bac_mousse",  "Bac à mousse",          "U"),
    ]),
    ("step", "STEP", [
        ("pomp_dessableur",  "Pompage nettoyage dessableur",       "U"),
        ("pomp_decanteur",   "Pompage nettoyage décanteur",        "U"),
        ("pomp_degrilleur",  "Pompage nettoyage panier dégrilleur", "U"),
    ]),
    ("fosse", "Fosse", [
        ("fosse_septique",    "Pompage nettoyage fosse septique",     "U"),
        ("fosse_toutes_eaux", "Pompage nettoyage fosse toutes eaux",  "U"),
        ("fosse_etanche",     "Pompage nettoyage fosse étanche",      "U"),
        ("microstation",      "Pompage nettoyage micro-station (jusqu'à 3 m3)", "U"),
        ("m3_supp",           "M3 supplémentaire(s) au-delà de 3 m3", "m3"),
        ("descente_fosse",    "Descente opérateur en fosse",          "U"),
        ("desinfection_fosse","Désinfection fosse",                   "m3"),
        ("lavage_filtre",     "Lavage filtre fosse",                  "U"),
        ("lavage_filtre_complet", "Lavage complet filtre fosse",      "U"),
        ("filtre_coco",       "Changement de filtre coco",            "H"),
        ("visite_fosse",      "Visite contrôle entretien fosse",      "U"),
        ("colonne_asp",       "Pose dépose colonne d'aspiration",     "U"),
    ]),
    ("itv", "ITV", [
        ("itv_inspection", "Inspection télévisée",              "ML"),
        ("itv_branchements", "Branchements",                    "U"),
        ("itv_fourgon",    "Prise en charge fourgon ITV / tests", "H"),
        ("itv_tests",      "Tests",                             "H"),
        ("itv_rech_fuite", "Recherche de fuite",                "H"),
        ("itv_rech_reseau","Recherche de réseau",               "H"),
        ("itv_rapport",    "Rapport",                           "U"),
        ("aerogommage",    "Aérogommage",                       "m²"),
    ]),
    ("autre", "Autre", [
        ("main_oeuvre",   "Main d'œuvre",                    "H"),
        ("lavage_mat",    "Lavage matériel",                 "U"),
        ("mad_materiel",  "Mise à disposition de matériel",  "U"),
        ("deratisation",  "Dératisation",                    "U"),
        ("autre",         "Autre (préciser dans commentaires)", "H"),
    ]),
]
# Liste à plat (POST + chatter) — compatibilité avec le code existant
TRAVAUX = [row for _k, _lbl, rows in TRAVAUX_TABS for row in rows]
DECHETS = [
    ("sable",   "Sable"),
    ("graisse", "Graisse"),
    ("refus",   "Refus dégrillage"),
    ("vidange", "Matières de vidanges"),
    ("autre",   "Autre"),
]
# Destinations possibles des déchets (liste déroulante de la fiche)
DESTINATIONS = ["Assainissement", "Châtellerault", "Chinon",
                "Dépotage sur place", "Loches", "Metha Ingrandes", "Saché"]

@bp.app_context_processor
def inject_base():
    # "/protec" quand monté sous prefix, "" en exécution seule
    return {"BASE": url_for("protec.week_view").rstrip("/")}

# ─── Helpers Odoo ─────────────────────────────────────────────────────────────
_uid_cache = {"uid": None}

def odoo_connect():
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    if _uid_cache["uid"]:
        return _uid_cache["uid"], models
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
    if not uid:
        raise ValueError("Authentification Odoo échouée")
    _uid_cache["uid"] = uid
    return uid, models

def x(models, uid, model, method, *params, **kw):
    return models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, model, method, list(params), kw)

# Base de test (validation des tarifs clients avant bascule en production)
TEST_ODOO_URL = os.environ.get("PROTEC_TEST_ODOO_URL", "https://testprotec070826.odoo.com")
TEST_ODOO_DB  = os.environ.get("PROTEC_TEST_ODOO_DB",  "testprotec070826")
_uid_cache_test = {"uid": None}

def odoo_connect_src(src=""):
    """Connexion Odoo production (défaut) ou base de test (src='test').
    Renvoie (uid, models, db) — utiliser avec xq() qui prend le db explicite."""
    if src != "test":
        uid, models = odoo_connect()
        return uid, models, ODOO_DB
    models = xmlrpc.client.ServerProxy(f"{TEST_ODOO_URL}/xmlrpc/2/object")
    if not _uid_cache_test["uid"]:
        common = xmlrpc.client.ServerProxy(f"{TEST_ODOO_URL}/xmlrpc/2/common")
        uid = common.authenticate(TEST_ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
        if not uid:
            raise ValueError("Authentification Odoo (base test) échouée")
        _uid_cache_test["uid"] = uid
    return _uid_cache_test["uid"], models, TEST_ODOO_DB

def xq(db, models, uid, model, method, *params, **kw):
    return models.execute_kw(db, uid, ODOO_PASSWORD, model, method, list(params), kw)

def utc_to_local(dt_str):
    if not dt_str:
        return None
    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=UTC)
    return dt.astimezone(TZ)

def local_range_to_utc(d_start: date, d_end: date):
    """[d_start 00:00, d_end 00:00) heure de Paris → chaînes UTC Odoo."""
    s = datetime(d_start.year, d_start.month, d_start.day, tzinfo=TZ).astimezone(UTC)
    e = datetime(d_end.year, d_end.month, d_end.day, tzinfo=TZ).astimezone(UTC)
    return s.strftime("%Y-%m-%d %H:%M:%S"), e.strftime("%Y-%m-%d %H:%M:%S")

def monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())

def emp_color(emp_id: int) -> str:
    if not emp_id:
        return "#64748b"
    return CMAP.get(emp_id, PALETTE[emp_id % len(PALETTE)])

MOIS_FR  = ["janvier","février","mars","avril","mai","juin","juillet","août",
            "septembre","octobre","novembre","décembre"]
MOIS_ABR = ["jan","fév","mar","avr","mai","juin","juil","août","sep","oct","nov","déc"]
JOURS_FR = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
JOURS_LONG = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]

# ─── Sécurité HMAC ────────────────────────────────────────────────────────────
def _sign(payload: str) -> str:
    return hmac.new(SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:20]

def driver_sig(emp_id: int) -> str:
    return _sign(f"driver:{emp_id}")

def fiche_token(slot_id: int) -> str:
    return _sign(f"fdt:{slot_id}")

def check_sig(payload: str, sig: str) -> bool:
    return bool(SECRET) and hmac.compare_digest(_sign(payload), sig or "")

# ─── Données ──────────────────────────────────────────────────────────────────
SLOT_FIELDS = ["name", "start_datetime", "end_datetime",
               "partner_id", "partner_city", "partner_zip", "partner_street",
               "partner_phone", "employee_ids", "role_id", "x_fdt_fait"]

_emp_cache = {"t": 0.0, "data": None}

def get_all_employees(uid, models):
    now = time.time()
    if _emp_cache["data"] is not None and now - _emp_cache["t"] < 600:
        return _emp_cache["data"]
    emps = x(models, uid, "hr.employee", "search_read",
        [["active", "=", True], ["category_ids", "in", [CHAUFFEUR_TAG_ID]]],
        fields=["id", "name"], order="name asc", limit=100)
    data = sorted([e for e in emps if e["id"] not in EXCLUDED_EMPLOYEE_IDS],
                  key=lambda e: (OMAP.get(e["id"], 99), e["name"]))
    _emp_cache["t"], _emp_cache["data"] = now, data
    return data

def emp_name_map(uid, models):
    return {e["id"]: e["name"] for e in get_all_employees(uid, models)}

_veh_cache = {"t": 0.0, "data": None}

def get_fleet_vehicles(uid, models):
    """Noms des camions du parc (format « Surnom (Plaque) »), pour la liste
    déroulante véhicule de la fiche. Tolérant aux erreurs d'accès."""
    now = time.time()
    if _veh_cache["data"] is not None and now - _veh_cache["t"] < 600:
        return _veh_cache["data"]
    try:
        vs = x(models, uid, "fleet.vehicle", "search_read",
               [["active", "=", True]], fields=["name"], order="name asc", limit=200)
        data = [v["name"] for v in vs if v.get("name")]
    except Exception:
        data = []
    _veh_cache["t"], _veh_cache["data"] = now, data
    return data

def fetch_slots(uid, models, d_start: date, d_end: date, emp_id=None):
    s_utc, e_utc = local_range_to_utc(d_start, d_end)
    domain = [["start_datetime", ">=", s_utc], ["start_datetime", "<", e_utc]]
    if emp_id:
        domain.append(["employee_ids", "in", [emp_id]])
    return x(models, uid, "planning.slot", "search_read", domain,
             fields=SLOT_FIELDS, order="start_datetime asc", limit=2000)

def slot_to_card(s):
    sl = utc_to_local(s["start_datetime"])
    el = utc_to_local(s["end_datetime"])
    return {
        "id":     s["id"],
        "ref":    (s["name"] or "").strip(),
        "day":    sl.date().isoformat(),
        "start":  sl.strftime("%H:%M"),
        "end":    el.strftime("%H:%M") if el else "",
        "client": s["partner_id"][1] if s["partner_id"] else "—",
        "street": s.get("partner_street") or "",
        "zip":    s.get("partner_zip") or "",
        "ville":  s.get("partner_city") or "",
        "tel":    (s.get("partner_phone") or "").strip(),
        "role":   s["role_id"][1] if s.get("role_id") else "",
        "fdt":    bool(s.get("x_fdt_fait")),
        "emp_ids": [e for e in s.get("employee_ids", []) if e not in EXCLUDED_EMPLOYEE_IDS],
        "token":  fiche_token(s["id"]),
    }

def build_matrix(slots, columns):
    """matrix[(day_iso, emp_id)] = [cards] ; emp 0 = non assigné."""
    matrix, counts = {}, {}
    col_ids = {c["id"] for c in columns}
    for s in slots:
        card = slot_to_card(s)
        targets = [e for e in card["emp_ids"] if e in col_ids] or [0]
        for eid in targets:
            if eid not in col_ids:
                continue
            matrix.setdefault((card["day"], eid), []).append(card)
            counts[eid] = counts.get(eid, 0) + 1
    for c in columns:
        c["total"] = counts.get(c["id"], 0)
    return matrix

def make_columns(all_emps, emp_filter=None, with_unassigned=True):
    cols = [{"id": e["id"], "name": e["name"], "color": emp_color(e["id"])}
            for e in all_emps if not emp_filter or e["id"] == emp_filter]
    if with_unassigned and not emp_filter:
        cols.append({"id": 0, "name": "Non assigné", "color": "#ef4444"})
    return cols

def week_label(ws: date) -> str:
    we = ws + timedelta(days=5)
    num = ws.isocalendar()[1]
    if ws.month == we.month:
        return f"S{num} — {ws.day} au {we.day} {MOIS_ABR[we.month-1]} {we.year}"
    return f"S{num} — {ws.day} {MOIS_ABR[ws.month-1]} au {we.day} {MOIS_ABR[we.month-1]} {we.year}"

def common_nav(emp_filter, all_emps):
    return {
        "emp_filter": emp_filter or 0,
        "drivers": [{"id": e["id"], "name": e["name"], "color": emp_color(e["id"])} for e in all_emps],
    }

# ─── VUE SEMAINE ──────────────────────────────────────────────────────────────
@bp.route("/")
def week_view():
    week_str = request.args.get("week")
    emp_filter = request.args.get("c", type=int)
    try:
        ws = datetime.strptime(week_str, "%Y-%m-%d").date() if week_str else date.today()
    except ValueError:
        ws = date.today()
    week_start = monday_of(ws)
    days = [week_start + timedelta(days=i) for i in range(6)]  # Lun → Sam

    error, columns, matrix = None, [], {}
    all_emps = []
    try:
        uid, models = odoo_connect()
        all_emps = get_all_employees(uid, models)
        columns = make_columns(all_emps, emp_filter)
        slots = fetch_slots(uid, models, week_start, week_start + timedelta(days=7), emp_filter)
        matrix = build_matrix(slots, columns)
    except Exception as e:
        error = str(e)

    day_labels = [{"iso": d.isoformat(),
                   "dow": JOURS_FR[d.weekday()], "num": f"{d.day}/{d.month}",
                   "is_today": d == date.today()} for d in days]

    qs_c = f"&c={emp_filter}" if emp_filter else ""
    return render_template("week.html",
        days=day_labels, columns=columns, matrix=matrix,
        cur_week=week_start.isoformat(),
        week_label=week_label(week_start),
        prev_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        this_week=monday_of(date.today()).isoformat(),
        is_current=(week_start == monday_of(date.today())),
        month_str=f"{week_start.year}-{week_start.month:02d}",
        qs_c=qs_c, error=error,
        total_slots=sum(len(v) for v in matrix.values()),
        **common_nav(emp_filter, all_emps))

# ─── VUE MOIS (bandeaux par semaine) ─────────────────────────────────────────
@bp.route("/mois")
def month_view():
    m_str = request.args.get("m")
    emp_filter = request.args.get("c", type=int)
    today = date.today()
    try:
        y, m = (int(v) for v in m_str.split("-")) if m_str else (today.year, today.month)
    except (ValueError, AttributeError):
        y, m = today.year, today.month
    first = date(y, m, 1)
    last = (date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)) - timedelta(days=1)

    # Semaines couvrant le mois
    weeks = []
    wcur = monday_of(first)
    while wcur <= last:
        weeks.append(wcur)
        wcur += timedelta(days=7)

    error, columns, matrix = None, [], {}
    all_emps = []
    try:
        uid, models = odoo_connect()
        all_emps = get_all_employees(uid, models)
        columns = make_columns(all_emps, emp_filter)
        slots = fetch_slots(uid, models, weeks[0], weeks[-1] + timedelta(days=7), emp_filter)
        matrix = build_matrix(slots, columns)
    except Exception as e:
        error = str(e)

    week_blocks = []
    for wstart in weeks:
        days = []
        for i in range(6):  # Lun → Sam
            d = wstart + timedelta(days=i)
            days.append({"iso": d.isoformat(), "dow": JOURS_FR[d.weekday()],
                         "num": f"{d.day}/{d.month}",
                         "in_month": d.month == m, "is_today": d == today})
        total = sum(len(matrix.get((dd["iso"], c["id"]), [])) for dd in days for c in columns)
        week_blocks.append({"label": week_label(wstart), "days": days, "total": total,
                            "week_iso": wstart.isoformat()})

    prev_m = date(y - 1, 12, 1) if m == 1 else date(y, m - 1, 1)
    next_m = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
    qs_c = f"&c={emp_filter}" if emp_filter else ""

    return render_template("mois.html",
        weeks=week_blocks, columns=columns, matrix=matrix,
        month_label=f"{MOIS_FR[m-1].capitalize()} {y}",
        month_str=f"{y}-{m:02d}",
        prev_month=f"{prev_m.year}-{prev_m.month:02d}",
        next_month=f"{next_m.year}-{next_m.month:02d}",
        this_month=f"{today.year}-{today.month:02d}",
        is_current=(y == today.year and m == today.month),
        this_week=monday_of(today).isoformat(),
        qs_c=qs_c, error=error,
        total_slots=sum(len(v) for v in matrix.values()),
        **common_nav(emp_filter, all_emps))

# ─── MA TOURNÉE (mobile chauffeur) ───────────────────────────────────────────
@bp.route("/ma-tournee")
def ma_tournee():
    emp_id = request.args.get("c", type=int)
    sig = request.args.get("s", "")
    if not emp_id or not check_sig(f"driver:{emp_id}", sig):
        abort(403)

    error, emp_name, dayblocks = None, "", []
    try:
        uid, models = odoo_connect()
        emp_name = emp_name_map(uid, models).get(emp_id) or f"Chauffeur {emp_id}"

        today = date.today()
        # Fenêtre : 15 jours avant → 15 jours après (borne de fin exclusive)
        slots = fetch_slots(uid, models, today - timedelta(days=15), today + timedelta(days=16), emp_id)
        by_day = {}
        cards = []
        for s in slots:
            card = slot_to_card(s)
            by_day.setdefault(card["day"], []).append(card)
            cards.append(card)

        # Lieu d'intervention / interlocuteur / objet depuis la commande liée
        refs = list({c["ref"].split()[0] for c in cards if c["ref"]})
        somap = {}
        if refs:
            sos = x(models, uid, "sale.order", "search_read",
                [["name", "in", refs]],
                fields=["name", "x_studio_lieu_dintervention_2",
                        "x_studio_interlocuteur_1", "x_studio_lieu_dintervention_1"])
            somap = {o["name"]: o for o in sos}
        # camion renseigné sur le BI lié (origine = référence commande)
        bimap = {}
        if refs:
            try:
                picks = x(models, uid, "stock.picking", "search_read",
                          [["origin", "in", refs], ["x_camion_id", "!=", False]],
                          fields=["origin", "x_camion_id"])
                bimap = {p["origin"]: p["x_camion_id"][1] for p in picks}
            except Exception:
                pass
        for c in cards:
            so = somap.get(c["ref"].split()[0]) if c["ref"] else None
            c["lieu"] = (so or {}).get("x_studio_lieu_dintervention_2") or ""
            c["inter"] = (so or {}).get("x_studio_interlocuteur_1") or ""
            c["objet"] = (so or {}).get("x_studio_lieu_dintervention_1") or ""
            c["camion"] = bimap.get(c["ref"].split()[0], "") if c["ref"] else ""

        # Toujours afficher aujourd'hui même vide
        all_days = sorted(set(by_day.keys()) | {today.isoformat()})
        for iso in all_days:
            d = date.fromisoformat(iso)
            dayblocks.append({
                "iso": iso,
                "label": f"{JOURS_LONG[d.weekday()]} {d.day} {MOIS_ABR[d.month-1]}",
                "is_today": d == today,
                "cards": by_day.get(iso, []),
            })
    except Exception as e:
        error = str(e)

    return render_template("tournee.html",
        emp_name=emp_name, emp_id=emp_id, sig=sig,
        dayblocks=dayblocks, today_iso=date.today().isoformat(), error=error)

# ─── FICHE DE FIN DE TRAVAUX ─────────────────────────────────────────────────
FDT_FIELDS = ["x_fdt_fait", "x_fdt_date", "x_fdt_vehicule", "x_fdt_heure_arrivee",
              "x_fdt_heure_depart", "x_fdt_temps_trajet", "x_fdt_operateurs",
              "x_fdt_commentaires", "x_fdt_data", "x_fdt_signataire",
              "x_fdt_prochain_entretien"]

@bp.route("/fiche/<int:slot_id>", methods=["GET", "POST"])
def fiche(slot_id):
    token = request.args.get("t", "")
    if not check_sig(f"fdt:{slot_id}", token):
        abort(403)

    uid, models = odoo_connect()
    recs = x(models, uid, "planning.slot", "read", [slot_id],
             fields=SLOT_FIELDS + FDT_FIELDS + ["partner_street2"])
    if not recs:
        abort(404)
    s = recs[0]
    card = slot_to_card(s)

    emp_names = ""
    if card["emp_ids"]:
        nmap = emp_name_map(uid, models)
        emp_names = ", ".join(nmap.get(i, "") for i in card["emp_ids"] if nmap.get(i))

    if request.method == "POST":
        f = request.form
        data = {"travaux": {}, "dechets": {}}
        for code, _label, _unit in TRAVAUX:
            qte = f.get(f"trav_{code}_qte", "").strip()
            temps = f.get(f"trav_{code}_temps", "").strip()
            if qte or temps:
                data["travaux"][code] = {"qte": qte, "temps": temps}
        for code, _label in DECHETS:
            row = {
                "volume":       f.get(f"dech_{code}_vol", "").strip(),
                "destination":  f.get(f"dech_{code}_dest", "").strip(),
                "temps":        f.get(f"dech_{code}_temps", "").strip(),
            }
            if any(row.values()):
                data["dechets"][code] = row

        # Tranches de curage (lignes dynamiques rue par rue)
        data["curage_commune"] = f.get("curage_commune", "").strip()
        curage_rows = []
        idxs = sorted({int(mo.group(1)) for k in f.keys()
                       if (mo := re.match(r"curage_rue_(\d+)$", k))})
        for i in idxs:
            rue = f.get(f"curage_rue_{i}", "").strip()
            longueur = f.get(f"curage_longueur_{i}", "").strip()
            if not rue and not longueur:
                continue
            curage_rows.append({
                "rue": rue,
                "longueur": longueur,
                "amiante": bool(f.get(f"curage_amiante_{i}")),
                "probleme": f.get(f"curage_probleme_{i}", "").strip(),
            })
        data["curage"] = curage_rows

        def bin_val(field_val):
            if field_val and field_val.startswith("data:"):
                return field_val.split(",", 1)[1]
            return None

        # « pause » : tout est sauvegardé mais la fiche reste à reprendre
        # (pas de badge FDT, pas de message au bureau)
        pause = f.get("save_mode") == "pause"
        vals = {
            "x_fdt_fait": bool(s.get("x_fdt_fait")) if pause else True,
            "x_fdt_date": datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S"),
            "x_fdt_vehicule":      f.get("vehicule", "").strip(),
            "x_fdt_heure_arrivee": f.get("heure_arrivee", "").strip(),
            "x_fdt_heure_depart":  f.get("heure_depart", "").strip(),
            "x_fdt_temps_trajet":  f.get("temps_trajet", "").strip(),
            "x_fdt_operateurs":    f.get("operateurs", "").strip(),
            "x_fdt_commentaires":  f.get("commentaires", "").strip(),
            "x_fdt_data":          json.dumps(data, ensure_ascii=False),
            "x_fdt_signataire":    f.get("signataire", "").strip(),
            "x_fdt_prochain_entretien": f.get("prochain_entretien", "").strip(),
        }
        # Corps HTML du rapport d'intervention (lieu/interlocuteur/objet
        # repris de la commande liée quand elle existe)
        _d = date.fromisoformat(card["day"])
        _adresse = ", ".join(p for p in [card["street"],
                   f"{card['zip']} {card['ville']}".strip()] if p)
        _inter, _objet = "", ""
        _ref0 = card["ref"].split()[0] if card["ref"] else ""
        if _ref0:
            try:
                _so = x(models, uid, "sale.order", "search_read",
                        [["name", "=", _ref0]],
                        fields=["x_studio_lieu_dintervention_2",
                                "x_studio_interlocuteur_1",
                                "x_studio_lieu_dintervention_1"], limit=1)
                if _so:
                    _adresse = _so[0].get("x_studio_lieu_dintervention_2") or _adresse
                    _inter = _so[0].get("x_studio_interlocuteur_1") or ""
                    _objet = _so[0].get("x_studio_lieu_dintervention_1") or ""
            except Exception:
                pass
        vals["x_fdt_report_body"] = build_report_body(
            card["client"], _adresse, f"{_d.day:02d}/{_d.month:02d}/{_d.year}",
            vals["x_fdt_vehicule"], vals["x_fdt_operateurs"],
            vals["x_fdt_heure_arrivee"], vals["x_fdt_heure_depart"],
            vals["x_fdt_temps_trajet"], vals["x_fdt_commentaires"], data,
            interlocuteur=_inter, objet=_objet,
            prochain=vals["x_fdt_prochain_entretien"])
        sig = bin_val(f.get("signature", ""))
        if sig:
            vals["x_fdt_signature"] = sig
        nphotos = 0
        for i in (1, 2, 3, 4):
            pv = f.get(f"photo_{i}", "")
            b = bin_val(pv)
            if b:
                vals[f"x_fdt_photo_{i}"] = b
                nphotos += 1
            elif pv == "KEEP":
                nphotos += 1
        x(models, uid, "planning.slot", "write", [slot_id], vals)

        # Message chatter pour le bureau (mail.message direct : message_post
        # échappe le HTML passé par XML-RPC en v19) — pas de message en pause
        if not pause:
            try:
                html = render_fiche_html(card, vals, data, nphotos)
                subtype = x(models, uid, "ir.model.data", "check_object_reference", "mail", "mt_note")
                x(models, uid, "mail.message", "create", [{
                    "model": "planning.slot", "res_id": slot_id,
                    "message_type": "comment", "subtype_id": subtype[1],
                    "body": html,
                }])
            except Exception:
                pass

        back = request.args.get("back", "")
        return redirect(url_for(".fiche", slot_id=slot_id, t=token, back=back,
                                ok=2 if pause else 1))

    # GET : présence signature/photos sans télécharger les binaires
    bins = x(models, uid, "planning.slot", "read", [slot_id],
             fields=["x_fdt_signature", "x_fdt_photo_1", "x_fdt_photo_2",
                     "x_fdt_photo_3", "x_fdt_photo_4"],
             context={"bin_size": True})[0]
    has = {
        "sig": bool(bins.get("x_fdt_signature")),
        "p1": bool(bins.get("x_fdt_photo_1")),
        "p2": bool(bins.get("x_fdt_photo_2")),
        "p3": bool(bins.get("x_fdt_photo_3")),
        "p4": bool(bins.get("x_fdt_photo_4")),
    }

    # GET : préremplissage
    saved_data = {}
    if s.get("x_fdt_data"):
        try:
            saved_data = json.loads(s["x_fdt_data"])
        except (ValueError, TypeError):
            saved_data = {}

    # Nb de lignes remplies par onglet (badge)
    _sv_trav = saved_data.get("travaux", {}) if isinstance(saved_data, dict) else {}
    tab_counts = {tkey: sum(1 for code, _l, _u in rows if code in _sv_trav)
                  for tkey, _tl, rows in TRAVAUX_TABS}
    curage_rows = saved_data.get("curage", []) if isinstance(saved_data, dict) else []
    curage_commune = (saved_data.get("curage_commune") if isinstance(saved_data, dict) else "") or card["ville"]

    prefill = {
        "vehicule":      s.get("x_fdt_vehicule") or "",
        "heure_arrivee": s.get("x_fdt_heure_arrivee") or card["start"],
        "heure_depart":  s.get("x_fdt_heure_depart") or card["end"],
        "temps_trajet":  s.get("x_fdt_temps_trajet") or "",
        "operateurs":    _first_names(s.get("x_fdt_operateurs") or emp_names),
        "commentaires":  s.get("x_fdt_commentaires") or "",
        "prochain_entretien": s.get("x_fdt_prochain_entretien") or "",
    }
    d = date.fromisoformat(card["day"])
    adresse = ", ".join(p for p in [card["street"], f"{card['zip']} {card['ville']}".strip()] if p)
    ref0 = card["ref"].split()[0] if card["ref"] else ""
    if ref0:
        so_l = x(models, uid, "sale.order", "search_read",
                 [["name", "=", ref0]], fields=["x_studio_lieu_dintervention_2"], limit=1)
        if so_l and so_l[0].get("x_studio_lieu_dintervention_2"):
            adresse = so_l[0]["x_studio_lieu_dintervention_2"]

    # Préremplissage du véhicule depuis le camion renseigné sur le BI
    if not prefill["vehicule"] and ref0:
        try:
            bi = x(models, uid, "stock.picking", "search_read",
                   [["origin", "=", ref0], ["x_camion_id", "!=", False]],
                   fields=["x_camion_id"], limit=1)
            if bi and bi[0].get("x_camion_id"):
                prefill["vehicule"] = bi[0]["x_camion_id"][1]
        except Exception:
            pass

    return render_template("fiche.html",
        card=card, slot_id=slot_id, token=token,
        date_label=f"{d.day:02d}/{d.month:02d}/{d.year}",
        adresse=adresse, prefill=prefill,
        travaux=TRAVAUX, travaux_tabs=TRAVAUX_TABS, tab_counts=tab_counts,
        curage_rows=curage_rows, curage_commune=curage_commune,
        dechets=DECHETS, destinations=DESTINATIONS,
        vehicules=get_fleet_vehicles(uid, models),
        saved=saved_data,
        has=has,
        signataire=s.get("x_fdt_signataire") or "",
        deja_fait=bool(s.get("x_fdt_fait")),
        en_pause=bool(s.get("x_fdt_data")) and not s.get("x_fdt_fait"),
        ok=request.args.get("ok"),
        back=request.args.get("back", ""))

def render_fiche_html(card, vals, data, nphotos=0):
    """Résumé HTML de la fiche pour le chatter Odoo."""
    rows_t = ""
    labels_t = {c: (l, u) for c, l, u in TRAVAUX}
    for code, v in data.get("travaux", {}).items():
        label, unit = labels_t.get(code, (code, ""))
        rows_t += f"<tr><td>{label}</td><td>{v.get('qte','')} {unit}</td><td>{v.get('temps','')} h</td></tr>"
    rows_d = ""
    labels_d = dict(DECHETS)
    for code, v in data.get("dechets", {}).items():
        rows_d += (f"<tr><td>{labels_d.get(code, code)}</td>"
                   f"<td>{v.get('volume','')}</td><td>{v.get('destination','')}</td><td>{v.get('temps','')}</td></tr>")
    rows_c = ""
    total_ml = 0
    for v in data.get("curage", []):
        try:
            total_ml += float((v.get("longueur") or "0").replace(",", "."))
        except ValueError:
            pass
        rows_c += (f"<tr><td>{v.get('rue','')}</td><td>{v.get('longueur','')}</td>"
                   f"<td>{'Oui' if v.get('amiante') else 'Non'}</td><td>{v.get('probleme','')}</td></tr>")
    html = (
        f"<p><b>📝 Fiche de fin de travaux</b> — {card['client']}</p>"
        f"<p>Véhicule : {vals['x_fdt_vehicule'] or '—'} · "
        f"Arrivée : {vals['x_fdt_heure_arrivee'] or '—'} · Départ : {vals['x_fdt_heure_depart'] or '—'} · "
        f"Trajet A/R : {vals['x_fdt_temps_trajet'] or '—'} · Opérateur(s) : {vals['x_fdt_operateurs'] or '—'}</p>"
        + (f"<p>✍️ Signé par : <b>{vals.get('x_fdt_signataire')}</b></p>" if vals.get('x_fdt_signataire') else "")
        + (f"<p>📷 {nphotos} photo(s) jointe(s) — voir l'onglet Fiche de fin de travaux</p>" if nphotos else "")
    )
    if rows_t:
        html += ("<table border='1' cellpadding='3'><tr><th>Nature des travaux</th><th>Quantité</th><th>Temps</th></tr>"
                 + rows_t + "</table>")
    if vals["x_fdt_commentaires"]:
        html += f"<p><b>Commentaires :</b><br/>{vals['x_fdt_commentaires']}</p>"
    if rows_d:
        html += ("<table border='1' cellpadding='3'><tr><th>Déchets</th><th>Vol. m³</th>"
                 "<th>Destination</th><th>Tps dépotage</th></tr>" + rows_d + "</table>")
    if rows_c:
        commune = data.get("curage_commune") or ""
        html += (f"<p><b>Tranches de curage</b>{' — ' + commune if commune else ''} "
                 f"— total {total_ml:g} ml</p>"
                 "<table border='1' cellpadding='3'><tr><th>Rue</th><th>Longueur (ml)</th>"
                 "<th>Amiante</th><th>Problèmes rencontrés</th></tr>" + rows_c + "</table>")
    return html

def _labelize(code):
    """Libellé lisible pour un code inconnu (anciennes fiches)."""
    return code.replace("_", " ").capitalize()

def _first_names(s):
    """Ne garde que le prénom (1er mot) de chaque opérateur, ex.
    « Alexandre GIRAUDEAU, Didier BELLIARD » → « Alexandre, Didier »."""
    parts = [p.strip() for p in (s or "").split(",") if p.strip()]
    return ", ".join(p.split()[0] for p in parts if p.split())

def build_report_body(client, adresse, date_label, vehicule, operateurs,
                      h_arrivee, h_depart, trajet, commentaires, data,
                      interlocuteur="", objet="", prochain=""):
    """Corps HTML du rapport d'intervention client. Enrobé côté Odoo par le
    rapport QWeb (en-tête société + logo, titre, photos, signatures)."""
    from markupsafe import escape
    def E(v): return str(escape(v or ""))
    H2 = ("background:#0b7285;color:#fff;font-size:12px;padding:6px 12px;"
          "border-radius:3px;margin:16px 0 6px;letter-spacing:0.4px;")
    th = ("padding:6px 10px;text-align:left;background:#e6f4f1;color:#0b7285;"
          "font-size:11px;border:1px solid #d3e5e2;")
    td = "padding:6px 10px;border:1px solid #e5e7eb;font-size:12px;"
    lbl = ("padding:6px 8px;font-weight:700;color:#0b7285;width:145px;"
           "vertical-align:top;border-bottom:1px solid #eef2f4;font-size:12px;")
    val = "padding:6px 8px;border-bottom:1px solid #eef2f4;font-size:12px;"

    def info(k, v):
        return f"<tr><td style='{lbl}'>{k}</td><td style='{val}'>{E(v) or '—'}</td></tr>"

    html = [f"<h3 style='{H2}'>INFORMATIONS</h3>"]
    html.append("<table style='width:100%;border-collapse:collapse;margin-bottom:6px;'>")
    html.append(info("Client", client))
    if interlocuteur:
        html.append(info("Interlocuteur", interlocuteur))
    html.append(info("Lieu d'intervention", adresse))
    html.append(info("Date d'intervention", date_label))
    html.append(info("Technicien(s)", _first_names(operateurs)))
    html.append(info("Véhicule", vehicule))
    html.append(info("Arrivée / Départ", f"{h_arrivee or '—'}  →  {h_depart or '—'}"))
    html.append(info("Temps de trajet A/R", trajet))
    html.append("</table>")

    labels_t = {c: (l, u) for c, l, u in TRAVAUX}
    rows_t = data.get("travaux", {})
    if rows_t:
        html.append(f"<h3 style='{H2}'>PRESTATIONS RÉALISÉES</h3>")
        html.append("<table style='width:100%;border-collapse:collapse;margin-bottom:6px;'>")
        html.append(f"<tr><th style='{th}'>Prestation</th><th style='{th};width:100px;'>Quantité</th>"
                    f"<th style='{th};width:80px;'>Temps</th></tr>")
        for code, v in rows_t.items():
            label, unit = labels_t.get(code, (_labelize(code), ""))
            qte = f"{E(v.get('qte',''))} {unit}".strip()
            html.append(f"<tr><td style='{td}'>{label}</td><td style='{td}'>{qte}</td>"
                        f"<td style='{td}'>{E(v.get('temps',''))} {'h' if v.get('temps') else ''}</td></tr>")
        html.append("</table>")

    labels_d = dict(DECHETS)
    rows_d = data.get("dechets", {})
    if rows_d:
        html.append(f"<h3 style='{H2}'>MATIÈRES ÉVACUÉES — TRAÇABILITÉ</h3>")
        html.append("<table style='width:100%;border-collapse:collapse;margin-bottom:6px;'>")
        html.append(f"<tr><th style='{th}'>Déchet</th><th style='{th};width:100px;'>Volume (m³)</th>"
                    f"<th style='{th}'>Destination de traitement</th>"
                    f"<th style='{th};width:120px;'>Tps dépotage (h)</th></tr>")
        for code, v in rows_d.items():
            html.append(f"<tr><td style='{td}'>{labels_d.get(code, _labelize(code))}</td>"
                        f"<td style='{td}'>{E(v.get('volume',''))}</td>"
                        f"<td style='{td}'>{E(v.get('destination',''))}</td>"
                        f"<td style='{td}'>{E(v.get('temps',''))}</td></tr>")
        html.append("</table>")

    rows_c = data.get("curage", [])
    if rows_c:
        commune = data.get("curage_commune") or ""
        total_ml = 0
        for v in rows_c:
            try:
                total_ml += float((v.get("longueur") or "0").replace(",", "."))
            except ValueError:
                pass
        title = "TRANCHES DE CURAGE" + (f" — {E(commune).upper()}" if commune else "")
        html.append(f"<h3 style='{H2}'>{title}</h3>")
        html.append("<table style='width:100%;border-collapse:collapse;margin-bottom:6px;'>")
        html.append(f"<tr><th style='{th}'>Rue</th><th style='{th};width:100px;'>Longueur (ml)</th>"
                    f"<th style='{th};width:80px;'>Amiante</th><th style='{th}'>Problèmes rencontrés</th></tr>")
        for v in rows_c:
            html.append(f"<tr><td style='{td}'>{E(v.get('rue',''))}</td>"
                        f"<td style='{td}'>{E(v.get('longueur',''))}</td>"
                        f"<td style='{td}'>{'Oui' if v.get('amiante') else 'Non'}</td>"
                        f"<td style='{td}'>{E(v.get('probleme',''))}</td></tr>")
        html.append(f"<tr><td style='{td};font-weight:700;'>TOTAL</td>"
                    f"<td style='{td};font-weight:700;'>{total_ml:g} ml</td>"
                    f"<td style='{td}'></td><td style='{td}'></td></tr>")
        html.append("</table>")

    if commentaires or prochain:
        html.append(f"<h3 style='{H2}'>OBSERVATIONS &amp; RECOMMANDATIONS</h3>")
        html.append("<div style='border:1px solid #e5e7eb;border-left:4px solid #0b7285;"
                    "background:#f8fafc;padding:9px 13px;font-size:12px;"
                    "white-space:pre-wrap;margin-bottom:6px;'>")
        if commentaires:
            html.append(E(commentaires))
        if prochain:
            html.append(("<br/><br/>" if commentaires else "")
                        + f"<b>Prochain entretien conseillé : {E(prochain)}</b>")
        html.append("</div>")

    return "".join(html)

# ─── BON D'INTERVENTION (PDF chiffré, via session web du compte technique) ──
# NB : le rapport « Fiche de fin de travaux » est consultable/envoyable depuis
# la commande, la facture et le BI via des smart boutons Odoo (rendu à la volée).
BI_REPORT = "protec_custom.report_deliveryslip_priced"
_web_session = {"opener": None}

def _web_opener():
    if _web_session["opener"] is not None:
        return _web_session["opener"]
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    payload = json.dumps({"jsonrpc": "2.0", "method": "call",
        "params": {"db": ODOO_DB, "login": ODOO_USER, "password": ODOO_PASSWORD}}).encode()
    req = urllib.request.Request(f"{ODOO_URL}/web/session/authenticate", data=payload,
                                 headers={"Content-Type": "application/json"})
    opener.open(req, timeout=30)
    _web_session["opener"] = opener
    return opener

@bp.route("/bi/<int:slot_id>")
def bi_pdf(slot_id):
    token = request.args.get("t", "")
    if not check_sig(f"fdt:{slot_id}", token):
        abort(403)
    uid, models = odoo_connect()
    rec = x(models, uid, "planning.slot", "read", [slot_id], fields=["name"])
    if not rec:
        abort(404)
    ref = (rec[0]["name"] or "").strip().split()[0] if rec[0].get("name") else ""
    if not ref:
        return "Pas de commande liée à cette intervention.", 404
    picks = x(models, uid, "stock.picking", "search_read",
              [["origin", "=", ref], ["picking_type_id.sequence_code", "=", "ASS/BI/"]],
              fields=["id", "name", "company_id"], order="id desc", limit=1)
    if not picks:
        return "Pas de bon d'intervention pour cette intervention.", 404
    pid, pname = picks[0]["id"], picks[0]["name"]
    cid = picks[0]["company_id"][0] if picks[0].get("company_id") else 2
    # allowed_company_ids dans l'URL : requis pour les enregistrements hors
    # société par défaut de la session (multi-sociétés)
    import urllib.parse as _up
    ctx = _up.quote(json.dumps({"allowed_company_ids": [cid]}))
    for _attempt in (1, 2):
        try:
            op = _web_opener()
            r = op.open(f"{ODOO_URL}/report/pdf/{BI_REPORT}/{pid}?context={ctx}", timeout=60)
            pdf = r.read()
            if pdf[:4] == b"%PDF":
                fn = pname.replace("/", "_")
                return Response(pdf, mimetype="application/pdf",
                                headers={"Content-Disposition": f'inline; filename="BI_{fn}.pdf"'})
        except Exception:
            pass
        _web_session["opener"] = None  # session expirée : ré-authentifier et réessayer
    return "Erreur de génération du PDF — réessayez.", 502

# ─── PLEIN DE GASOIL (TICPE) ────────────────────────────────────────────────
# Chaque plein saisi par un chauffeur devient un relevé compteur du module
# Parc automobile (fleet.vehicle.odometer + litres) → suivi TICPE automatisé.
def _num(f, name):
    v = (f.get(name) or "").replace(",", ".").replace(" ", "")
    return float(v) if v else 0.0

def _save_plein(uid, models, vid, km, litres, jour, chauffeur, compteur):
    """Crée le relevé odomètre + litrage. Lève ValueError si saisie invalide."""
    if not vid or km <= 0 or litres <= 0:
        raise ValueError("Camion, compteur km et litres sont obligatoires.")
    prev = x(models, uid, "fleet.vehicle.odometer", "search_read",
             [["vehicle_id", "=", vid]], fields=["value", "date"],
             order="value desc", limit=1)
    vals = {"vehicle_id": vid, "date": jour, "value": km,
            "x_litres": litres, "x_chauffeur": chauffeur}
    if compteur > 0:
        vals["x_compteur_cuve"] = compteur
    x(models, uid, "fleet.vehicle.odometer", "create", [vals])
    result = {"litres": litres, "km": km}
    if prev and km > prev[0]["value"] > 0:
        dist = km - prev[0]["value"]
        result["distance"] = dist
        result["conso"] = round(litres / dist * 100, 1)
    return result

def _write_plein(uid, models, rec_id, vid, km, litres, jour, chauffeur, compteur):
    """Met à jour un relevé existant. Lève ValueError si saisie invalide."""
    if not vid or km <= 0 or litres <= 0:
        raise ValueError("Camion, compteur km et litres sont obligatoires.")
    vals = {"vehicle_id": vid, "date": jour, "value": km,
            "x_litres": litres, "x_chauffeur": chauffeur}
    if compteur > 0:
        vals["x_compteur_cuve"] = compteur
    x(models, uid, "fleet.vehicle.odometer", "write", [rec_id], vals)
    return {"litres": litres, "km": km}

def _last_cuve(uid, models):
    last_cuve = x(models, uid, "fleet.vehicle.odometer", "search_read",
                  [["x_compteur_cuve", ">", 0]], fields=["x_compteur_cuve"],
                  order="x_compteur_cuve desc", limit=1)
    return last_cuve[0]["x_compteur_cuve"] if last_cuve else None

@bp.route("/plein", methods=["GET", "POST"])
def plein():
    emp_id = request.args.get("c", type=int)
    sig = request.args.get("s", "")
    if not emp_id or not check_sig(f"driver:{emp_id}", sig):
        abort(403)
    uid, models = odoo_connect()
    emp_name = emp_name_map(uid, models).get(emp_id) or f"Chauffeur {emp_id}"

    result, error = None, None
    if request.method == "POST":
        f = request.form
        try:
            vid = int(f.get("vehicule") or 0)
            jour = f.get("date") or date.today().isoformat()
            result = _save_plein(uid, models, vid, _num(f, "km"), _num(f, "litres"),
                                  jour, emp_name.split()[0], _num(f, "compteur"))
        except ValueError as ex:
            error = str(ex) or "Saisie invalide — vérifiez les valeurs."
        except Exception:
            error = "Erreur d'enregistrement — réessayez."

    vehicles = x(models, uid, "fleet.vehicle", "search_read",
                 [["active", "=", True]], fields=["id", "name", "odometer"],
                 order="name asc")
    return render_template("plein.html", emp_id=emp_id, sig=sig, emp_name=emp_name,
                           vehicles=vehicles, today=date.today().isoformat(),
                           last_cuve=_last_cuve(uid, models), result=result, error=error)

# ─── ASTREINTE (opération non prévue) ────────────────────────────────────────
# Le chauffeur crée lui-même la tâche planning depuis son téléphone, puis
# enchaîne directement sur la fiche de fin de travaux correspondante.
@bp.route("/astreinte", methods=["GET", "POST"])
def astreinte():
    emp_id = request.args.get("c", type=int)
    sig = request.args.get("s", "")
    if not emp_id or not check_sig(f"driver:{emp_id}", sig):
        abort(403)
    uid, models = odoo_connect()
    emp_name = emp_name_map(uid, models).get(emp_id) or f"Chauffeur {emp_id}"

    error = None
    if request.method == "POST":
        f = request.form
        try:
            desc = (f.get("description") or "").strip()
            if not desc:
                raise ValueError("Décrivez l'intervention (client, lieu…).")
            jour = date.fromisoformat(f.get("date") or date.today().isoformat())
            h1 = f.get("heure_debut") or "08:00"
            h2 = f.get("heure_fin") or ""
            hh1, mm1 = int(h1[:2]), int(h1[3:5])
            start_l = datetime(jour.year, jour.month, jour.day, hh1, mm1, tzinfo=TZ)
            if h2:
                hh2, mm2 = int(h2[:2]), int(h2[3:5])
                end_l = datetime(jour.year, jour.month, jour.day, hh2, mm2, tzinfo=TZ)
                if end_l <= start_l:
                    end_l += timedelta(days=1)
            else:
                end_l = start_l + timedelta(hours=2)
            emp = x(models, uid, "hr.employee", "read", [emp_id],
                    fields=["resource_id", "company_id"])[0]
            vals = {
                "name": f"ASTREINTE — {desc}",
                "start_datetime": start_l.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S"),
                "end_datetime": end_l.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S"),
                "company_id": emp["company_id"][0] if emp.get("company_id") else 2,
            }
            if emp.get("resource_id"):
                vals["resource_ids"] = [(6, 0, [emp["resource_id"][0]])]
            created = x(models, uid, "planning.slot", "create", [vals])
            slot_id = created[0] if isinstance(created, list) else created
            back = url_for(".ma_tournee", c=emp_id, s=sig)
            return redirect(url_for(".fiche", slot_id=slot_id,
                                    t=fiche_token(slot_id), back=back))
        except ValueError as ex:
            error = str(ex) or "Saisie invalide — vérifiez les valeurs."
        except Exception:
            error = "Erreur d'enregistrement — réessayez."

    now_l = datetime.now(TZ)
    return render_template("astreinte.html", emp_id=emp_id, sig=sig, emp_name=emp_name,
                           today=date.today().isoformat(),
                           now_h=now_l.strftime("%H:%M"), error=error)

def _shift_month(mois, delta):
    y, mo = int(mois[:4]), int(mois[5:7])
    mo += delta
    y += (mo - 1) // 12
    mo = (mo - 1) % 12 + 1
    return f"{y:04d}-{mo:02d}"

# ─── SAISIE MANUELLE (admin/secrétariat) ─────────────────────────────────────
# Backup de /plein pour Manon : si un chauffeur n'a pas pu saisir son plein
# depuis son téléphone, la secrétaire peut le faire ici pour n'importe quel
# camion / chauffeur, avec la même logique d'enregistrement.
@bp.route("/plein-manuel", methods=["GET", "POST"])
def plein_manuel():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    uid, models = odoo_connect()
    employees = get_all_employees(uid, models)

    result, error = None, None
    if request.method == "POST":
        f = request.form
        try:
            vid = int(f.get("vehicule") or 0)
            chauffeur = (f.get("chauffeur_autre") or "").strip() or (f.get("chauffeur") or "").strip()
            if not chauffeur:
                raise ValueError("Le nom du chauffeur est obligatoire.")
            jour = f.get("date") or date.today().isoformat()
            edit_id = f.get("edit_id", type=int)
            if edit_id:
                result = _write_plein(uid, models, edit_id, vid, _num(f, "km"), _num(f, "litres"),
                                       jour, chauffeur, _num(f, "compteur"))
                result["updated"] = True
            else:
                result = _save_plein(uid, models, vid, _num(f, "km"), _num(f, "litres"),
                                      jour, chauffeur, _num(f, "compteur"))
            result["vehicule"] = f.get("vehicule_label", "")
            result["chauffeur"] = chauffeur
        except ValueError as ex:
            error = str(ex) or "Saisie invalide — vérifiez les valeurs."
        except Exception:
            error = "Erreur d'enregistrement — réessayez."

    edit_rec = None
    edit_id = request.args.get("edit", type=int)
    if edit_id and request.method == "GET":
        rec = x(models, uid, "fleet.vehicle.odometer", "read", [edit_id],
                fields=["date", "vehicle_id", "value", "x_litres", "x_chauffeur", "x_compteur_cuve"])
        if rec:
            edit_rec = rec[0]

    vehicles = x(models, uid, "fleet.vehicle", "search_read",
                 [["active", "=", True]], fields=["id", "name", "odometer"],
                 order="name asc")

    mois = request.args.get("mois") or date.today().strftime("%Y-%m")
    d1, d2 = _ebp_month_range(mois)
    veh_filter = request.args.get("veh", type=int)
    dom = [["x_litres", ">", 0], ["date", ">=", d1], ["date", "<=", d2]]
    if veh_filter:
        dom.append(["vehicle_id", "=", veh_filter])
    historique = x(models, uid, "fleet.vehicle.odometer", "search_read",
                   dom, fields=["date", "vehicle_id", "value", "x_litres", "x_chauffeur"],
                   order="date desc, id desc")
    mois_label = f"{MOIS_FR[int(mois[5:7]) - 1].capitalize()} {mois[:4]}"

    return render_template("plein_manuel.html", token=SECRET, vehicles=vehicles,
                           employees=employees, today=date.today().isoformat(),
                           last_cuve=_last_cuve(uid, models), historique=historique,
                           mois=mois, mois_label=mois_label, mois_prev=_shift_month(mois, -1),
                           mois_next=_shift_month(mois, 1), veh_filter=veh_filter,
                           edit_rec=edit_rec, result=result, error=error)

@bp.route("/plein-manuel/supprimer", methods=["POST"])
def plein_manuel_supprimer():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    uid, models = odoo_connect()
    rec_id = request.form.get("id", type=int)
    if rec_id:
        x(models, uid, "fleet.vehicle.odometer", "unlink", [rec_id])
    mois = request.form.get("mois") or date.today().strftime("%Y-%m")
    veh_filter = request.form.get("veh", type=int)
    return redirect(url_for(".plein_manuel", token=SECRET, mois=mois, veh=veh_filter or None))

# ─── PWA : hors connexion pour les chauffeurs ────────────────────────────────
SW_JS = """
const CACHE = 'protec-tournee-v1';
self.addEventListener('install', function (e) { self.skipWaiting(); });
self.addEventListener('activate', function (e) { e.waitUntil(clients.claim()); });
self.addEventListener('fetch', function (e) {
  const req = e.request;
  if (req.method !== 'GET') return;
  const p = new URL(req.url).pathname;
  if (p.indexOf('/ma-tournee') === -1 && p.indexOf('/fiche/') === -1 &&
      p.indexOf('/bi/') === -1 && p.indexOf('/icon-') === -1) return;
  e.respondWith(
    fetch(req).then(function (r) {
      if (r.ok) { var cp = r.clone(); caches.open(CACHE).then(function (c) { c.put(req, cp); }); }
      return r;
    }).catch(function () { return caches.match(req); })
  );
});
"""

@bp.route("/sw.js")
def sw_js():
    return Response(SW_JS, mimetype="application/javascript",
                    headers={"Cache-Control": "no-cache"})

@bp.route("/manifest.json")
def manifest():
    c = request.args.get("c", "")
    sg = request.args.get("s", "")
    base = BASE_PREFIX()
    start = f"{base}/ma-tournee?c={c}&s={sg}" if c else f"{base}/"
    return json.dumps({
        "name": "Ma tournée PROTEC", "short_name": "Tournée",
        "start_url": start, "scope": base + "/",
        "display": "standalone", "background_color": "#3a5a99",
        "theme_color": "#3a5a99",
        "icons": [{"src": f"{base}/icon-192.png", "sizes": "192x192", "type": "image/png"},
                  {"src": f"{base}/icon-512.png", "sizes": "512x512", "type": "image/png"}],
    }), 200, {"Content-Type": "application/manifest+json"}

def BASE_PREFIX():
    return url_for("protec.week_view").rstrip("/")

_icon_cache = {}

@bp.route("/icon-<int:size>.png")
def icon(size):
    if size not in (192, 512):
        abort(404)
    if size not in _icon_cache:
        from PIL import Image, ImageDraw, ImageFont
        import io
        img = Image.new("RGB", (size, size), "#3a5a99")
        d = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", int(size * 0.55))
        except Exception:
            font = ImageFont.load_default()
        d.text((size / 2, size / 2), "P", fill="#ffffff", font=font, anchor="mm")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        _icon_cache[size] = buf.getvalue()
    return Response(_icon_cache[size], mimetype="image/png",
                    headers={"Cache-Control": "public, max-age=86400"})

# ─── TARIFS CLIENTS (admin/bureau) ───────────────────────────────────────────
# Page ouverte par les smart boutons « Tarifs client » (fiche contact + devis).
# Tarif standard = liste de prix par défaut PROTEC ; tarifs spécifiques =
# liste de prix propre au client (items à prix fixe, avec plage de validité).
# La page permet aussi d'ajouter un tarif spécifique sur une plage donnée.
TARIFS_COMPANY_ID = 2

def _tarifs_paris_to_utc(day_str, end=False):
    """'2026-08-07' → chaîne datetime UTC pour date_start/date_end d'un item."""
    d = date.fromisoformat(day_str)
    t = datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=TZ) if end else \
        datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=TZ)
    return t.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S")

def _tarifs_fmt_dt(dt_str):
    dt = utc_to_local(dt_str)
    return dt.strftime("%d/%m/%Y") if dt else ""

@bp.route("/tarifs", methods=["GET", "POST"])
def tarifs_client():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    pid = request.args.get("c", type=int)
    src = request.args.get("src", "")
    if not pid:
        abort(404)
    uid, models, db = odoo_connect_src(src)
    def q(model, method, *p, **kw):
        return xq(db, models, uid, model, method, *p, **kw)
    CTX = {"allowed_company_ids": [TARIFS_COMPANY_ID]}

    partner = q("res.partner", "read", [pid],
                fields=["name", "property_product_pricelist"], context=CTX)
    if not partner:
        abort(404)
    partner = partner[0]
    default_pl = q("product.pricelist", "search",
                   [["name", "=", "Liste de prix par défaut"],
                    ["company_id", "=", TARIFS_COMPANY_ID]], limit=1)
    default_pl_id = default_pl[0] if default_pl else 0
    prop_pl = partner.get("property_product_pricelist")
    client_pl_id = prop_pl[0] if prop_pl else 0
    # une liste "générique" (défaut PROTEC ou d'une autre société) ≠ tarif spécifique
    client_pl_name = prop_pl[1] if prop_pl else ""
    if client_pl_id:
        pl_info = q("product.pricelist", "read", [client_pl_id], fields=["company_id"])
        if client_pl_id == default_pl_id or \
           (pl_info and pl_info[0]["company_id"] and
            pl_info[0]["company_id"][0] != TARIFS_COMPANY_ID):
            client_pl_id = 0

    error, ok = None, request.args.get("ok")
    if request.method == "POST":
        f = request.form
        try:
            tmpl_id = int(f.get("tmpl") or 0)
            prix = float((f.get("prix") or "").replace(",", ".").replace(" ", ""))
            du, au = f.get("du", "").strip(), f.get("au", "").strip()
            if not tmpl_id or prix <= 0:
                raise ValueError("Produit et prix sont obligatoires.")
            if not client_pl_id:
                created = q("product.pricelist", "create",
                    [{"name": partner["name"], "company_id": TARIFS_COMPANY_ID}])
                client_pl_id = created[0] if isinstance(created, list) else created
                q("res.partner", "write", [pid],
                  {"property_product_pricelist": client_pl_id}, context=CTX)
            vals = {"pricelist_id": client_pl_id, "applied_on": "1_product",
                    "product_tmpl_id": tmpl_id, "compute_price": "fixed",
                    "fixed_price": prix}
            if du:
                vals["date_start"] = _tarifs_paris_to_utc(du)
            if au:
                vals["date_end"] = _tarifs_paris_to_utc(au, end=True)
            q("product.pricelist.item", "create", [vals])
            args = {"token": SECRET, "c": pid, "ok": 1}
            if src:
                args["src"] = src
            return redirect(url_for(".tarifs_client", **args))
        except ValueError as ex:
            error = str(ex) or "Saisie invalide."
        except Exception:
            error = "Erreur d'enregistrement — réessayez."

    # ── catalogue + prix standard ──
    products = q("product.product", "search_read",
                 [["sale_ok", "=", True], ["active", "=", True],
                  "|", ["company_id", "=", False], ["company_id", "=", TARIFS_COMPANY_ID]],
                 fields=["display_name", "default_code", "list_price",
                         "product_tmpl_id", "categ_id"])
    std_var, std_tmpl = {}, {}
    if default_pl_id:
        for it in q("product.pricelist.item", "search_read",
                    [["pricelist_id", "=", default_pl_id], ["compute_price", "=", "fixed"]],
                    fields=["applied_on", "product_id", "product_tmpl_id", "fixed_price"]):
            if it["applied_on"] == "0_product_variant" and it["product_id"]:
                std_var[it["product_id"][0]] = it["fixed_price"]
            elif it["product_tmpl_id"]:
                std_tmpl[it["product_tmpl_id"][0]] = it["fixed_price"]

    # ── items spécifiques du client (avec plages de validité) ──
    cli_var, cli_tmpl = {}, {}
    if client_pl_id:
        for it in q("product.pricelist.item", "search_read",
                    [["pricelist_id", "=", client_pl_id]],
                    fields=["applied_on", "product_id", "product_tmpl_id",
                            "compute_price", "fixed_price", "percent_price",
                            "date_start", "date_end", "min_quantity"],
                    order="date_start desc, id desc"):
            if it["applied_on"] == "0_product_variant" and it["product_id"]:
                cli_var.setdefault(it["product_id"][0], []).append(it)
            elif it["applied_on"] == "1_product" and it["product_tmpl_id"]:
                cli_tmpl.setdefault(it["product_tmpl_id"][0], []).append(it)

    # ── historique facturé (postées, société PROTEC) ──
    hist_by_prod = {}
    for l in q("account.move.line", "search_read",
               [["parent_state", "=", "posted"],
                ["move_id.move_type", "in", ["out_invoice", "out_refund"]],
                ["partner_id", "child_of", pid],
                ["company_id", "=", TARIFS_COMPANY_ID],
                ["product_id", "!=", False]],
               fields=["date", "move_name", "product_id", "quantity",
                       "price_unit", "discount"],
               order="date desc, id desc", limit=3000):
        hist_by_prod.setdefault(l["product_id"][0], []).append(l)

    now_utc = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for p in products:
        vid, tid = p["id"], p["product_tmpl_id"][0]
        std = std_var.get(vid, std_tmpl.get(tid, p["list_price"]))
        items = cli_var.get(vid, []) + cli_tmpl.get(tid, [])
        current = next((i for i in items
                        if (not i["date_start"] or i["date_start"] <= now_utc)
                        and (not i["date_end"] or i["date_end"] >= now_utc)), None)
        future = min((i for i in items if i["date_start"] and i["date_start"] > now_utc),
                     key=lambda i: i["date_start"], default=None)
        shown = current or future
        cli_price, validity = None, ""
        if shown:
            if shown["compute_price"] == "fixed":
                cli_price = shown["fixed_price"]
            elif shown["compute_price"] == "percentage":
                cli_price = round(std * (1 - shown["percent_price"] / 100), 2)
            ds, de = shown["date_start"], shown["date_end"]
            if shown is future:
                validity = f"à partir du {_tarifs_fmt_dt(ds)}"
            elif ds and de:
                validity = f"du {_tarifs_fmt_dt(ds)} au {_tarifs_fmt_dt(de)}"
            elif de:
                validity = f"jusqu'au {_tarifs_fmt_dt(de)}"
            elif ds:
                validity = f"depuis le {_tarifs_fmt_dt(ds)}"
        hist = hist_by_prod.get(vid, [])
        items_edit = []
        for it in items:
            if it["compute_price"] != "fixed":
                continue
            du_l = utc_to_local(it["date_start"])
            au_l = utc_to_local(it["date_end"])
            lbl = f"{it['fixed_price']:.2f} €".replace(".", ",")
            if du_l:
                lbl += f" — du {du_l.strftime('%d/%m/%Y')}"
            if au_l:
                lbl += f" au {au_l.strftime('%d/%m/%Y')}"
            items_edit.append({"id": it["id"], "prix": f"{it['fixed_price']:.2f}",
                                "du": du_l.date().isoformat() if du_l else "",
                                "au": au_l.date().isoformat() if au_l else "",
                                "label": lbl})
        rows.append({
            "vid": vid, "tmpl": tid, "name": p["display_name"],
            "code": p["default_code"] or "",
            "categ": p["categ_id"][1] if p.get("categ_id") else "Autre",
            "std": std,
            "items_edit": items_edit,
            "specific": bool(items), "cli_price": cli_price, "validity": validity,
            "nb_items": len(items), "items": items,
            "ecart": round((cli_price - std) / std * 100, 1) if (cli_price is not None and std) else None,
            "hist": [{"date": h["date"], "piece": h["move_name"],
                      "qte": h["quantity"], "pu": h["price_unit"],
                      "remise": h["discount"]} for h in hist[:30]],
            "hist_all": hist,
        })
    rows.sort(key=lambda r: (not r["specific"], r["name"].lower()))
    nb_spec = sum(1 for r in rows if r["specific"])
    nb_fact = sum(1 for r in rows if r["hist"])
    categories = sorted({r["categ"] for r in rows}, key=str.lower)

    # ── onglet Évolution : prix par année (tarif spécifique, sinon facturé) ──
    def _year_of(dt_str):
        return int(dt_str[:4]) if dt_str else None
    years_set = set()
    for r in rows:
        for it in r["items"]:
            for k in ("date_start", "date_end"):
                y = _year_of(it[k])
                if y:
                    years_set.add(y)
        for h in r["hist_all"]:
            y = _year_of(h["date"])
            if y:
                years_set.add(y)
    this_year = date.today().year
    years = sorted(y for y in years_set if y <= this_year)[-8:] or [this_year]
    evo_rows = []
    for r in rows:
        if not r["items"] and not r["hist_all"]:
            continue
        cells = []
        has_data = False
        for y in years:
            y_start, y_end = f"{y}-01-01 00:00:00", f"{y}-12-31 23:59:59"
            tarif = None
            applicable = [it for it in r["items"]
                          if it["compute_price"] == "fixed"
                          and (not it["date_start"] or it["date_start"] <= y_end)
                          and (not it["date_end"] or it["date_end"] >= y_start)]
            if applicable:
                applicable.sort(key=lambda it: it["date_start"] or "", reverse=True)
                tarif = applicable[0]["fixed_price"]
            pus = [h["price_unit"] for h in r["hist_all"] if _year_of(h["date"]) == y]
            fact = round(sum(pus) / len(pus), 2) if pus else None
            if tarif is not None or fact is not None:
                has_data = True
            cells.append({"tarif": tarif, "fact": fact})
        if has_data:
            evo_rows.append({"name": r["name"], "code": r["code"],
                              "categ": r["categ"], "specific": r["specific"],
                              "cells": cells})

    edit_map = {r["vid"]: r["items_edit"] for r in rows if r["items_edit"]}
    return render_template("tarifs.html", partner=partner, rows=rows,
                           nb_spec=nb_spec, nb_fact=nb_fact, pid=pid, src=src,
                           categories=categories, years=years, evo_rows=evo_rows,
                           edit_map=edit_map,
                           client_pl_name=client_pl_name if client_pl_id else "",
                           token=SECRET, ok=ok, error=error,
                           today=date.today().isoformat())

# ── Modification / suppression d'un tarif spécifique existant ───────────────
def _tarifs_item_ok(q, item_id):
    """L'item existe et appartient bien à une liste de prix PROTEC."""
    it = q("product.pricelist.item", "read", [item_id], fields=["pricelist_id"])
    if not it:
        return False
    pl = q("product.pricelist", "read", [it[0]["pricelist_id"][0]], fields=["company_id"])
    return bool(pl and pl[0]["company_id"] and pl[0]["company_id"][0] == TARIFS_COMPANY_ID)

@bp.route("/tarifs/item/modifier", methods=["POST"])
def tarifs_item_modifier():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    pid = request.args.get("c", type=int)
    src = request.args.get("src", "")
    uid, models, db = odoo_connect_src(src)
    def q(model, method, *p, **kw):
        return xq(db, models, uid, model, method, *p, **kw)
    f = request.form
    item_id = int(f.get("item_id") or 0)
    ok = "mod"
    try:
        prix = float((f.get("prix") or "").replace(",", ".").replace(" ", ""))
        if not item_id or prix <= 0 or not _tarifs_item_ok(q, item_id):
            raise ValueError()
        du, au = f.get("du", "").strip(), f.get("au", "").strip()
        q("product.pricelist.item", "write", [item_id], {
            "fixed_price": prix,
            "date_start": _tarifs_paris_to_utc(du) if du else False,
            "date_end": _tarifs_paris_to_utc(au, end=True) if au else False,
        })
    except (ValueError, TypeError):
        ok = None
    args = {"token": SECRET, "c": pid}
    if ok:
        args["ok"] = ok
    if src:
        args["src"] = src
    return redirect(url_for(".tarifs_client", **args))

@bp.route("/tarifs/item/supprimer", methods=["POST"])
def tarifs_item_supprimer():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    pid = request.args.get("c", type=int)
    src = request.args.get("src", "")
    uid, models, db = odoo_connect_src(src)
    def q(model, method, *p, **kw):
        return xq(db, models, uid, model, method, *p, **kw)
    item_id = int(request.form.get("item_id") or 0)
    if item_id and _tarifs_item_ok(q, item_id):
        q("product.pricelist.item", "unlink", [item_id])
    args = {"token": SECRET, "c": pid, "ok": "del"}
    if src:
        args["src"] = src
    return redirect(url_for(".tarifs_client", **args))

# ── Import Excel des tarifs client (Manon, client par client) ───────────────
def _tarifs_norm(s):
    """Normalise un libellé pour le rapprochement (minuscules, sans accents,
    abréviations métier développées)."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9 ]", " ", s.lower())
    s = re.sub(r"\bwe\b", "week end", s)
    s = re.sub(r"\bp en c\b", "prise en charge", s)
    s = re.sub(r"\bdmv\b", "matieres de vidange", s)
    return re.sub(r"\s+", " ", s).strip()

_TARIFS_STOPWORDS = {"de", "des", "du", "la", "le", "les", "l", "d", "et", "en",
                     "sur", "a", "au", "aux", "pour", "avec", "ou"}

def _tarifs_score(nlabel, nname):
    """Score de rapprochement : séquence + proportion de mots-clés retrouvés."""
    from difflib import SequenceMatcher
    seq = SequenceMatcher(None, nlabel, nname).ratio()
    ltoks = {t for t in nlabel.split() if t not in _TARIFS_STOPWORDS}
    ntoks = {t for t in nname.split() if t not in _TARIFS_STOPWORDS}
    common = {t for t in ltoks & ntoks if len(t) > 2}
    tok = len(common) / len(ltoks) if ltoks and common else 0.0
    return max(seq, tok)

def _tarifs_header_year(v):
    """Année portée par une cellule d'entête : 2026, n° de série Excel
    (44743 → 2022) ou texte contenant une année (« du 1er mai au 31/12/2025 »)."""
    try:
        n = float(v)
        if 1990 <= n <= 2100:
            return int(n)
        if 30000 <= n <= 60000:  # date Excel (jours depuis 30/12/1899)
            return (datetime(1899, 12, 30) + timedelta(days=int(n))).year
    except (TypeError, ValueError):
        pass
    mo = re.search(r"\b(19|20)\d{2}\b", str(v or ""))
    return int(mo.group(0)) if mo else None

def _tarifs_parse_excel(data, filename):
    """Extrait les blocs tarifaires du fichier « Évolution des tarifs ».
    Chaque feuille peut contenir plusieurs blocs (un par client/contrat),
    délimités par leur ligne d'entête NATURE PRESTATION précédée d'une ligne
    « Client : … ». Pour chaque prestation : prix de l'année la plus récente.
    Renvoie [(feuille, client, [(libellé, unité, prix, année), …]), …]"""
    import io
    grids = []
    if filename.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        for ws in wb.sheets():
            if ws.nrows:
                grids.append((ws.name, [[ws.cell_value(r, c) for c in range(ws.ncols)]
                                         for r in range(ws.nrows)]))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        for ws in wb.worksheets:
            if ws.max_row:
                grids.append((ws.title, [[c.value for c in row] for row in ws.iter_rows()]))

    blocks = []
    for sheet_name, grid in grids:
        headers = []
        for i, row in enumerate(grid):
            cells = [str(v or "").strip().upper() for v in row]
            if any("NATURE" in c and "PRESTATION" in c for c in cells):
                headers.append(i)
        for hi, header_i in enumerate(headers):
            end_i = headers[hi + 1] if hi + 1 < len(headers) else len(grid)
            # ligne « Client : … » au-dessus de l'entête
            client = ""
            for k in range(header_i - 1, max(header_i - 4, -1), -1):
                c0 = str(grid[k][0] or "").strip()
                if c0.lower().startswith("client"):
                    client = c0.split(":", 1)[-1].strip().split("\n")[0]
                    break
            hrow = grid[header_i]
            year_cols, unit_col = [], None
            for j, v in enumerate(hrow):
                sv = str(v or "").strip().lower()
                if sv.startswith("unit"):
                    unit_col = j
                    continue
                if sv in ("ced", "onu", "etiq.", "danger", "g.e."):
                    continue
                y = _tarifs_header_year(v)
                if y:
                    year_cols.append((y, j))
            # année la plus récente d'abord ; à année égale, colonne la plus à gauche
            year_cols.sort(key=lambda t: (-t[0], t[1]))
            rows = []
            for row in grid[header_i + 1:end_i]:
                label = str(row[0] or "").strip()
                if not label or label.lower().startswith("client"):
                    continue
                unit = str(row[unit_col] or "").strip() \
                    if unit_col is not None and unit_col < len(row) else ""
                price, year = None, None
                for y, j in year_cols:
                    if j < len(row):
                        try:
                            v = float(str(row[j]).replace(",", ".").replace("€", "").replace(" ", ""))
                            if v > 0:
                                price, year = v, y
                                break
                        except (TypeError, ValueError):
                            continue
                if price is not None:
                    rows.append((label, unit, price, year))
            if rows:
                blocks.append((sheet_name, client, rows))
    return blocks

@bp.route("/tarifs/import", methods=["POST"])
def tarifs_import():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    pid = request.args.get("c", type=int)
    src = request.args.get("src", "")
    f = request.files.get("fichier")
    uid, models, db = odoo_connect_src(src)
    def q(model, method, *p, **kw):
        return xq(db, models, uid, model, method, *p, **kw)
    CTX = {"allowed_company_ids": [TARIFS_COMPANY_ID]}
    partner = q("res.partner", "read", [pid], fields=["name"], context=CTX)[0]

    if not f or not f.filename:
        return redirect(url_for(".tarifs_client", token=SECRET, c=pid,
                                **({"src": src} if src else {})))
    try:
        blocks = _tarifs_parse_excel(f.read(), f.filename)
    except Exception:
        blocks = []

    # catalogue (templates) + prix standard pour le rapprochement
    products = q("product.template", "search_read",
                 [["sale_ok", "=", True], ["active", "=", True],
                  "|", ["company_id", "=", False], ["company_id", "=", TARIFS_COMPANY_ID]],
                 fields=["name", "default_code", "list_price"])
    default_pl = q("product.pricelist", "search",
                   [["name", "=", "Liste de prix par défaut"],
                    ["company_id", "=", TARIFS_COMPANY_ID]], limit=1)
    std_tmpl = {}
    if default_pl:
        for it in q("product.pricelist.item", "search_read",
                    [["pricelist_id", "=", default_pl[0]], ["compute_price", "=", "fixed"]],
                    fields=["product_tmpl_id", "fixed_price"]):
            if it["product_tmpl_id"]:
                std_tmpl.setdefault(it["product_tmpl_id"][0], it["fixed_price"])
    catalog = sorted(
        [{"tmpl": p["id"], "name": p["name"], "code": p["default_code"] or "",
          "std": std_tmpl.get(p["id"], p["list_price"])} for p in products],
        key=lambda c: c["name"].lower())

    # rapprochement automatique libellé Excel → produit Odoo, bloc par bloc
    norm_cat = [(c, _tarifs_norm(c["name"])) for c in catalog]
    review = []
    for bi, (sheet_name, client, lignes) in enumerate(blocks):
        bloc_label = f"Feuille « {sheet_name} »" + (f" — {client}" if client else "")
        for label, unit, price, year in lignes:
            nlabel = _tarifs_norm(label.split("\n")[0])
            best, best_score = None, 0.0
            for c, nname in norm_cat:
                score = _tarifs_score(nlabel, nname)
                if score > best_score:
                    best, best_score = c, score
            matched = best if best_score >= 0.45 else None
            std = matched["std"] if matched else None
            differs = matched is not None and abs(price - std) > 0.005
            review.append({"bloc": bi, "bloc_label": bloc_label,
                            "label": label, "unit": unit, "prix": price, "annee": year,
                            "tmpl": matched["tmpl"] if matched else 0,
                            "score": round(best_score * 100),
                            "std": std, "differs": differs})
    return render_template("tarifs_import.html", partner=partner, pid=pid, src=src,
                           token=SECRET, review=review, nb_blocs=len(blocks),
                           catalog=catalog, filename=f.filename,
                           today=date.today().isoformat())

@bp.route("/tarifs/import/valider", methods=["POST"])
def tarifs_import_valider():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    pid = request.args.get("c", type=int)
    src = request.args.get("src", "")
    uid, models, db = odoo_connect_src(src)
    def q(model, method, *p, **kw):
        return xq(db, models, uid, model, method, *p, **kw)
    CTX = {"allowed_company_ids": [TARIFS_COMPANY_ID]}
    f = request.form
    partner = q("res.partner", "read", [pid],
                fields=["name", "property_product_pricelist"], context=CTX)[0]
    default_pl = q("product.pricelist", "search",
                   [["name", "=", "Liste de prix par défaut"],
                    ["company_id", "=", TARIFS_COMPANY_ID]], limit=1)
    default_pl_id = default_pl[0] if default_pl else 0
    prop = partner.get("property_product_pricelist")
    client_pl_id = prop[0] if prop and prop[0] != default_pl_id else 0
    if client_pl_id:
        pli = q("product.pricelist", "read", [client_pl_id], fields=["company_id"])
        if pli and pli[0]["company_id"] and pli[0]["company_id"][0] != TARIFS_COMPANY_ID:
            client_pl_id = 0
    n = int(f.get("n") or 0)
    du = f.get("du", "").strip()
    created = 0
    for i in range(n):
        if not f.get(f"chk_{i}"):
            continue
        tmpl = int(f.get(f"tmpl_{i}") or 0)
        try:
            prix = float((f.get(f"prix_{i}") or "").replace(",", "."))
        except ValueError:
            continue
        if not tmpl or prix <= 0:
            continue
        if not client_pl_id:
            created_pl = q("product.pricelist", "create",
                [{"name": partner["name"], "company_id": TARIFS_COMPANY_ID}])
            client_pl_id = created_pl[0] if isinstance(created_pl, list) else created_pl
            q("res.partner", "write", [pid],
              {"property_product_pricelist": client_pl_id}, context=CTX)
        vals = {"pricelist_id": client_pl_id, "applied_on": "1_product",
                "product_tmpl_id": tmpl, "compute_price": "fixed",
                "fixed_price": prix}
        if du:
            vals["date_start"] = _tarifs_paris_to_utc(du)
        q("product.pricelist.item", "create", [vals])
        created += 1
    args = {"token": SECRET, "c": pid, "ok": created}
    if src:
        args["src"] = src
    return redirect(url_for(".tarifs_client", **args))

# ─── LIENS CHAUFFEURS (admin) ────────────────────────────────────────────────
@bp.route("/liens")
def liens():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    error, rows = None, []
    try:
        uid, models = odoo_connect()
        for e in get_all_employees(uid, models):
            url = url_for(".ma_tournee", c=e["id"], s=driver_sig(e["id"]), _external=True)
            rows.append({"id": e["id"], "name": e["name"], "color": emp_color(e["id"]),
                         "url": url})
    except Exception as ex:
        error = str(ex)
    return render_template("liens.html", rows=rows, error=error, token=SECRET)

# ─── EXPORT COMPTABLE EBP (admin) ────────────────────────────────────────────
# Journaux Odoo (PROTEC) → EBP : la comptable importe les CSV dans EBP SaaS.
# À terme : poussée directe via l'API EBP (couche push_ebp_api à câbler quand
# la clé d'abonnement et la doc des endpoints seront fournies).
EBP_COMPANY_ID = 2

def _ebp_month_range(mois):
    """'2026-06' → ('2026-06-01', '2026-06-30')."""
    import calendar
    y, mo = int(mois[:4]), int(mois[5:7])
    return (f"{y:04d}-{mo:02d}-01",
            f"{y:04d}-{mo:02d}-{calendar.monthrange(y, mo)[1]:02d}")

def _ebp_lines(uid, models, mois, journal_code=None):
    d1, d2 = _ebp_month_range(mois)
    dom = [["company_id", "=", EBP_COMPANY_ID], ["parent_state", "=", "posted"],
           ["account_id", "!=", False],  # exclut notes/sections sans compte
           "|", ["debit", "!=", 0], ["credit", "!=", 0],  # EBP refuse les lignes à 0
           ["date", ">=", d1], ["date", "<=", d2]]
    if journal_code:
        dom.append(["journal_id.code", "=", journal_code])
    return x(models, uid, "account.move.line", "search_read", dom,
             fields=["date", "move_id", "account_id", "partner_id", "name",
                     "debit", "credit", "journal_id"],
             order="date, move_id", limit=50000)

@bp.route("/ebp")
def ebp_page():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    mois = request.args.get("mois") or date.today().strftime("%Y-%m")
    error, rows, months = None, [], []
    try:
        uid, models = odoo_connect()
        lines = _ebp_lines(uid, models, mois)
        agg = {}
        for l in lines:
            code = l["journal_id"][1]
            jid = l["journal_id"][0]
            a = agg.setdefault(jid, {"name": code, "code": "", "nlines": 0,
                                     "debit": 0.0, "pieces": set()})
            a["nlines"] += 1
            a["debit"] += l["debit"] or 0
            a["pieces"].add(l["move_id"][0])
        if agg:
            jinfo = x(models, uid, "account.journal", "read", list(agg.keys()),
                      fields=["code", "name", "type"])
            for j in jinfo:
                agg[j["id"]]["code"] = j["code"]
                agg[j["id"]]["name"] = j["name"]
                agg[j["id"]]["jtype"] = j["type"]
        # regroupement : Ventes / Achats / Banque / TVA / Autres
        def cat(r):
            if r["code"] == "TVAE":
                return (3, "TVA")
            return {"sale": (0, "Ventes"), "purchase": (1, "Achats"),
                    "bank": (2, "Banque"), "cash": (2, "Banque")}.get(
                    r.get("jtype"), (4, "Autres"))
        for r in agg.values():
            r["corder"], r["categorie"] = cat(r)
        rows = sorted(agg.values(), key=lambda r: (r["corder"], r["code"]))
        for r in rows:
            r["npieces"] = len(r.pop("pieces"))
        d = date.today()
        for i in range(8):
            mm = (d.year * 12 + d.month - 1 - i)
            months.append(f"{mm // 12:04d}-{mm % 12 + 1:02d}")
    except Exception as ex:
        error = str(ex)
    return render_template("ebp.html", mois=mois, months=months, rows=rows,
                           mois_debut=_ebp_month_range(mois)[0],
                           mois_fin=_ebp_month_range(mois)[1],
                           error=error, token=request.args.get("token"))

def _ebp_csv_body(uid, models, mois, code):
    """Contenu CSV (format import EBP) d'un journal pour un mois."""
    lines = _ebp_lines(uid, models, mois, journal_code=code or None)
    acc_ids = list({l["account_id"][0] for l in lines})
    # code comptable : champ multi-société en v19 → lecture en contexte PROTEC
    accs = {a["id"]: a["code"] for a in x(models, uid, "account.account", "read",
            acc_ids, fields=["code"],
            context={"allowed_company_ids": [EBP_COMPANY_ID]})} if acc_ids else {}
    def fr(v):
        return f"{round(v or 0, 2):.2f}".replace(".", ",")
    out = ["Journal;Date;N° pièce;Compte;Libellé;Débit;Crédit;Partenaire"]
    for l in lines:
        lib = (l["name"] or "").replace(";", ",").replace("\n", " ")[:60]
        out.append(";".join([
            str(code or l["journal_id"][1]), str(l["date"]), str(l["move_id"][1]),
            str(accs.get(l["account_id"][0]) or ""), lib,
            fr(l["debit"]), fr(l["credit"]),
            (l["partner_id"][1] if l["partner_id"] else "").replace(";", ","),
        ]))
    return "﻿" + "\r\n".join(out)

@bp.route("/ebp.csv")
def ebp_csv():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    mois = request.args.get("mois") or date.today().strftime("%Y-%m")
    code = request.args.get("journal") or ""
    uid, models = odoo_connect()
    body = _ebp_csv_body(uid, models, mois, code)
    fn = f"Export_{code or 'TOUS'}_{mois}.csv"
    return Response(body, mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})

@bp.route("/ebp.zip")
def ebp_zip():
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    mois = request.args.get("mois") or date.today().strftime("%Y-%m")
    uid, models = odoo_connect()
    # journaux mouvementés du mois
    lines = _ebp_lines(uid, models, mois)
    jids = list({l["journal_id"][0] for l in lines})
    codes = sorted({j["code"] for j in x(models, uid, "account.journal", "read",
                    jids, fields=["code"])}) if jids else []
    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for code in codes:
            z.writestr(f"Export_{code}_{mois}.csv",
                       _ebp_csv_body(uid, models, mois, code).encode("utf-8"))
    buf.seek(0)
    return Response(buf.read(), mimetype="application/zip",
                    headers={"Content-Disposition":
                             f'attachment; filename="Export_EBP_{mois}.zip"'})

@bp.route("/ebp-fec")
def ebp_fec():
    """Export type FEC (18 colonnes, séparateur |) hors journaux d'ouverture."""
    if not SECRET or request.args.get("token") != SECRET:
        abort(403)
    mois = request.args.get("mois") or date.today().strftime("%Y-%m")
    d1 = request.args.get("du") or _ebp_month_range(mois)[0]
    d2 = request.args.get("au") or _ebp_month_range(mois)[1]
    uid, models = odoo_connect()
    CTX = {"allowed_company_ids": [EBP_COMPANY_ID]}

    # écriture d'ouverture de la société + journaux d'à-nouveaux exclus
    co = x(models, uid, "res.company", "read", [EBP_COMPANY_ID],
           fields=["account_opening_move_id"], context=CTX)[0]
    open_move = co["account_opening_move_id"][0] if co.get("account_opening_move_id") else 0

    dom = [["company_id", "=", EBP_COMPANY_ID], ["parent_state", "=", "posted"],
           ["account_id", "!=", False], ["date", ">=", d1], ["date", "<=", d2],
           "|", ["debit", "!=", 0], ["credit", "!=", 0],  # EBP refuse les lignes à 0
           ["journal_id.code", "not in", ["AN", "OUV"]]]
    if open_move:
        dom.append(["move_id", "!=", open_move])
    lines = x(models, uid, "account.move.line", "search_read", dom,
        fields=["journal_id", "move_id", "date", "account_id", "partner_id",
                "name", "debit", "credit", "matching_number"],
        order="date, move_id", limit=100000)

    jids = list({l["journal_id"][0] for l in lines})
    jmap = {j["id"]: j for j in x(models, uid, "account.journal", "read", jids,
            fields=["code", "name"])} if jids else {}
    aids = list({l["account_id"][0] for l in lines})
    amap = {a["id"]: a for a in x(models, uid, "account.account", "read", aids,
            fields=["code", "name", "account_type"], context=CTX)} if aids else {}
    mids = list({l["move_id"][0] for l in lines})
    mmap = {mv["id"]: mv for mv in x(models, uid, "account.move", "read", mids,
            fields=["name", "ref", "date", "invoice_date"])} if mids else {}
    pids = list({l["partner_id"][0] for l in lines if l["partner_id"]})
    pmap = {p["id"]: p for p in x(models, uid, "res.partner", "read", pids,
            fields=["ref", "name"])} if pids else {}

    def dt(s):  # 2026-06-30 -> 20260630
        return (s or "").replace("-", "")
    def fr(v):
        return f"{round(v or 0, 2):.2f}".replace(".", ",")
    out = ["JournalCode|JournalLib|EcritureNum|EcritureDate|CompteNum|CompteLib|"
           "CompAuxNum|CompAuxLib|PieceRef|PieceDate|EcritureLib|Debit|Credit|"
           "EcritureLet|DateLet|ValidDate|Montantdevise|Idevise"]
    for l in lines:
        j = jmap.get(l["journal_id"][0], {})
        a = amap.get(l["account_id"][0], {})
        mv = mmap.get(l["move_id"][0], {})
        aux_n, aux_l = "", ""
        if l["partner_id"] and a.get("account_type") in ("asset_receivable",
                                                         "liability_payable"):
            p = pmap.get(l["partner_id"][0], {})
            aux_n = p.get("ref") or f"P{l['partner_id'][0]}"
            aux_l = p.get("name") or ""
        def clean(s):
            return str(s or "").replace("|", "/").replace("\n", " ").strip()
        out.append("|".join([
            clean(j.get("code")), clean(j.get("name")), clean(mv.get("name")),
            dt(l["date"]), clean(a.get("code")), clean(a.get("name")),
            clean(aux_n), clean(aux_l),
            clean(mv.get("ref") or mv.get("name")),
            dt(mv.get("invoice_date") or mv.get("date")),
            clean(l["name"]) or "/",
            fr(l["debit"]), fr(l["credit"]),
            clean(l.get("matching_number") or ""), "",
            dt(mv.get("date")), "", "",
        ]))
    body = "﻿" + "\r\n".join(out)
    fn = f"FEC_PROTEC_{dt(d1)}_{dt(d2)}.txt"
    return Response(body, mimetype="text/plain; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})

@bp.route("/health")
def health():
    return {"status": "ok", "app": "protec-planning"}

# ─── Exécution seule (tests locaux) ──────────────────────────────────────────
def create_app():
    app = Flask(__name__)
    app.register_blueprint(bp)
    return app

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    create_app().run(host="0.0.0.0", port=port, debug=False)
