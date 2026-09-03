# -*- coding: utf-8 -*-
"""Export comptable Sage — multi-sociétés (même esprit que la page /protec/ebp).

Remplace les actions serveur Odoo 1459 (« Export journal pour logiciel compta
extérieur ») et 1671 (« Export journal Caisse pour Sage ») pour pouvoir
supprimer ce code payant d'Odoo. Formats de fichiers STRICTEMENT identiques
aux exports historiques validés par le cabinet (CSV `;`, cp1252) :
- ventes / banque : écritures groupées par (compte, analytique) — logique 1459 ;
- caisse : ligne à ligne (tickets comptoir) — logique 1671.

Page /export-compta : société + mois → aperçu par journal (pièces, lignes,
total débit) puis téléchargement par journal ou ZIP complet. Le ZIP contient
aussi `nouveaux_clients_*.txt` : les clients des écritures exportées jamais
transmis à Sage (`x_sage_envoye_le` vide), à importer dans le dossier — avec
option pour les marquer transmis.

Jeton : ir.config_parameter `maquignon.compta_key` (?token=...).
"""
import calendar
import io
import os
import zipfile
from datetime import date

import xmlrpc.client
from flask import Blueprint, request, abort, Response

bp = Blueprint("export_compta", __name__)

ODOO_URL = os.environ.get("ODOO_URL", "")
ODOO_DB = os.environ.get("ODOO_DB", "")
ODOO_USER = os.environ.get("ODOO_USER", "")
ODOO_PASSWORD = os.environ.get("ODOO_PASSWORD", "")

_conn = {}


def _q(model, method, *params, **kw):
    if "uid" not in _conn:
        common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
        _conn["uid"] = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
        _conn["models"] = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)
    return _conn["models"].execute_kw(ODOO_DB, _conn["uid"], ODOO_PASSWORD,
                                      model, method, list(params), kw)


def _check_token():
    import hmac
    tok = request.args.get("token", "") or request.form.get("token", "")
    ref = _q("ir.config_parameter", "get_param", "maquignon.compta_key") or ""
    if not (tok and ref and hmac.compare_digest(tok, ref)):
        abort(403)


ENTETE = ("Numéro de pièce;Numéro facture;Code journal;Date facture;Code client;"
          "Référence client;Nom client;Code compte;Libellé compte;Date échéance;"
          "Débit;Crédit;URL Facture;Plan;Analytique;Type écriture")


def _mois_bornes(mois):
    y, mo = int(mois[:4]), int(mois[5:7])
    return (f"{y:04d}-{mo:02d}-01", f"{y:04d}-{mo:02d}-{calendar.monthrange(y, mo)[1]:02d}")


def _periode(args):
    """Bornes (du, au) : plage libre du/au (YYYY-MM-DD, prioritaire — rapprochements
    en cours de mois) sinon le mois complet. Renvoie aussi le libellé affiché."""
    import re
    du, au = (args.get("du") or "").strip(), (args.get("au") or "").strip()
    if (re.match(r"^\d{4}-\d{2}-\d{2}$", du) and re.match(r"^\d{4}-\d{2}-\d{2}$", au)
            and du <= au):
        return du, au, "du %s au %s" % (du, au)
    mois = args.get("mois") or ""
    du, au = _mois_bornes(mois)
    return du, au, mois


def _ddmmyy(iso):
    return (iso[8:10] + iso[5:7] + iso[2:4]) if iso else ""


def _piece(numero):
    chiffres = "".join(c for c in (numero or "") if c.isdigit())
    return chiffres[-7:] if chiffres else ""


def _referentiels(line_rows, comp):
    acc_ids = list({l["account_id"][0] for l in line_rows if l["account_id"]})
    part_ids = list({l["partner_id"][0] for l in line_rows if l["partner_id"]})
    ana_ids = set()
    for l in line_rows:
        for k in (l.get("analytic_distribution") or {}):
            for part in str(k).split(","):
                if part.strip().isdigit():
                    ana_ids.add(int(part))
    # account.account.code est company-dependent : sans le contexte de la société
    # exportée, les comptes des autres sociétés ressortent sans code et toutes
    # leurs lignes sont écartées (fichier vide pour Haims/Châtel).
    comptes = {a["id"]: a for a in _q("account.account", "read", acc_ids,
                                      fields=["code", "name"],
                                      context={"allowed_company_ids": [comp]})} if acc_ids else {}
    partenaires = {p["id"]: p for p in _q("res.partner", "read", part_ids,
                                          fields=["ref", "name"])} if part_ids else {}
    # sections Sage uniquement : le plan « Project » (id 1 — Demande de transport,
    # Interne, Service sur site) est de l'analytique opérationnelle des modules
    # projet/transport et ne doit jamais partir vers Sage.
    analytiques = {a["id"]: a["name"]
                   for a in _q("account.analytic.account", "read", list(ana_ids),
                               fields=["name", "plan_id"])
                   if a["plan_id"] and a["plan_id"][0] != PLAN_PROJECT} if ana_ids else {}
    return comptes, partenaires, analytiques


PLAN_PROJECT = 1


def _ana_libelle(line, analytiques):
    """Sections Sage de la ligne. Les clés composées (« 13,2 » = section Sage +
    compte du plan Project) sont décomposées pour garder la section — l'action
    historique les faisait échouer et exportait un analytique vide."""
    dist = line.get("analytic_distribution") or {}
    ids, vus = [], set()
    for k in dist:
        for part in str(k).split(","):
            part = part.strip()
            if part.isdigit() and int(part) not in vus:
                vus.add(int(part))
                ids.append(int(part))
    return " | ".join(n for n in (analytiques.get(i, "") for i in ids) if n)


def _sections_pct(line, analytiques):
    """{id section Sage: %} de la ligne (clés composées décomposées : chaque
    section d'une clé reçoit le % entier de la clé, comme dans Odoo)."""
    out = {}
    for k, pct in (line.get("analytic_distribution") or {}).items():
        for part in str(k).split(","):
            part = part.strip()
            if part.isdigit() and int(part) in analytiques:
                out[int(part)] = out.get(int(part), 0.0) + (pct or 0.0)
    return out


def _lignes_a(base, analytiques, sections, tot_debit, tot_credit):
    """Lignes analytiques « A » : une par section Sage, montants au prorata des %.
    Une seule section = montants du groupe tels quels (comportement historique) ;
    plusieurs sections = ventilation, la dernière absorbe l'écart d'arrondi.
    base(debit, credit, analytique) rend la ligne sans le type d'écriture final."""
    ids = sorted(sections)
    lignes, cum_d, cum_c = [], 0.0, 0.0
    for i, sid in enumerate(ids):
        sd, sc = sections[sid]
        if i == len(ids) - 1:
            sd, sc = tot_debit - cum_d, tot_credit - cum_c
        else:
            sd, sc = round(sd, 2), round(sc, 2)
            cum_d += sd
            cum_c += sc
        if sd > 0 and sc > 0:
            legs = [(sd, 0), (0, sc)]
        elif sd or sc:
            legs = [(sd, sc)]
        else:
            continue
        for d, c in legs:
            lignes.append(base(d, c, analytiques[sid]) + "A")
    return lignes


def _export_ventes(journal, du, au, base_url, clients_vus, comp):
    """Format action 1459 : écritures groupées par (compte, analytique)."""
    moves = _q("account.move", "search_read",
               [("journal_id", "=", journal["id"]), ("state", "=", "posted"),
                 ("date", ">=", du), ("date", "<=", au)],
               fields=["name", "date", "invoice_date_due"], limit=0)  # ordre par défaut d'Odoo, comme l'action 1459
    if not moves:
        return None
    mids = [m["id"] for m in moves]
    lines = _q("account.move.line", "search_read", [("move_id", "in", mids)],
               fields=["move_id", "account_id", "debit", "credit", "partner_id",
                       "ref", "analytic_distribution"], limit=0, order="id asc")
    comptes, partenaires, analytiques = _referentiels(lines, comp)
    par_move = {}
    for l in lines:
        par_move.setdefault(l["move_id"][0], []).append(l)

    out = [ENTETE]
    for m in moves:
        numero = m["name"] or ""
        piece = _piece(numero)
        d_fac, d_ech = _ddmmyy(m["date"]), _ddmmyy(m["invoice_date_due"])
        groupes = {}
        for l in par_move.get(m["id"], []):
            debit, credit = l["debit"] or 0, l["credit"] or 0
            if debit == 0 and credit == 0:
                continue
            compte = comptes.get(l["account_id"] and l["account_id"][0]) or {}
            code_compte = compte.get("code") or ""
            if not code_compte:
                continue
            ana = _ana_libelle(l, analytiques)
            cle = (code_compte, ana)
            if cle not in groupes:
                part = partenaires.get(l["partner_id"] and l["partner_id"][0]) or {}
                if code_compte.startswith("411"):
                    code_client = part.get("ref") or ""
                    url = "%s/report/pdf/account.report_invoice/%s" % (base_url, m["id"])
                    if l["partner_id"]:
                        clients_vus.add(l["partner_id"][0])
                else:
                    code_client, url = "", ""
                groupes[cle] = {"code_client": code_client, "ref_ligne": l["ref"] or "",
                                "nom": part.get("name") or "", "url": url,
                                "debit": 0, "credit": 0, "sections": {}}
            groupes[cle]["debit"] += debit
            groupes[cle]["credit"] += credit
            for sid, pct in _sections_pct(l, analytiques).items():
                s = groupes[cle]["sections"].setdefault(sid, [0.0, 0.0])
                s[0] += debit * pct / 100.0
                s[1] += credit * pct / 100.0
        for (code_compte, ana), g in groupes.items():
            libelle = "%s %s" % (numero, g["nom"])
            if g["debit"] > 0 and g["credit"] > 0:
                montants = [(g["debit"], 0), (0, g["credit"])]
            else:
                montants = [(g["debit"], g["credit"])]
            def base(d, c, a, g=g, code_compte=code_compte, libelle=libelle):
                return "%s;%s;%s;%s;%s;%s;%s;%s;%s;%s;%.2f;%.2f;%s;1;%s;" % (
                    piece, numero, journal["code"] or "", d_fac,
                    g["code_client"], g["ref_ligne"], g["nom"],
                    code_compte, libelle, d_ech, d, c, g["url"], a)
            for debit, credit in montants:
                out.append(base(debit, credit, ana) + "G")
            if ana and code_compte.startswith("7"):
                out.extend(_lignes_a(base, analytiques, g["sections"],
                                     round(g["debit"], 2), round(g["credit"], 2)))
    return "\n".join(out)


def _export_caisse(journal, du, au, clients_vus, comp):
    """Format action 1671 : ligne à ligne (tickets comptoir)."""
    lines = _q("account.move.line", "search_read",
               [("journal_id", "=", journal["id"]), ("parent_state", "=", "posted"),
                 ("date", ">=", du), ("date", "<=", au)],
               fields=["move_id", "date", "account_id", "debit", "credit", "partner_id",
                       "name", "analytic_distribution"], limit=0,
               order="date asc, move_id asc, id asc")
    if not lines:
        return None
    mids = list({l["move_id"][0] for l in lines})
    moves = {m["id"]: m for m in _q("account.move", "read", mids,
                                    fields=["name", "invoice_date_due", "partner_id"])}
    comptes, partenaires, analytiques = _referentiels(lines, comp)
    extra = {m["partner_id"][0] for m in moves.values() if m["partner_id"]} - set(partenaires)
    if extra:
        for p in _q("res.partner", "read", list(extra), fields=["ref", "name"]):
            partenaires[p["id"]] = p
    soeurs = {}
    for l in lines:
        soeurs.setdefault(l["move_id"][0], []).append(l)

    out = [ENTETE]
    for l in lines:
        m = moves[l["move_id"][0]]
        numero = m["name"] or ""
        piece = _piece(numero)
        d_fac, d_ech = _ddmmyy(l["date"]), _ddmmyy(m["invoice_date_due"])
        compte = comptes.get(l["account_id"] and l["account_id"][0]) or {}
        code_compte = compte.get("code") or ""
        ana = _ana_libelle(l, analytiques)
        part = partenaires.get(l["partner_id"] and l["partner_id"][0]) or {}
        if code_compte.startswith("411"):
            code_client = part.get("ref") or ""
            nom = (part.get("name") or "").replace(";", ",")
            if l["partner_id"]:
                clients_vus.add(l["partner_id"][0])
        else:
            code_client, nom = "", ""
        libelle = (l["name"] or "").replace(";", ",")
        if not libelle:
            repli = part or (partenaires.get(m["partner_id"] and m["partner_id"][0]) or {})
            if not repli:
                for sib in soeurs.get(l["move_id"][0], []):
                    if sib["partner_id"]:
                        repli = partenaires.get(sib["partner_id"][0]) or {}
                        break
            libelle = (repli.get("name") or "").replace(";", ",")
        debit, credit = l["debit"] or 0, l["credit"] or 0

        def base(d, c, a, code_client=code_client, nom=nom,
                 code_compte=code_compte, libelle=libelle, numero=numero,
                 piece=piece, d_fac=d_fac, d_ech=d_ech):
            return "%s;%s;%s;%s;%s;%s;%s;%s;%s;%s;%.2f;%.2f;;1;%s;" % (
                piece, numero, journal["code"] or "", d_fac, code_client, numero,
                nom, code_compte, libelle, d_ech, d, c, a)
        out.append(base(debit, credit, ana) + "G")
        if ana and code_compte.startswith("7"):
            sections = {sid: [debit * pct / 100.0, credit * pct / 100.0]
                        for sid, pct in _sections_pct(l, analytiques).items()}
            out.extend(_lignes_a(base, analytiques, sections,
                                 round(debit, 2), round(credit, 2)))
    return "\n".join(out)


def _nouveaux_clients(clients_vus, marquer):
    """Clients des écritures exportées jamais transmis à Sage (fiches mères)."""
    if not clients_vus:
        return None, 0
    ps = _q("res.partner", "read", list(clients_vus),
            fields=["ref", "name", "street", "street2", "zip", "city", "vat",
                    "company_registry", "email", "phone", "x_sage_envoye_le", "parent_id"])
    nouveaux = [p for p in ps if p["ref"] and not p["x_sage_envoye_le"] and not p["parent_id"]]
    if not nouveaux:
        return None, 0
    out = ["Code Client;Intitulé Client;Adresse;Code Postal;Ville;Siret;"
           "N° TVA intracommunautaire;Email;Téléphone"]
    for p in sorted(nouveaux, key=lambda x: x["ref"]):
        adresse = " ".join(x for x in (p["street"], p["street2"]) if x)
        out.append(";".join((v or "").replace(";", ",") for v in (
            p["ref"], p["name"], adresse, p["zip"], p["city"],
            p["company_registry"], p["vat"], p["email"], p["phone"])))
    if marquer:
        _q("res.partner", "write", [p["id"] for p in nouveaux],
           {"x_sage_envoye_le": date.today().isoformat()})
    return "\n".join(out), len(nouveaux)


def _journaux(comp):
    """Journaux à exporter : ventes/banque/caisse + journaux de session PoS
    (ex. « Cloture CAISSE » Châtel, type OD) — ces derniers au format caisse."""
    # pour l'instant le cabinet n'importe que les VENTES et la CAISSE
    js = _q("account.journal", "search_read",
            [("company_id", "=", comp), ("type", "in", ["sale", "cash"])],
            fields=["name", "code", "type"], order="type, id")
    pos = _q("pos.config", "search_read", [("company_id", "=", comp)],
             fields=["journal_id"])
    pos_ids = {p["journal_id"][0] for p in pos if p["journal_id"]}
    deja = {j["id"] for j in js}
    manquants = [i for i in pos_ids if i not in deja]
    if manquants:
        js += _q("account.journal", "read", manquants, fields=["name", "code", "type"])
    for j in js:
        j["format_caisse"] = j["type"] == "cash" or j["id"] in pos_ids
        if j["id"] in pos_ids:
            j["explication"] = "Ventes comptoir du point de vente (tickets, TVA, REP) — le fichier « caisse » historique"
        elif j["type"] == "cash":
            j["explication"] = "Mouvements d'espèces : encaissements en liquide, remises en banque"
        elif j["type"] == "sale":
            j["explication"] = "Factures et avoirs clients (hors comptoir)"
        else:
            j["explication"] = "Relevés bancaires"
    return js


def _apercu(comp, du, au, journaux):
    """Une requête : pièces / lignes / total débit par journal + nouveaux clients."""
    lines = _q("account.move.line", "search_read",
               [("company_id", "=", comp), ("parent_state", "=", "posted"),
                 ("date", ">=", du), ("date", "<=", au),
                 ("journal_id", "in", [j["id"] for j in journaux])],
               fields=["journal_id", "move_id", "debit", "account_id", "partner_id"],
               limit=0)
    agg, clients = {}, set()
    acc_ids = list({l["account_id"][0] for l in lines if l["account_id"]})
    codes = {}
    for i in range(0, len(acc_ids), 800):
        for a in _q("account.account", "read", acc_ids[i:i + 800], fields=["code"],
                    context={"allowed_company_ids": [comp]}):
            codes[a["id"]] = a["code"] or ""
    for l in lines:
        jid = l["journal_id"][0]
        a = agg.setdefault(jid, {"nlignes": 0, "debit": 0.0, "pieces": set()})
        a["nlignes"] += 1
        a["debit"] += l["debit"] or 0
        a["pieces"].add(l["move_id"][0])
        if l["partner_id"] and codes.get(l["account_id"] and l["account_id"][0], "").startswith("411"):
            clients.add(l["partner_id"][0])
    nouveaux = []
    if clients:
        ps = _q("res.partner", "read", list(clients),
                fields=["ref", "name", "x_sage_envoye_le", "parent_id"])
        nouveaux = sorted((p["ref"], p["name"]) for p in ps
                          if p["ref"] and not p["x_sage_envoye_le"] and not p["parent_id"])
    return agg, nouveaux


def _controle_analytique(comp, du, au, journaux):
    """Contrôle avant export (MAQUIGNON) : lignes de produits sur comptes 7*
    sans section analytique Sage (plans hors « Project »). Renvoie une liste
    [(pièce, compte, montant)] agrégée par pièce + compte."""
    if comp != 1 or not journaux:
        return []
    lines = _q("account.move.line", "search_read",
               [("company_id", "=", comp), ("parent_state", "=", "posted"),
                 ("date", ">=", du), ("date", "<=", au),
                 ("journal_id", "in", [j["id"] for j in journaux]),
                 ("display_type", "=", "product")],
               fields=["move_name", "account_id", "analytic_distribution",
                       "debit", "credit"], limit=0)
    acc_ids = list({l["account_id"][0] for l in lines if l["account_id"]})
    codes = {}
    for i in range(0, len(acc_ids), 800):
        for a in _q("account.account", "read", acc_ids[i:i + 800], fields=["code"],
                    context={"allowed_company_ids": [comp]}):
            codes[a["id"]] = a["code"] or ""
    sage_ids = {a["id"] for a in _q("account.analytic.account", "search_read",
                                    [("plan_id", "!=", PLAN_PROJECT)], fields=["id"], limit=0)}
    manq = {}
    for l in lines:
        code = codes.get(l["account_id"] and l["account_id"][0], "")
        if not code.startswith("7"):
            continue
        ids = {int(x) for k in (l.get("analytic_distribution") or {})
               for x in str(k).split(",") if x.strip().isdigit()}
        if ids & sage_ids:
            continue
        cle = (l["move_name"] or "?", code)
        manq[cle] = manq.get(cle, 0.0) + (l["credit"] or 0) - (l["debit"] or 0)
    return sorted((p, c, m) for (p, c), m in manq.items())


TYPES = {"sale": "Ventes", "bank": "Banque", "cash": "Caisse"}

PAGE = """<!doctype html><html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Export comptable Sage</title><style>
body{font-family:system-ui,sans-serif;background:#f1f5f9;margin:0;padding:24px;color:#0f172a;}
.card{background:#fff;border-radius:14px;box-shadow:0 1px 6px rgba(0,0,0,.09);max-width:760px;margin:0 auto;padding:24px;}
h1{font-size:20px;margin:0 0 4px;}p.sub{color:#64748b;font-size:13px;margin:0 0 16px;}
form.bar{display:flex;gap:10px;flex-wrap:wrap;align-items:end;margin-bottom:16px;}
label{display:block;font-weight:700;font-size:12.5px;color:#334155;margin-bottom:3px;}
select,input[type=date]{padding:8px;border:1.5px solid #cbd5e1;border-radius:9px;font-size:14px;}
button,a.btn{border:none;border-radius:9px;background:#0f172a;color:#fff;font-weight:800;font-size:13px;
padding:9px 14px;cursor:pointer;text-decoration:none;display:inline-block;}
button:hover,a.btn:hover{filter:brightness(1.2);}
a.btn.sec{background:#e2e8f0;color:#0f172a;}
table{border-collapse:collapse;width:100%;font-size:13.5px;}
th{background:#0f172a;color:#fff;padding:7px 9px;text-align:left;font-size:12px;}
td{border-bottom:1px solid #e2e8f0;padding:7px 9px;}
td.num{text-align:right;font-variant-numeric:tabular-nums;}
.nc{background:#fefce8;border:1.5px solid #fde68a;border-radius:10px;padding:10px 13px;margin:14px 0;font-size:13.5px;font-weight:600;color:#713f12;}
.chk{display:flex;gap:8px;align-items:center;font-size:13px;font-weight:600;color:#334155;margin:10px 0;}
</style></head><body><div class="card">
<h1>📤 Export comptable Sage</h1>
<p class="sub">Fichiers d'écritures au format d'import du cabinet (identiques aux exports
historiques) + fichier des nouveaux clients à créer dans Sage. Un dossier Sage par société.</p>
<p class="sub">🏦 <a href="rappro?token=__TOKEN__">Import du rapprochement bancaire Sage</a> — passe les factures pointées par Charlotte à « Payé ».</p>
<form class="bar" method="get" action="">
<input type="hidden" name="token" value="__TOKEN__">
<div><label>Société</label><select name="societe" onchange="this.form.submit()">__SOCIETES__</select></div>
<div><label>Mois complet</label><select name="mois" onchange="this.form.du.value='';this.form.au.value='';this.form.submit()">__MOIS__</select></div>
<div><label>ou du</label><input type="date" name="du" value="__DU__" onchange="if(this.form.au.value)this.form.submit()"></div>
<div><label>au</label><input type="date" name="au" value="__AU__" onchange="if(this.form.du.value)this.form.submit()"></div>
</form>
__CORPS__
</div></body></html>"""


@bp.route("/", methods=["GET"])
def page():
    _check_token()
    token = request.args.get("token", "")
    comps = _q("res.company", "search_read", [], fields=["name"], order="id")
    comp = int(request.args.get("societe") or comps[0]["id"])
    d = date.today()
    mois_prec = f"{(d.year * 12 + d.month - 2) // 12:04d}-{(d.year * 12 + d.month - 2) % 12 + 1:02d}"
    mois = request.args.get("mois") or mois_prec
    args = dict(request.args)
    args["mois"] = mois
    du, au, libelle = _periode(args)
    libre = libelle != mois
    periode_qs = ("du=%s&au=%s" % (du, au)) if libre else ("mois=%s" % mois)

    opts_soc = "".join('<option value="%s"%s>%s</option>' % (
        c["id"], " selected" if c["id"] == comp else "", c["name"]) for c in comps)
    opts_mois = ""
    for i in range(12):
        mm = d.year * 12 + d.month - 1 - i
        v = f"{mm // 12:04d}-{mm % 12 + 1:02d}"
        opts_mois += '<option value="%s"%s>%s</option>' % (v, " selected" if v == mois else "", v)

    journaux = _journaux(comp)
    agg, nouveaux = _apercu(comp, du, au, journaux)
    lignes_html = ""
    for j in journaux:
        a = agg.get(j["id"])
        if not a:
            continue
        url = "fichier?token=%s&journal=%s&%s" % (token, j["id"], periode_qs)
        lignes_html += ("<tr><td>%s</td><td>%s<div style='font-size:11px;color:#64748b;font-weight:400;'>%s</div></td>"
                        "<td>%s</td><td class='num'>%s</td>"
                        "<td class='num'>%s</td><td class='num'>%.2f €</td>"
                        "<td><a class='btn sec' href='%s'>⬇ .txt</a></td></tr>" % (
                            j["code"] or "", j["name"], j.get("explication", ""),
                            TYPES.get(j["type"], j["type"]),
                            len(a["pieces"]), a["nlignes"], a["debit"], url))
    if not lignes_html:
        corps = "<p><b>Aucune écriture validée sur %s pour cette société.</b></p>" % libelle
    else:
        corps = ("<table><tr><th>Code</th><th>Journal</th><th>Type</th><th>Pièces</th>"
                 "<th>Lignes</th><th>Total débit</th><th></th></tr>%s</table>" % lignes_html)
        detail_nc = ""
        if nouveaux:
            detail_nc = ("<div style='margin-top:6px;font-size:12px;'>%s</div>"
                         % " · ".join("<b>%s</b> %s" % (r, n) for r, n in nouveaux))
        corps += ("<div class='nc'>👤 <b>%s</b> client(s) des écritures de la période jamais "
                  "transmis à Sage — inclus dans le ZIP (nouveaux_clients_*.txt).%s</div>"
                  % (len(nouveaux), detail_nc))
        manq = _controle_analytique(comp, du, au, journaux)
        if manq:
            det = " · ".join("<b>%s</b> %s (%.2f €)" % (c, p, m) for p, c, m in manq[:20])
            if len(manq) > 20:
                det += " · … et %d autre(s)" % (len(manq) - 20)
            corps += ("<div class='nc' style='background:#fef2f2;border-color:#fecaca;color:#7f1d1d;'>"
                      "⚠️ <b>%d</b> ligne(s) de vente <b>sans section analytique</b> — elles partiront "
                      "dans Sage sans analytique. À compléter dans Odoo (ou à ignorer si c'est voulu, "
                      "ex. cession de matériel) puis recharger cette page."
                      "<div style='margin-top:6px;font-size:12px;font-weight:400;'>%s</div></div>"
                      % (len(manq), det))
        champs = ("<input type='hidden' name='du' value='%s'>"
                  "<input type='hidden' name='au' value='%s'>" % (du, au)) if libre \
            else "<input type='hidden' name='mois' value='%s'>" % mois
        corps += ("<form method='post' action='export?token=%s'>"
                  "<input type='hidden' name='societe' value='%s'>%s"
                  "<label class='chk'><input type='checkbox' name='marquer' value='1' checked> "
                  "Marquer les nouveaux clients comme transmis à Sage</label>"
                  "<button type='submit'>📦 Télécharger le ZIP complet (%s)</button>"
                  "</form>" % (token, comp, champs, libelle))
    html = (PAGE.replace("__SOCIETES__", opts_soc).replace("__MOIS__", opts_mois)
                .replace("__DU__", du if libre else "").replace("__AU__", au if libre else "")
                .replace("__TOKEN__", token).replace("__CORPS__", corps))
    return Response(html, mimetype="text/html")


def _fichier_journal(j, du, au, base_url, clients_vus, comp):
    if j.get("format_caisse") or j["type"] == "cash":
        return _export_caisse(j, du, au, clients_vus, comp), "export_tickets_comptoir"
    return _export_ventes(j, du, au, base_url, clients_vus, comp), "export_journal"


@bp.route("/fichier", methods=["GET"])
def fichier():
    _check_token()
    jid = int(request.args["journal"])
    du, au, _lib = _periode(request.args)
    j = _q("account.journal", "read", [jid], fields=["name", "code", "type", "company_id"])[0]
    pos = _q("pos.config", "search_read", [("journal_id", "=", jid)], fields=["id"])
    j["format_caisse"] = j["type"] == "cash" or bool(pos)
    base_url = _q("ir.config_parameter", "get_param", "web.base.url") or ""
    contenu, prefixe = _fichier_journal(j, du, au, base_url, set(), j["company_id"][0])
    if contenu is None:
        return Response("Aucune écriture.", mimetype="text/plain; charset=utf-8", status=404)
    slug = j["company_id"][1].replace(" ", "_").replace("/", "_")
    nom = "%s_%s_%s_du_%s_au_%s.txt" % (prefixe, slug, j["name"].replace(" ", "_"),
                                        du.replace("-", ""), au.replace("-", ""))
    return Response(contenu.encode("cp1252", errors="replace"),
                    mimetype="text/plain; charset=windows-1252",
                    headers={"Content-Disposition": "attachment; filename=%s" % nom})


@bp.route("/export", methods=["POST"])
def export():
    _check_token()
    comp = int(request.form["societe"])
    du, au, _lib = _periode(request.form)
    marquer = request.form.get("marquer") == "1"
    comp_nom = _q("res.company", "read", [comp], fields=["name"])[0]["name"]
    base_url = _q("ir.config_parameter", "get_param", "web.base.url") or ""
    slug = comp_nom.replace(" ", "_").replace("/", "_")
    suffixe = "du_%s_au_%s" % (du.replace("-", ""), au.replace("-", ""))
    clients_vus = set()
    buf = io.BytesIO()
    n_fichiers = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for j in _journaux(comp):
            contenu, prefixe = _fichier_journal(j, du, au, base_url, clients_vus, comp)
            if contenu is None:
                continue
            nom = "%s_%s_%s_%s.txt" % (prefixe, slug, j["name"].replace(" ", "_"), suffixe)
            z.writestr(nom, contenu.encode("cp1252", errors="replace"))
            n_fichiers += 1
        contenu_nc, n_nouveaux = _nouveaux_clients(clients_vus, marquer)
        if contenu_nc:
            z.writestr("nouveaux_clients_%s_%s.txt" % (slug, suffixe),
                       contenu_nc.encode("cp1252", errors="replace"))
    if n_fichiers == 0:
        return Response("Aucune écriture sur la période pour cette société.",
                        mimetype="text/plain; charset=utf-8", status=404)
    buf.seek(0)
    return Response(buf.read(), mimetype="application/zip",
                    headers={"Content-Disposition":
                             "attachment; filename=export_compta_%s_%s.zip" % (slug, suffixe)})


# ─── IMPORT DU RAPPROCHEMENT BANCAIRE SAGE (Maquignon & Haims) ───────────────
# Charlotte lettre dans Sage ; l'état « Rapprochement bancaire » imprimé vers
# Excel est déposé ici. Chaque encaissement pointé est rapproché des paiements
# « En paiement » d'Odoo (montant exact ou combinaison — remises de chèques),
# ou d'une facture ouverte (création du paiement). L'application crée
# l'écriture 512/511900 à la date de l'écriture pointée et lettre le compte
# d'attente → paiement « Payé », facture « Payée ». Décaissements ignorés
# (fournisseurs, hors périmètre). Sociétés : Maquignon, Châtel'Granulats,
# Haims. Distri Béton : rapprochement Odoo.
import json as _json
from itertools import combinations as _combi


def _rappro_parse(fichier):
    """Lit l'état Sage « Rapprochement bancaire » (impression vers Excel)."""
    import openpyxl
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb.active
    lignes, date_rappro, compte = [], "", ""
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if any(c and "Date de rapprochement" in str(c) for c in cells):
            for c in cells:
                if c and str(c)[:4].isdigit() and "-" in str(c):
                    date_rappro = str(c)[:10]
        if cells[0] and str(cells[0]).startswith("512"):
            compte = str(cells[0]).strip()
            debit = float(cells[13] or cells[14] or 0)
            credit = float(cells[16] or cells[17] or 0)
            lignes.append({"compte": compte, "date": str(cells[2])[:10],
                           "piece": str(cells[4] or ""), "lib": str(cells[6] or "").strip(),
                           "debit": round(debit, 2), "credit": round(credit, 2)})
    return lignes, date_rappro


def _norm_rappro(s):
    import unicodedata
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode().upper()
    return "".join(c for c in s if c.isalnum())


def _rappro_journal(comp, compte):
    """Journal Odoo dont le compte par défaut correspond au compte Sage."""
    js = _q("account.journal", "search_read",
            [("company_id", "=", comp), ("type", "=", "bank")],
            fields=["name", "code", "default_account_id"])
    cible = "".join(c for c in compte if c.isdigit())
    for j in js:
        code = "".join(c for c in (j["default_account_id"][1] if j["default_account_id"] else "") if c.isdigit())
        if code and (code.startswith(cible) or cible.startswith(code[:6])):
            return j
    return js[0] if js else None


def _rappro_analyse(comp, lignes, journal_id=None):
    """Propositions de lettrage pour les encaissements (débits 512)."""
    dom = [("company_id", "=", comp), ("state", "=", "in_process"),
           ("payment_type", "=", "inbound"), ("move_id", "!=", False)]
    if journal_id:
        dom.append(("journal_id", "=", journal_id))
    pays = _q("account.payment", "search_read", dom,
              fields=["name", "partner_id", "amount", "date", "move_id"], limit=0)
    invs = _q("account.move", "search_read",
              [("company_id", "=", comp), ("move_type", "=", "out_invoice"),
               ("state", "=", "posted"),
               ("payment_state", "in", ["not_paid", "partial"])],
              fields=["name", "partner_id", "amount_residual"], limit=0)
    props = []
    for l in lignes:
        if not l["debit"]:
            continue
        m, nlib = l["debit"], _norm_rappro(l["lib"])
        prop = {"ligne": l, "type": "inconnu", "detail": [], "ids": []}
        exacts = [p for p in pays if abs(p["amount"] - m) < 0.01]
        if len(exacts) == 1:
            p = exacts[0]
            prop.update(type="paiements", ids=[p["id"]],
                        detail=["%s — %s (%.2f €)" % (p["name"], p["partner_id"][1] if p["partner_id"] else "?", p["amount"])])
        elif len(exacts) > 1:
            nommes = [p for p in exacts if p["partner_id"] and (_norm_rappro(p["partner_id"][1]) in nlib or nlib in _norm_rappro(p["partner_id"][1]))]
            p = (nommes or exacts)[0]
            prop.update(type="paiements", ids=[p["id"]],
                        detail=["%s — %s (%.2f €)" % (p["name"], p["partner_id"][1] if p["partner_id"] else "?", p["amount"])])
        else:
            # combinaison : UNIQUEMENT pour les remises groupees (cheques, effets)
            import re as _re
            est_remise = bool(_re.search(r"remise|ch[eè]q|effet|lcr", l["lib"], _re.I))
            mts = [round(p["amount"], 2) for p in pays] if est_remise else []
            trouve = None
            for r in range(2, min(7, len(mts) + 1)):
                for c in _combi(range(len(mts)), r):
                    if abs(sum(mts[i] for i in c) - m) < 0.01:
                        trouve = c
                        break
                if trouve:
                    break
            if trouve:
                prop.update(type="paiements", ids=[pays[i]["id"] for i in trouve],
                            detail=["%s — %s (%.2f €)" % (pays[i]["name"], pays[i]["partner_id"][1] if pays[i]["partner_id"] else "?", pays[i]["amount"]) for i in trouve])
            else:
                cands = [i for i in invs if abs(i["amount_residual"] - m) < 0.01
                         and i["partner_id"] and (_norm_rappro(i["partner_id"][1]) in nlib or nlib in _norm_rappro(i["partner_id"][1]))]
                if len(cands) == 1:
                    i = cands[0]
                    prop.update(type="facture", ids=[i["id"]],
                                detail=["%s — %s (reste %.2f €)" % (i["name"], i["partner_id"][1], i["amount_residual"])])
        props.append(prop)
    ignores = [l for l in lignes if l["credit"]]
    return props, ignores


def _erreur_propre(s):
    """Traduit les erreurs techniques en messages lisibles par la comptable."""
    s = str(s)
    if "cannot marshal None" in s:
        return ("la réconciliation n'a pas abouti (écart de montant probable, "
                "quelques centimes ?) — à vérifier sur la pièce dans Odoo")
    if "Fault" in s or "Traceback" in s:
        lignes = [l.strip() for l in s.replace("\\n", "\n").split("\n") if l.strip()]
        for l in reversed(lignes):
            if "File \"" not in l and not l.startswith(("^", "~", "raise ")):
                return l.split(":", 1)[-1].strip().rstrip("'>\"") or "erreur Odoo"
    return s[-160:]


def _compte_ecart(comp, sens):
    """Compte d'écart de règlement : charges (658*) ou produits (758*)."""
    pref = "658" if sens == "charge" else "758"
    a = _q("account.account", "search",
           [("code", "=like", pref + "%"), ("company_ids", "in", [comp])],
           limit=1, context={"allowed_company_ids": [comp]})
    return a[0] if a else None


def _infos_journal(jid):
    """(compte d'attente, compte banque) d'un journal — None si non configuré."""
    j = _q("account.journal", "read", [jid], fields=["name", "default_account_id"])[0]
    pml = _q("account.payment.method.line", "search_read",
             [("journal_id", "=", jid), ("payment_type", "=", "inbound"),
              ("payment_account_id", "!=", False)],
             fields=["payment_account_id"], limit=1)
    if not pml:
        return None, None, j["name"]
    return pml[0]["payment_account_id"][0], j["default_account_id"][0], j["name"]


def _rappro_applique(comp, journal, props):
    """Crée l'écriture 512/511 et lettre — la primitive validée sur la base de test.
    Chaque proposition peut porter son propre journal (journal_id), sinon celui
    passé en paramètre (cas de l'état de rapprochement, un seul journal)."""
    jcache = {}

    def infos(jid):
        if jid not in jcache:
            jcache[jid] = _infos_journal(jid)
        return jcache[jid]

    faits, erreurs = 0, []
    for prop in props:
        l = prop["ligne"]
        jid = prop.get("journal_id") or journal["id"]
        compte_attente, compte_banque, jnom = infos(jid)
        if prop["type"] == "releve":
            # v2 : la banque connectée a déjà l'écriture (ligne de relevé) —
            # on crée le paiement manquant puis on réconcilie l'attente du/des
            # paiement(s) avec la contrepartie du relevé. Zéro écriture créée.
            try:
                pay_ids = list(prop.get("pay_ids") or [])
                if prop.get("ids"):
                    ctx = {"active_model": "account.move", "active_ids": prop["ids"]}
                    wid = _q("account.payment.register", "create",
                             [{"journal_id": jid, "payment_date": l["date"],
                               "group_payment": True, "communication": l["lib"][:60]}],
                             context=ctx)
                    wid = wid[0] if isinstance(wid, list) else wid
                    try:
                        _q("account.payment.register", "action_create_payments", [wid], context=ctx)
                    except Exception:
                        pass
                    nouveaux = _q("account.payment", "search",
                                  [("company_id", "=", comp), ("journal_id", "=", jid),
                                   ("state", "=", "in_process"),
                                   ("reconciled_invoice_ids", "in", prop["ids"])],
                                  limit=5, order="id desc")
                    pay_ids += nouveaux
                moves = [p["move_id"][0] for p in _q("account.payment", "read", pay_ids, fields=["move_id"]) if p["move_id"]]
                l_att_rows = _q("account.move.line", "search_read",
                                [("move_id", "in", moves), ("account_id", "=", compte_attente), ("reconciled", "=", False)],
                                fields=["debit"])
                l_att = [r["id"] for r in l_att_rows]
                l_stmt, tot_stmt = [], 0.0
                for sm in prop.get("stmt_moves") or []:
                    for ml in _q("account.move.line", "search_read",
                                 [("move_id", "=", sm), ("reconciled", "=", False)],
                                 fields=["account_id", "credit"]):
                        if ml["account_id"] and ml["account_id"][0] != compte_banque and ml["credit"] > 0:
                            if ml["account_id"][0] != compte_attente:
                                _q("account.move.line", "write", [ml["id"]], {"account_id": compte_attente})
                            l_stmt.append(ml["id"])
                            tot_stmt += ml["credit"]
                if not l_att or not l_stmt:
                    erreurs.append("%s : lignes à réconcilier introuvables" % l["lib"][:40])
                    continue
                ids = l_att + l_stmt
                # écart de quelques centimes entre facture(s) et virement réel
                # (lettrage toléré par Sage) : passé en écart de règlement
                ecart = round(sum(r["debit"] for r in l_att_rows) - tot_stmt, 2)
                if 0 < abs(ecart) <= 0.05:
                    cpt = _compte_ecart(comp, "charge" if ecart > 0 else "produit")
                    if cpt:
                        m_ec = _q("account.move", "create", [{
                            "journal_id": jid, "date": l["date"],
                            "ref": "Écart de règlement Sage — %s" % l["lib"][:50],
                            "line_ids": [
                                (0, 0, {"account_id": cpt if ecart > 0 else compte_attente,
                                        "debit": abs(ecart), "credit": 0.0,
                                        "name": "Écart de règlement (%.2f €)" % ecart}),
                                (0, 0, {"account_id": compte_attente if ecart > 0 else cpt,
                                        "debit": 0.0, "credit": abs(ecart),
                                        "name": "Écart de règlement (%.2f €)" % ecart}),
                            ]}], context={"allowed_company_ids": [comp]})
                        m_ec = m_ec[0] if isinstance(m_ec, list) else m_ec
                        _q("account.move", "action_post", [m_ec])
                        ids += _q("account.move.line", "search",
                                  [("move_id", "=", m_ec), ("account_id", "=", compte_attente)])
                try:
                    _q("account.move.line", "reconcile", ids)
                except Exception:
                    verif = _q("account.move.line", "read", ids, fields=["reconciled"])
                    if not all(v["reconciled"] for v in verif):
                        raise
                faits += 1
            except Exception as exc:
                erreurs.append("%s : %s" % (l["lib"][:40], _erreur_propre(exc)))
            continue
        if not compte_attente:
            erreurs.append("journal %s : compte d'attente non configuré" % jnom)
            continue
        try:
            pay_ids = list(prop["ids"])
            if prop["type"] == "facture":
                ctx = {"active_model": "account.move", "active_ids": prop["ids"]}
                wid = _q("account.payment.register", "create",
                         [{"journal_id": jid, "amount": l["debit"], "payment_date": l["date"]}],
                         context=ctx)
                wid = wid[0] if isinstance(wid, list) else wid
                try:
                    _q("account.payment.register", "action_create_payments", [wid], context=ctx)
                except Exception:
                    pass  # l'action renvoyée peut contenir des None non sérialisables
                pay_ids = _q("account.payment", "search",
                             [("company_id", "=", comp), ("journal_id", "=", jid),
                              ("amount", "=", l["debit"]), ("date", "=", l["date"])],
                             limit=1, order="id desc")
                if not pay_ids:
                    erreurs.append("%s %.2f € : le paiement n'a pas pu être créé" % (l["lib"][:30], l["debit"]))
                    continue
            # lignes d'attente des paiements
            moves = [p["move_id"][0] for p in _q("account.payment", "read", pay_ids, fields=["move_id"]) if p["move_id"]]
            l_att = _q("account.move.line", "search",
                       [("move_id", "in", moves), ("account_id", "=", compte_attente), ("reconciled", "=", False)])
            if not l_att:
                erreurs.append("%s %.2f € : lignes d'attente introuvables" % (l["lib"][:30], l["debit"]))
                continue
            mid = _q("account.move", "create", [{
                "journal_id": jid, "date": l["date"],
                "ref": "Rapprochement Sage — %s" % (l["lib"][:60] or l["piece"]),
                "line_ids": [
                    (0, 0, {"account_id": compte_banque, "debit": l["debit"], "credit": 0.0,
                            "name": "Rapprochement Sage %s" % l["piece"]}),
                    (0, 0, {"account_id": compte_attente, "debit": 0.0, "credit": l["debit"],
                            "name": l["lib"][:60] or "Rapprochement Sage"}),
                ]}])
            mid = mid[0] if isinstance(mid, list) else mid
            _q("account.move", "action_post", [mid])
            l_rap = _q("account.move.line", "search", [("move_id", "=", mid), ("account_id", "=", compte_attente)])
            try:
                _q("account.move.line", "reconcile", l_att + l_rap)
            except Exception:
                # le contrôleur XML-RPC d'Odoo SaaS ne sait pas sérialiser le
                # None renvoyé par reconcile() : le lettrage a bien eu lieu
                # côté serveur — on vérifie l'état réel des lignes.
                verif = _q("account.move.line", "read", l_att + l_rap, fields=["reconciled"])
                if not all(v["reconciled"] for v in verif):
                    raise
            faits += 1
        except Exception as exc:
            erreurs.append("%s %.2f € : %s" % (l["lib"][:30], l["debit"], _erreur_propre(exc)))
    return faits, erreurs


# ─── GRAND-LIVRE DES TIERS (Sage) : lettrage par lettres ─────────────────────
# Fichier bien plus riche que l'état de rapprochement : chaque groupe portant
# la même lettre relie la ou les factures (n° Odoo dans le libellé des ventes)
# à leur(s) règlement(s) datés, journal par journal. À-nouveaux (RAN) inclus.
def _gl_detecte(fichier):
    """True si le fichier est un « Grand-livre des tiers » (sinon état 512)."""
    import openpyxl
    wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 8:
            break
        if any(c and "Grand-livre des tiers" in str(c) for c in row):
            return True
    return False


def _gl_parse(fichier):
    """Par client : écritures {date, journal, piece, lib, lettre, debit, credit}."""
    import openpyxl
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb.active
    clients, cur = [], None
    for row in ws.iter_rows(values_only=True):
        c = list(row) + [None] * (19 - len(row))
        c0 = c[0]
        if c[5] == "Total du tiers" or (c0 and "Sage" in str(c0)) or c0 == "Date":
            continue
        est_date = hasattr(c0, "year")
        if c0 and not est_date and not c[1] and c[3]:
            cur = {"code": str(c0).strip(), "nom": str(c[3]).strip(), "ecritures": []}
            clients.append(cur)
            continue
        if cur is None or not est_date:
            continue
        cur["ecritures"].append({
            "date": str(c0)[:10], "journal": str(c[1] or "").strip(),
            "piece": str(c[2] or "").strip(), "lib": str(c[5] or "").strip(),
            "lettre": str(c[8] or "").strip(),
            "debit": round(float(c[12] or 0), 2), "credit": round(float(c[15] or 0), 2)})
    return [cl for cl in clients if cl["ecritures"]]


def _gl_analyse(comp, clients):
    """Groupes (client, lettre) équilibrés → propositions par facture Odoo."""
    import re as _re
    from collections import defaultdict
    js = _q("account.journal", "search_read",
            [("company_id", "=", comp), ("type", "=", "bank")], fields=["name", "code"])
    # journaux réellement câblés sur le compte d'attente (les journaux « CB /
    # chèques à encaisser » de Châtel restent sur leur circuit Odoo propre)
    cables = set()
    for j in js:
        pml = _q("account.payment.method.line", "search",
                 [("journal_id", "=", j["id"]), ("payment_type", "=", "inbound"),
                  ("payment_account_id", "!=", False)], limit=1)
        if pml:
            cables.add(j["id"])

    def journal_pour(code):
        code = (code or "").upper()
        for j in js:
            if code and (code == (j["code"] or "").upper() or code in (j["name"] or "").upper()):
                return j["id"]
        return js[0]["id"] if js else 0

    def journal_nom(jid2):
        for j in js:
            if j["id"] == jid2:
                return j["name"]
        return "?"

    noms = set()
    for cl in clients:
        for e in cl["ecritures"]:
            m = _re.search(r"(FAC/[0-9]{4}/[0-9]+|FAC/[0-9]{2}-[0-9]{2}/[0-9]+)", e["lib"])
            e["fac"] = m.group(1) if m else ""
            if e["fac"]:
                noms.add(e["fac"])
    invs = {}
    noms = sorted(noms)
    for i in range(0, len(noms), 400):
        for r in _q("account.move", "search_read",
                    [("company_id", "=", comp), ("name", "in", noms[i:i + 400])],
                    fields=["name", "payment_state", "amount_residual", "state"]):
            invs[r["name"]] = r
    # paiements « en paiement » existants : facture -> paiement à lettrer
    pay_par_inv = {}
    for p in _q("account.payment", "search_read",
                [("company_id", "=", comp), ("state", "=", "in_process"),
                 ("payment_type", "=", "inbound"), ("move_id", "!=", False)],
                fields=["name", "amount", "reconciled_invoice_ids"], limit=0):
        for iid in (p.get("reconciled_invoice_ids") or []):
            pay_par_inv.setdefault(iid, p)
    # v2 : lignes de relevé non rapprochées (banques connectées) — la vraie
    # écriture de banque est déjà dans Odoo, on la réconcilie au lieu d'en créer
    stmt = _q("account.bank.statement.line", "search_read",
              [("company_id", "=", comp), ("is_reconciled", "=", False), ("amount", ">", 0)],
              fields=["date", "amount", "journal_id", "move_id", "payment_ref"], limit=0)
    stmt_uses = set()

    def stmt_pour(jid, montant, dstr):
        import datetime as _dt
        try:
            dref = _dt.date.fromisoformat(dstr[:10])
        except Exception:
            return None
        cands = []
        for s2 in stmt:
            if s2["id"] in stmt_uses or s2["journal_id"][0] != jid:
                continue
            if abs(s2["amount"] - montant) > 0.005:
                continue
            ecart = abs((_dt.date.fromisoformat(str(s2["date"])[:10]) - dref).days)
            if ecart <= 5:
                cands.append((ecart, s2))
        if not cands:
            return None
        return sorted(cands, key=lambda c: c[0])[0][1]

    props, anomalies, deja = [], [], 0
    for cl in clients:
        groupes = defaultdict(list)
        for e in cl["ecritures"]:
            if e["lettre"]:
                groupes[e["lettre"]].append(e)
        for lettre, es in sorted(groupes.items()):
            sd = round(sum(e["debit"] for e in es), 2)
            sc = round(sum(e["credit"] for e in es), 2)
            facs = [e for e in es if e["fac"] and e["debit"] > 0]
            regs = [e for e in es if e["credit"] > 0]
            if not facs or not regs:
                continue
            if abs(sd - sc) > 0.01:
                anomalies.append("%s lettre %s : débits %.2f ≠ crédits %.2f (lettrage partiel ?)"
                                 % (cl["code"], lettre, sd, sc))
                continue
            date_reg = max(e["date"] for e in regs)
            jids_regs = {journal_pour(e["journal"]) for e in regs}
            if any(j2 not in cables for j2 in jids_regs):
                hors = ", ".join(sorted({journal_nom(j2) for j2 in jids_regs if j2 not in cables}))
                anomalies.append("%s lettre %s : règlement via « %s » — suivi par le circuit CB/chèques d'Odoo, non traité ici"
                                 % (cl["code"], lettre, hors))
                continue
            jid = journal_pour(regs[-1]["journal"])
            # ── v2 : le groupe entier se réconcilie-t-il avec des lignes de relevé ? ──
            etats = [invs.get(e["fac"]) for e in facs]
            if all(i and i["state"] == "posted" for i in etats) \
               and not any(i["payment_state"] in ("paid", "reversed") for i in etats):
                matches = []
                for e in regs:
                    s2 = stmt_pour(journal_pour(e["journal"]), e["credit"], e["date"])
                    if s2 is None:
                        matches = None
                        break
                    matches.append(s2)
                ouverts_ok = all(
                    i["payment_state"] != "not_paid" or abs(i["amount_residual"] - e2["debit"]) < 0.01
                    for i, e2 in zip(etats, facs))
                a_payer = [i["id"] for i in etats if i["payment_state"] in ("not_paid", "partial")]
                pay_ids = sorted({pay_par_inv[i["id"]]["id"] for i in etats
                                  if i["payment_state"] == "in_payment" and i["id"] in pay_par_inv})
                complet = all(i["payment_state"] != "in_payment" or i["id"] in pay_par_inv
                              for i in etats) and (a_payer or pay_ids)
                if matches and ouverts_ok and complet:
                    for s2 in matches:
                        stmt_uses.add(s2["id"])
                    total = round(sum(e2["debit"] for e2 in facs), 2)
                    libs = ", ".join(e2["fac"] for e2 in facs)
                    props.append({
                        "ligne": {"date": date_reg, "piece": lettre, "credit": 0.0,
                                  "lib": "%s — %s (lettre %s)" % (libs[:40], cl["nom"][:24], lettre),
                                  "debit": total},
                        "type": "releve", "ids": a_payer, "pay_ids": pay_ids,
                        "stmt_moves": [s2["move_id"][0] for s2 in matches],
                        "journal_id": jid,
                        "detail": ["relevé %s : %s (%.2f €)" % (str(s2["date"])[:10],
                                   (s2["payment_ref"] or "")[:45], s2["amount"]) for s2 in matches]})
                    continue
            for e in facs:
                inv = invs.get(e["fac"])
                if not inv or inv["state"] != "posted":
                    anomalies.append("%s : facture %s introuvable dans Odoo" % (cl["code"], e["fac"]))
                    continue
                if inv["payment_state"] in ("paid", "reversed"):
                    deja += 1
                    continue
                lib = "%s — %s (lettre %s)" % (e["fac"], cl["nom"][:30], lettre)
                pay = pay_par_inv.get(inv["id"])
                if inv["payment_state"] == "in_payment" and pay:
                    props.append({"ligne": {"date": date_reg, "lib": lib, "piece": lettre,
                                            "debit": round(pay["amount"], 2), "credit": 0.0},
                                  "type": "paiements", "ids": [pay["id"]], "journal_id": jid,
                                  "detail": ["paiement %s (%.2f €) réglé le %s"
                                             % (pay["name"], pay["amount"], date_reg)]})
                elif inv["payment_state"] in ("not_paid", "partial"):
                    montant = min(e["debit"], round(inv["amount_residual"], 2))
                    if montant <= 0:
                        deja += 1
                        continue
                    props.append({"ligne": {"date": date_reg, "lib": lib, "piece": lettre,
                                            "debit": montant, "credit": 0.0},
                                  "type": "facture", "ids": [inv["id"]], "journal_id": jid,
                                  "detail": ["aucun paiement saisi dans Odoo — cocher pour créer le paiement de %.2f € au %s et lettrer"
                                             % (montant, date_reg)]})
                else:
                    deja += 1
    return props, anomalies, deja


RAPPRO_PAGE = """<!doctype html><html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Import rapprochement Sage</title><style>
body{font-family:system-ui,sans-serif;background:#f1f5f9;margin:0;padding:24px;color:#0f172a;}
.card{background:#fff;border-radius:14px;box-shadow:0 1px 6px rgba(0,0,0,.09);max-width:900px;margin:0 auto;padding:24px;}
h1{font-size:20px;margin:0 0 4px;}p.sub{color:#64748b;font-size:13px;margin:0 0 16px;}
label{display:block;font-weight:700;font-size:12.5px;color:#334155;margin:10px 0 4px;}
select,input[type=file]{padding:8px;border:1.5px solid #cbd5e1;border-radius:9px;font-size:14px;}
button{border:none;border-radius:9px;background:#0f172a;color:#fff;font-weight:800;font-size:13px;padding:10px 16px;cursor:pointer;margin-top:14px;}
button:hover{filter:brightness(1.2);}
table{border-collapse:collapse;width:100%;font-size:13px;margin-top:12px;}
th{background:#0f172a;color:#fff;padding:6px 8px;text-align:left;font-size:11.5px;}
td{border-bottom:1px solid #e2e8f0;padding:6px 8px;vertical-align:top;}
.ok{color:#166534;font-weight:700;}.crt{color:#1d4ed8;font-weight:700;}.ko{color:#b91c1c;font-weight:700;}
.note{background:#fefce8;border:1.5px solid #fde68a;border-radius:10px;padding:9px 12px;font-size:12.5px;margin:12px 0;}
a.retour{font-size:12.5px;}
#att{display:none;position:fixed;inset:0;background:rgba(241,245,249,.85);z-index:50;
     align-items:center;justify-content:center;flex-direction:column;gap:14px;}
#att .rond{width:46px;height:46px;border:5px solid #cbd5e1;border-top-color:#0f172a;
     border-radius:50%;animation:tourne 0.9s linear infinite;}
#att .txt{font-weight:800;color:#0f172a;font-size:15px;}
@keyframes tourne{to{transform:rotate(360deg);}}
</style></head><body>
<div id="att"><div class="rond"></div><div class="txt" id="attTxt">Analyse du fichier…</div></div>
<script>
document.addEventListener('submit', function(ev){
  var f = ev.target;
  var b = f.querySelector('button[type=submit]');
  var texte = (f.action || '').indexOf('applique') >= 0 ? 'Application du lettrage…' : 'Analyse du fichier…';
  document.getElementById('attTxt').textContent = texte;
  document.getElementById('att').style.display = 'flex';
  if(b){ b.disabled = true; }
}, true);
</script>
<div class="card">
<h1>🏦 Import du rapprochement bancaire Sage</h1>
<p class="sub">Déposez l'état « Rapprochement bancaire » de Sage (imprimé vers Excel). Les
encaissements pointés passent les paiements et factures Odoo à « Payé », à la date du pointage.
Sociétés concernées : SARL Maquignon, Châtel'Granulats et Carrière d'Haims (Distri Béton
reste rapprochée dans Odoo).</p>
__CORPS__
<p style="margin-top:16px;"><a class="retour" href="./?token=__TOKEN__">← Retour à l'export Sage</a></p>
</div></body></html>"""


@bp.route("/rappro", methods=["GET"])
def rappro_page():
    _check_token()
    token = request.args.get("token", "")
    corps = ("""<form method="post" action="rappro/analyse?token=%s" enctype="multipart/form-data">
<label>Société</label><select name="societe">
<option value="1">SARL MAQUIGNON</option><option value="3">CHATEL'GRANULATS</option><option value="4">CARRIERE D'HAIMS</option></select>
<label>Fichier Sage (.xlsx) — état de rapprochement bancaire OU grand-livre des tiers avec lettrage (le type est détecté automatiquement)</label><input type="file" name="fichier" accept=".xlsx" required/>
<br/><button type="submit">🔎 Analyser</button></form>""" % token)
    return Response(RAPPRO_PAGE.replace("__CORPS__", corps).replace("__TOKEN__", token),
                    mimetype="text/html")


@bp.route("/rappro/analyse", methods=["POST"])
def rappro_analyse():
    _check_token()
    token = request.args.get("token", "")
    comp = int(request.form["societe"])
    buf = io.BytesIO(request.files["fichier"].read())
    if _gl_detecte(buf):
        # ── Grand-livre des tiers : lettrage par lettres ──
        buf.seek(0)
        clients = _gl_parse(buf)
        props, anomalies, deja = _gl_analyse(comp, clients)
        rows = ""
        for i, pr in enumerate(props):
            l = pr["ligne"]
            if pr["type"] == "releve":
                statut = "<span class='ok'>✓ rapprocher avec le relevé bancaire</span>"
                coche = "checked"
            elif pr["type"] == "paiements":
                statut = "<span class='ok'>✓ lettrer le paiement existant</span>"
                coche = "checked"
            else:
                # lettré dans Sage mais AUCUN paiement saisi dans Odoo : oubli
                # de saisie probable -> alerte, décochée par défaut
                statut = "<span style='color:#b45309;font-weight:700;'>⚠ paiement absent d'Odoo (oubli de saisie ?)</span>"
                coche = ""
            rows += ("<tr><td><input type='checkbox' name='sel' value='%d' %s/></td>"
                     "<td>%s</td><td>%s</td><td style='text-align:right;'>%.2f €</td>"
                     "<td>%s<div style='font-size:11.5px;color:#64748b;'>%s</div></td></tr>"
                     % (i, coche, l["date"], l["lib"][:60], l["debit"], statut, "<br/>".join(pr["detail"])))
        note_ano = ""
        if anomalies:
            note_ano = ("<div class='note'>⚠ Hors lettrage automatique :<br/>%s</div>"
                        % "<br/>".join(anomalies[:30]))
        js = _q("account.journal", "search", [("company_id", "=", comp), ("type", "=", "bank")], limit=1)
        if not props:
            corps = ("<div class='note'>Grand-livre lu : <b>%d</b> clients · <b>%d</b> facture(s) déjà à jour "
                     "dans Odoo · rien de nouveau à lettrer.</div>%s" % (len(clients), deja, note_ano))
        else:
            n_alertes = sum(1 for p in props if p["type"] == "facture")
            note_al = (" · <b style='color:#b45309;'>%d alerte(s) : lettré dans Sage sans paiement saisi dans Odoo</b>" % n_alertes) if n_alertes else ""
            corps = ("""<div class="note">Grand-livre des tiers lu : <b>%d</b> clients ·
<b>%d</b> facture(s) à passer « Payé » · %d déjà à jour dans Odoo%s.</div>%s
<form method="post" action="applique?token=%s">
<input type="hidden" name="societe" value="%d"/>
<input type="hidden" name="journal" value="%d"/>
<input type="hidden" name="props" value='%s'/>
<input type="hidden" name="stats" value='%s'/>
<table><tr><th></th><th>Réglée le</th><th>Facture — client</th><th>Montant</th><th>Action</th></tr>%s</table>
<button type="submit">✅ Appliquer le lettrage (%d)</button></form>"""
                     % (len(clients), len(props), deja, note_al, note_ano, token, comp,
                        js[0] if js else 0, _json.dumps(props).replace("'", "&#39;"),
                        _json.dumps({"deja": deja, "ano": len(anomalies)}),
                        rows, len(props)))
        return Response(RAPPRO_PAGE.replace("__CORPS__", corps).replace("__TOKEN__", token),
                        mimetype="text/html")
    buf.seek(0)
    lignes, date_rappro = _rappro_parse(buf)
    if not lignes:
        return Response(RAPPRO_PAGE.replace("__CORPS__", "<p class='ko'>Aucune écriture 512 trouvée dans ce fichier — est-ce bien l'état « Rapprochement bancaire » imprimé vers Excel ?</p>").replace("__TOKEN__", token), mimetype="text/html")
    journal = _rappro_journal(comp, lignes[0]["compte"])
    props, ignores = _rappro_analyse(comp, lignes, journal_id=journal["id"] if journal else None)
    rows = ""
    for i, pr in enumerate(props):
        l = pr["ligne"]
        if pr["type"] == "paiements":
            statut = "<span class='ok'>✓ %s paiement(s) reconnu(s)</span>" % len(pr["ids"])
        elif pr["type"] == "facture":
            statut = "<span class='crt'>➕ créer le paiement sur la facture</span>"
        else:
            statut = "<span class='ko'>? non reconnu — à traiter dans Odoo</span>"
        det = "<br/>".join(pr["detail"]) or "—"
        coche = "checked" if pr["type"] != "inconnu" else "disabled"
        rows += ("<tr><td><input type='checkbox' name='sel' value='%d' %s/></td>"
                 "<td>%s</td><td>%s</td><td style='text-align:right;'>%.2f €</td>"
                 "<td>%s<div style='font-size:11.5px;color:#64748b;'>%s</div></td></tr>"
                 % (i, coche, l["date"], l["lib"][:48], l["debit"], statut, det))
    n_auto = sum(1 for p in props if p["type"] != "inconnu")
    corps = ("""<div class="note">Journal identifié : <b>%s</b> · rapprochement du <b>%s</b> ·
%d encaissement(s) — <b>%d</b> lettrable(s) automatiquement · %d décaissement(s) ignoré(s) (fournisseurs).</div>
<form method="post" action="applique?token=%s">
<input type="hidden" name="societe" value="%d"/>
<input type="hidden" name="journal" value="%d"/>
<input type="hidden" name="props" value='%s'/>
<input type="hidden" name="stats" value='%s'/>
<table><tr><th></th><th>Date</th><th>Libellé Sage</th><th>Montant</th><th>Proposition</th></tr>%s</table>
<button type="submit">✅ Appliquer le lettrage (%d)</button></form>"""
             % (journal["name"] if journal else "?", date_rappro or "?", len(props), n_auto,
                len(ignores), token, comp, journal["id"] if journal else 0,
                _json.dumps(props).replace("'", "&#39;"),
                _json.dumps({"inconnu": len(props) - n_auto, "ignores": len(ignores)}),
                rows, n_auto))
    return Response(RAPPRO_PAGE.replace("__CORPS__", corps).replace("__TOKEN__", token),
                    mimetype="text/html")


@bp.route("/rappro/applique", methods=["POST"])
def rappro_applique():
    _check_token()
    token = request.args.get("token", "")
    comp = int(request.form["societe"])
    props = _json.loads(request.form["props"])
    sel = {int(i) for i in request.form.getlist("sel")}
    retenus = [p for i, p in enumerate(props) if i in sel and p["type"] != "inconnu"]
    journal = _q("account.journal", "read", [int(request.form["journal"])],
                 fields=["name", "code", "default_account_id"])[0]
    faits, erreurs = _rappro_applique(comp, journal, retenus)
    try:
        stats = _json.loads(request.form.get("stats") or "{}")
    except Exception:
        stats = {}
    n_lettres = sum(1 for p in retenus if p["type"] == "paiements")
    n_releves = sum(1 for p in retenus if p["type"] == "releve")
    n_crees = sum(1 for p in retenus if p["type"] == "facture")
    ecartes = len(props) - len(retenus)
    lignes_recap = ["<b>%d</b> appliqué(s) — factures passées « Payé »" % faits]
    if n_lettres:
        lignes_recap.append("dont %d lettrage(s) de paiements déjà saisis" % n_lettres)
    if n_releves:
        lignes_recap.append("dont %d groupe(s) rapprochés avec les relevés bancaires" % n_releves)
    if n_crees:
        lignes_recap.append("<span style='color:#b45309;'>⚠ dont %d paiement(s) créé(s) faute de saisie dans Odoo (oubli à vérifier)</span>" % n_crees)
    if stats.get("deja"):
        lignes_recap.append("%d facture(s) déjà à jour dans Odoo (rien à faire)" % stats["deja"])
    if ecartes:
        lignes_recap.append("%d proposition(s) laissée(s) de côté (non cochées ou non reconnues)" % ecartes)
    if stats.get("inconnu"):
        lignes_recap.append("%d ligne(s) non reconnue(s) — à traiter dans Odoo" % stats["inconnu"])
    if stats.get("ignores"):
        lignes_recap.append("%d décaissement(s) ignoré(s) (fournisseurs)" % stats["ignores"])
    if stats.get("ano"):
        lignes_recap.append("%d groupe(s) hors lettrage automatique (à vérifier dans Sage)" % stats["ano"])
    if erreurs:
        lignes_recap.append("<span class='ko'>%d erreur(s) — détail ci-dessous</span>" % len(erreurs))
    corps = ("<p class='ok'>✅ Lettrage appliqué.</p><div class='note'>📋 Récapitulatif :<br/>• "
             + "<br/>• ".join(lignes_recap) + "</div>")
    if erreurs:
        corps += "<div class='note'>⚠ À traiter manuellement :<br/>%s</div>" % "<br/>".join(erreurs)
    return Response(RAPPRO_PAGE.replace("__CORPS__", corps).replace("__TOKEN__", token),
                    mimetype="text/html")
